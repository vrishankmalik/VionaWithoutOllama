"""Two-tab enriched workbook builder.

Sheet 1 — "DPD + NOC + Patents":
  One row per DIN (DPD and NOC records joined), sorted ascending by DIN.
  NOC rows whose DIN is blank / "Not Applicable" are excluded entirely.
  Patent block: WIDE format — patent_1_number/filing/grant/expiry through
  patent_M_number/..., where M = max patents held by any DIN in the dataset.
  Labeling columns per DIN (from store). Data protection columns from the
  Register of Innovative Drugs (dp_6yr_no_file_date, pediatric_extension,
  data_protection_ends).

  Columns removed vs old format (Change 2):
    - All *_url columns (record_url, noc_record_url, labeling_pdf_url)
    - All *_page citation columns
    - Old combined patent_numbers / all_patents_detail / earliest_* / latest_* columns
  Columns kept:
    - _drug_code (internal DPD identifier, not a URL)
    - needs_ocr (live OCR-pipeline provenance flag)

Sheet 2 — "Generic Submissions":
  GSUR records filtered to the queried ingredient (same normalisation used
  elsewhere). Standalone — never joined to Sheet 1.

CLI:
  python -m app.enrichment.workbook --q "alpelisib" --field ingredient
"""
from __future__ import annotations

import io
import logging
import re
from typing import Any, Optional

import pandas as pd

from app.config import WORKBOOK_MIN_FILL_RATE

logger = logging.getLogger(__name__)

from app.enrichment.data_protection import (
    _match_data_protection_deterministic,
)
from app.enrichment.store import get_labeling_for_din, get_patents_for_din
from app.models import DrugRecord, SearchResponse

# DIN values that should be excluded from Sheet 1
_EXCLUDED_DIN_VALUES = {"", "not applicable", "n/a", "na", "none"}

# Supplement submission types to drop from Sheet 1 (SNDS / SANDS)
_SUPPLEMENT_TYPE_RE = re.compile(
    r"\bSNDS\b|\bSANDS\b|Supplement\s+to\s+(a\s+New|an\s+Abbreviated)",
    re.IGNORECASE,
)

# Sentinel dict for DPD DINs that have no matching NOC record
_NO_NOC_RECORD = {
    "noc_brand_name": "No NOC record",
    "noc_company": "No NOC record",
    "noc_date": "No NOC record",
    "noc_submission_type": "No NOC record",
    "noc_therapeutic_class": "No NOC record",
}

_LABELING_FIELDS = (
    "active_ingredient", "excipients_core", "excipients_coating",
    "preservatives", "pack_size", "pack_style",
    "colour", "shape", "size_mm", "weight", "ph",
)

# Columns that are NEVER pruned regardless of fill rate.
_NEVER_DROP_COLS = frozenset({
    # Identity / provenance
    "din", "_drug_code", "needs_ocr",
    # DPD core
    "brand_name", "company", "ingredient", "strength",
    "dosage_form", "route", "status",
    # Patent summary (patent_1_* is protected separately as the floor group)
    "patent_count",
    # NOC
    "noc_brand_name", "noc_company", "noc_date",
    "noc_submission_type", "noc_therapeutic_class",
    # Labeling
    "active_ingredient", "excipients_core", "excipients_coating", "preservatives",
    "pack_size", "pack_style",
    "colour", "shape", "size_mm", "weight", "ph",
    # Data protection (always present even when no record matches)
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
})

# Canonical Sheet 1 column order.  Patent N_* columns are inserted between the
# pre-patent and post-patent groups at assembly time (count varies per dataset).
_SHEET1_PRE_PATENT_COLS = (
    "din", "ingredient", "brand_name", "company", "strength", "dosage_form",
    "route", "status", "_drug_code", "_schedule", "_last_update",
    "noc_brand_name", "noc_company", "noc_date",
    "noc_submission_type", "noc_therapeutic_class",
    "patent_count",
)
_SHEET1_POST_PATENT_COLS = (
    "active_ingredient", "excipients_core", "excipients_coating", "preservatives",
    "pack_size", "pack_style", "colour", "shape", "size_mm",
    "weight", "ph", "needs_ocr",
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
)

# Regex to identify patent_N group columns (the four columns per patent slot).
_PATENT_GROUP_RE = re.compile(
    r"^patent_(\d+)_(number|filing_date|grant_date|expiry_date)$"
)


def _is_empty_for_fill(v: Any) -> bool:
    """True if v counts as empty for fill-rate purposes.

    None, NaN, "", and whitespace-only strings are empty.
    Sentinel strings ("No NOC record", "Not in PM", "No", …) are NOT empty —
    they represent a real, meaningful absence and protect the column.
    """
    if v is None:
        return True
    try:
        if pd.isna(v):
            return True
    except (TypeError, ValueError):
        pass
    return str(v).strip() == ""


def _col_fill_rate(series: "pd.Series[Any]", n_rows: int) -> float:
    return int(series.apply(lambda v: not _is_empty_for_fill(v)).sum()) / n_rows


def _prune_sparse_columns(
    df: pd.DataFrame,
    min_fill_rate: float = WORKBOOK_MIN_FILL_RATE,
) -> pd.DataFrame:
    """Drop Sheet 1 columns whose non-empty fill rate is at or below min_fill_rate.

    Patent groups (patent_N_number/filing/grant/expiry) are evaluated and
    dropped together so the wide layout stays aligned.  The report is printed
    to stdout so it appears in CLI and server logs.
    """
    if df.empty:
        return df

    n_rows = len(df)
    cols_before = len(df.columns)
    cols_to_drop: list[str] = []
    report_lines: list[str] = []

    # ── Collect patent groups ────────────────────────────────────────────────
    patent_groups: dict[int, list[str]] = {}
    for col in df.columns:
        m = _PATENT_GROUP_RE.match(col)
        if m:
            patent_groups.setdefault(int(m.group(1)), []).append(col)
    patent_all_cols: set[str] = {c for cols in patent_groups.values() for c in cols}

    # Prune patent tail groups: walk from the HIGHEST group downward, dropping
    # sparse groups until we reach a dense group (or hit the patent_1 floor).
    # patent_1 is NEVER pruned by the threshold — it is the minimum patent slot
    # and its dates represent real data for any DIN that has a patent.
    if patent_groups:
        for n in sorted(patent_groups, reverse=True):
            if n == 1:
                break  # patent_1 is the unconditional floor — stop here
            num_col = f"patent_{n}_number"
            if num_col not in df.columns:
                continue
            fr = _col_fill_rate(df[num_col], n_rows)
            if fr <= min_fill_rate:
                n_filled = round(fr * n_rows)
                cols_to_drop.extend(patent_groups[n])
                report_lines.append(
                    f"  patent_{n} group (4 cols): {n_filled}/{n_rows} = {fr:.1%} fill → dropped"
                )
            else:
                break  # Dense group found — keep this one and everything below it

    # ── Non-patent columns ───────────────────────────────────────────────────
    for col in df.columns:
        if col in patent_all_cols or col in _NEVER_DROP_COLS:
            continue
        fr = _col_fill_rate(df[col], n_rows)
        if fr <= min_fill_rate:
            n_filled = round(fr * n_rows)
            cols_to_drop.append(col)
            report_lines.append(
                f"  {col}: {n_filled}/{n_rows} = {fr:.1%} fill → dropped"
            )

    cols_after = cols_before - len(cols_to_drop)
    print(f"\n=== Workbook column cleanup (min_fill_rate={min_fill_rate:.1%}) ===")
    if report_lines:
        for line in report_lines:
            print(line)
    else:
        print("  (no columns dropped)")
    print(f"  Columns: {cols_before} → {cols_after}")
    print("=" * 52)

    return df.drop(columns=cols_to_drop)


# ── Sheet 1 helpers ───────────────────────────────────────────────────────────

def _is_excluded_din(din: Optional[str]) -> bool:
    return din is None or din.strip().lower() in _EXCLUDED_DIN_VALUES


def _collect_dpd_rows(records: list[DrugRecord]) -> dict[str, dict[str, Any]]:
    """Build DIN-keyed dict from DPD records. record_url excluded per Change 2."""
    out: dict[str, dict[str, Any]] = {}
    for r in records:
        if r.source != "DPD" or _is_excluded_din(r.din):
            continue
        din = r.din.strip()  # type: ignore[union-attr]
        out[din] = {
            "din": din,
            "brand_name": r.brand_name,
            "company": r.company,
            "ingredient": r.ingredient,
            "strength": r.strength,
            "dosage_form": r.dosage_form,
            "route": r.route,
            "status": r.status,
            "_drug_code": r.source_specific.get("drug_code"),
            "_schedule": r.source_specific.get("schedule"),
            "_last_update": r.source_specific.get("last_update_date"),
        }
    return out


def _collect_noc_rows(records: list[DrugRecord]) -> dict[str, dict[str, Any]]:
    """Build DIN-keyed dict from NOC records.

    Supplement rows (SNDS / SANDS) are dropped — only NDS, ANDS, and unknown
    types are included.  noc_record_url excluded per Change 2.
    """
    out: dict[str, dict[str, Any]] = {}
    for r in records:
        if r.source != "NOC" or _is_excluded_din(r.din):
            continue
        sub_type = r.source_specific.get("submission_type") or ""
        if _SUPPLEMENT_TYPE_RE.search(sub_type):
            continue  # drop SNDS / SANDS rows
        din = r.din.strip()  # type: ignore[union-attr]
        out[din] = {
            "noc_brand_name": r.brand_name,
            "noc_company": r.company,
            "noc_date": r.source_specific.get("noc_date"),
            "noc_submission_type": sub_type or None,
            "noc_therapeutic_class": r.source_specific.get("therapeutic_class"),
        }
    return out


def _aggregate_patents_wide(din: str, max_patents: int) -> dict[str, Any]:
    """Return wide patent block: patent_count + max_patents column groups of 4.

    Each group: patent_N_number, patent_N_filing_date, patent_N_grant_date,
    patent_N_expiry_date.  Groups beyond the DIN's actual count are all None.
    max_patents must be >= 1.
    """
    rows = get_patents_for_din(din)
    out: dict[str, Any] = {"patent_count": len(rows)}
    for i in range(1, max_patents + 1):
        if i <= len(rows):
            r = rows[i - 1]
            out[f"patent_{i}_number"] = r["patent_number"]
            out[f"patent_{i}_filing_date"] = r.get("filing_date")
            out[f"patent_{i}_grant_date"] = r.get("grant_date")
            out[f"patent_{i}_expiry_date"] = r.get("expiry_date")
        else:
            out[f"patent_{i}_number"] = None
            out[f"patent_{i}_filing_date"] = None
            out[f"patent_{i}_grant_date"] = None
            out[f"patent_{i}_expiry_date"] = None
    return out


def _get_labeling_cols(din: str) -> dict[str, Any]:
    """Return labeling fields for a DIN.

    *_page citation columns and labeling_pdf_url are excluded per Change 2.
    needs_ocr is kept as a live extraction-provenance flag.
    """
    row = get_labeling_for_din(din)
    out: dict[str, Any] = {}
    for field in _LABELING_FIELDS:
        out[field] = row.get(field) if row else None
        # _page columns intentionally omitted from workbook output
    out["needs_ocr"] = bool(row.get("needs_ocr")) if row else None
    # labeling_pdf_url intentionally omitted from workbook output
    return out


def _get_dp_cols(
    dpd_ingredient: Optional[str],
    dpd_company: Optional[str],
    dp_table: Optional[list[dict]],
) -> dict[str, Any]:
    """Return data protection fields for a DIN, or blanks when dp_table is None."""
    blank = {"dp_6yr_no_file_date": None, "pediatric_extension": None, "data_protection_ends": None}
    if dp_table is None:
        return blank
    matched = _match_data_protection_deterministic(
        dpd_ingredient or "", dpd_company or "", dp_table
    )
    if matched:
        return matched
    return blank


def _col_is_all_empty(series: pd.Series) -> bool:
    """True iff every value in the series is None/NaN/empty string/whitespace.

    Sentinel strings ("No NOC record", "No PM available", "Not in PM", …) are
    NOT empty — they carry real information and prevent the column from dropping.
    """
    for val in series:
        if val is None:
            continue
        try:
            if pd.isna(val):
                continue
        except (TypeError, ValueError):
            pass
        s = str(val).strip()
        if not s or s.lower() in ("none", "nan"):
            continue
        # Has at least one real value (including any sentinel string)
        return False
    return True


def _drop_empty_sheet1_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Drop all-empty columns from Sheet 1 before writing.

    Patent groups (patent_N_number + 3 date cols) are evaluated and dropped as
    a unit: only when ALL four columns in the group are all-empty for all rows.
    This keeps layout aligned when some patents have numbers but missing dates.

    All other columns are dropped individually when all-empty.

    "Empty" means None/NaN/''/whitespace only.  Sentinels like
    'No NOC record', 'No PM available', 'Not in PM' count as REAL data.
    """
    if df.empty:
        return df

    cols_to_drop: list[str] = []
    patent_group_cols: set[str] = set()

    # 1. Patent groups — drop only if the whole group is all-empty
    pat_nums = sorted({
        int(m.group(1))
        for c in df.columns
        if (m := re.match(r"^patent_(\d+)_", c))
    })
    for n in pat_nums:
        grp = [c for c in df.columns if c.startswith(f"patent_{n}_")]
        patent_group_cols.update(grp)
        if all(_col_is_all_empty(df[c]) for c in grp):
            cols_to_drop.extend(grp)

    # 2. Non-patent columns — drop individually if all-empty.
    # Protected schema columns (_NEVER_DROP_COLS) are always kept even when
    # all-empty, because they are part of the declared output contract.
    for col in df.columns:
        if col in patent_group_cols or col in _NEVER_DROP_COLS:
            continue
        if _col_is_all_empty(df[col]):
            cols_to_drop.append(col)

    if cols_to_drop:
        logger.info(
            "Dropping %d all-empty columns from Sheet 1: %s",
            len(cols_to_drop), cols_to_drop,
        )
        print(f"[workbook] Dropping {len(cols_to_drop)} all-empty Sheet 1 columns: {cols_to_drop}")

    return df.drop(columns=cols_to_drop)


def build_sheet1(
    response: SearchResponse,
    dp_table: Optional[list[dict]] = None,
) -> pd.DataFrame:
    """Build Sheet 1: one row per DIN, DPD + NOC + wide patents + labeling + data protection."""
    all_records = [r for s in response.sources for r in s.records]

    dpd_by_din = _collect_dpd_rows(all_records)
    noc_by_din = _collect_noc_rows(all_records)

    all_dins = sorted(set(dpd_by_din) | set(noc_by_din))
    if not all_dins:
        return pd.DataFrame()

    # Compute M = max patents across all DINs in this result set (at least 1)
    max_patents = max(
        (len(get_patents_for_din(din)) for din in all_dins),
        default=0,
    )
    max_patents = max(max_patents, 1)

    rows = []
    for din in all_dins:
        row: dict[str, Any] = {"din": din}
        row.update(dpd_by_din.get(din, {}))
        # DPD products with no NOC record (or whose NOC submission was filtered out)
        # receive explicit "No NOC record" labels so blanks are unambiguous.
        noc_data = noc_by_din.get(din)
        row.update(noc_data if noc_data is not None else _NO_NOC_RECORD)
        row.update(_aggregate_patents_wide(din, max_patents))
        row.update(_get_labeling_cols(din))
        dpd_rec = dpd_by_din.get(din, {})
        row.update(_get_dp_cols(dpd_rec.get("ingredient"), dpd_rec.get("company"), dp_table))
        rows.append(row)

    df = pd.DataFrame(rows)

    # Apply the canonical column order defined by _SHEET1_PRE/POST_PATENT_COLS.
    # Patent_N_* columns are dynamic (count varies), so they are inserted between
    # the two fixed groups, sorted by patent slot number.
    present = set(df.columns)
    patent_cols = sorted(
        (c for c in present if _PATENT_GROUP_RE.match(c)),
        key=lambda c: (int(_PATENT_GROUP_RE.match(c).group(1)),  # type: ignore[union-attr]
                       ("number", "filing_date", "grant_date", "expiry_date").index(
                           _PATENT_GROUP_RE.match(c).group(2))),  # type: ignore[union-attr]
    )
    ordered = (
        [c for c in _SHEET1_PRE_PATENT_COLS if c in present]
        + patent_cols
        + [c for c in _SHEET1_POST_PATENT_COLS if c in present]
    )
    # Append any remaining columns not yet listed (future-proofing)
    ordered += [c for c in df.columns if c not in set(ordered)]

    df = df[ordered].sort_values("din", kind="stable").reset_index(drop=True)
    df = _drop_empty_sheet1_cols(df)
    return _prune_sparse_columns(df)


# ── Sheet 2 helpers ───────────────────────────────────────────────────────────

def _ingredient_matches(record_ingredient: Optional[str], query: str) -> bool:
    """Return True if query is contained in the record's ingredient string."""
    if not record_ingredient:
        return False
    q = re.sub(r"\s+", " ", query.strip()).lower()
    ing = re.sub(r"\s+", " ", record_ingredient.strip()).lower()
    return q in ing


def build_sheet2(response: SearchResponse) -> pd.DataFrame:
    """Build Sheet 2: Generic Submissions filtered to the queried ingredient."""
    query = response.metadata.query
    rows = []
    for sr in response.sources:
        if sr.source != "GenericSubmissions":
            continue
        for r in sr.records:
            if not _ingredient_matches(r.ingredient, query):
                continue
            rows.append({
                "medicinal_ingredient": r.ingredient,
                "company": r.company,
                "therapeutic_area": r.source_specific.get("therapeutic_area"),
                "year_month_accepted": r.source_specific.get("date_accepted"),
                "status": r.status,
            })

    if not rows:
        return pd.DataFrame(
            columns=["medicinal_ingredient", "company", "therapeutic_area",
                     "year_month_accepted", "status"]
        )
    return pd.DataFrame(rows)


# ── Workbook assembly ─────────────────────────────────────────────────────────

def _style_sheet(worksheet: Any, df: pd.DataFrame) -> None:
    """Apply bold header, freeze row, autofilter, and autosized columns."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=False)

    worksheet.freeze_panes = "A2"

    if not df.empty:
        last_col = get_column_letter(len(df.columns))
        worksheet.auto_filter.ref = f"A1:{last_col}1"

    for i, col in enumerate(df.columns, 1):
        max_val_len = (
            df[col].fillna("").astype(str).str.len().max()
            if not df.empty else 0
        )
        width = min(max(len(str(col)) + 2, int(max_val_len or 0) + 2), 60)
        worksheet.column_dimensions[get_column_letter(i)].width = width

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Calibri", size=10)


def _build_status_sheet(
    writer: Any,
    response: SearchResponse,
    source_errors: dict[str, Optional[str]],
) -> None:
    """Write a per-source status warning sheet (used when allow_partial=True)."""
    rows = []
    for src in response.sources:
        rows.append({
            "source": src.source,
            "status": src.status,
            "record_count": src.count,
            "error_message": src.error_message or "",
            "warning": (
                "⚠ DATA MISSING FROM THIS EXPORT"
                if src.status == "error"
                else ""
            ),
        })
    df = pd.DataFrame(
        rows,
        columns=["source", "status", "record_count", "error_message", "warning"],
    )
    sheet_name = "⚠ Source Status"
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    _style_sheet(writer.sheets[sheet_name], df)


def build_workbook(
    response: SearchResponse,
    source_errors: Optional[dict[str, Optional[str]]] = None,
    dp_table: Optional[list[dict]] = None,
) -> bytes:
    """Assemble the enriched workbook and return XLSX bytes.

    source_errors: when provided (allow_partial=True path), appends a
    '⚠ Source Status' sheet that visibly flags every failed source.
    dp_table: pre-fetched active data protection register rows (from
    fetch_data_protection_table()); None means the three dp_* columns are blank.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        sheet1 = build_sheet1(response, dp_table=dp_table)
        sheet1.to_excel(writer, sheet_name="DPD + NOC + Patents", index=False)
        _style_sheet(writer.sheets["DPD + NOC + Patents"], sheet1)

        sheet2 = build_sheet2(response)
        sheet2.to_excel(writer, sheet_name="Generic Submissions", index=False)
        _style_sheet(writer.sheets["Generic Submissions"], sheet2)

        if source_errors is not None:
            _build_status_sheet(writer, response, source_errors)

    return buf.getvalue()


# ── Multi-product side-by-side workbook ──────────────────────────────────────

# Light-fill palette for product blocks.  8 distinct accessible colors;
# cycles for 9+ products.  All are light enough that black text remains
# readable; distinct enough that adjacent products are easy to tell apart.
_BLOCK_COLORS: list[str] = [
    "EDD6EB",  # light purple   (#AA55A0 ~15%)
    "CCF0F0",  # light teal     (#00A5A5 ~20%)
    "DDD5EE",  # light deep purple (#3D226E ~20%)
    "CCE7F2",  # light teal-dark   (#008BAD ~20%)
    "F3E5F2",  # pale purple
    "D5F2F2",  # pale teal
    "E5D8F0",  # lavender purple
    "C5EAEA",  # seafoam teal
]

# Spacer columns written between adjacent product blocks
_BLOCK_SPACER = 1

# Medium tints of Zydus purple/teal for banner rows (black text on these)
_BLOCK_BANNER_COLORS: list[str] = [
    "D4A8D0",  # medium purple   (#AA55A0 ~50%)
    "80CECE",  # medium teal     (#00A5A5 ~50%)
    "9D8AC4",  # medium deep purple (#3D226E ~50%)
    "7FC4D8",  # medium teal-dark   (#008BAD ~50%)
    "E0B8DC",  # light-medium purple
    "99D8D8",  # light-medium teal
    "C4AADA",  # lavender purple
    "88CCCC",  # teal variant
]


def _block_color(idx: int) -> str:
    return _BLOCK_COLORS[idx % len(_BLOCK_COLORS)]


def _block_banner_color(idx: int) -> str:
    return _BLOCK_BANNER_COLORS[idx % len(_BLOCK_BANNER_COLORS)]


def _safe_cell_val(val: Any) -> Any:
    """Convert a pandas/numpy value to a plain Python type for openpyxl."""
    import numpy as np  # numpy is a pandas dependency; always available
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, np.integer):
        return int(val)
    if isinstance(val, np.floating):
        return float(val)
    if isinstance(val, np.bool_):
        return bool(val)
    return val


def _write_multiblock_sheet(
    ws: Any,
    blocks: list[tuple[str, pd.DataFrame]],
    colors: list[str],
    banner_colors: list[str],
    spacer_cols: int = _BLOCK_SPACER,
) -> None:
    """Write N product DataFrames side-by-side on one openpyxl worksheet.

    Layout:
      Row 1  — KEY LEGEND: "PRODUCT KEY:" label + one colored cell per product.
      Row 2  — BANNER: merged-cell product name, colored, across each block.
      Row 3  — HEADERS: column names for each block (standard gray fill).
      Row 4+ — DATA: per-DIN rows, top-aligned (ragged heights are expected).

    freeze_panes is set to "A4" so the key, banner, and header rows stay
    visible when scrolling down.  No column freeze is applied since
    side-by-side blocks require horizontal scrolling.

    Autofilter is placed on the first block's header row (openpyxl supports
    only one autofilter per sheet).

    Between blocks, `spacer_cols` empty columns act as visual separators.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from openpyxl.styles import Border, Side

    KEY_ROW = 1
    BANNER_ROW = 2
    HEADER_ROW = 3
    DATA_START = 4

    # ── Style constants ────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill(start_color="3D226E", end_color="3D226E", fill_type="solid")
    KEY_LABEL_FILL = PatternFill(start_color="3D226E", end_color="3D226E", fill_type="solid")
    WHITE_FILL   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    _thin = Side(style="thin", color="D1D1D1")
    _med  = Side(style="medium", color="A0A0A0")
    CELL_BORDER   = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    HEADER_BORDER = Border(left=_thin, right=_thin, top=_med,  bottom=_med)

    # Status → font color (dark readable shades)
    _STATUS_COLOR = {
        "marketed":  "1A6B3C",
        "approved":  "1A6B3C",
        "cancelled": "9B1C1C",
        "dormant":   "7A4F00",
        "inactive":  "7A4F00",
    }

    # Minimum column widths for well-known short fields
    _MIN_WIDTHS = {
        "din": 12, "status": 13, "form": 13, "route": 13,
        "strength": 14, "noc_date": 13, "patent_count": 10,
    }

    # ── Compute column ranges for each block ──────────────────────────────────
    block_col_ranges: list[tuple[int, int]] = []
    col_cursor = 1
    for _, df in blocks:
        n_cols = len(df.columns) if not df.empty and len(df.columns) > 0 else 1
        block_col_ranges.append((col_cursor, col_cursor + n_cols - 1))
        col_cursor += n_cols + spacer_cols

    # ── Row 1: Key legend ─────────────────────────────────────────────────────
    label_cell = ws.cell(row=KEY_ROW, column=1)
    label_cell.value = "PRODUCT KEY:"
    label_cell.font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
    label_cell.fill = KEY_LABEL_FILL
    label_cell.alignment = Alignment(horizontal="left", vertical="center")
    label_cell.border = CELL_BORDER
    ws.column_dimensions["A"].width = 15

    for i, ((name, _df), color) in enumerate(zip(blocks, colors)):
        key_col = 2 + i
        cell = ws.cell(row=KEY_ROW, column=key_col)
        cell.value = name
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(key_col)].width = max(len(name) + 4, 16)

    ws.row_dimensions[KEY_ROW].height = 22

    # ── Row 2: Banners ────────────────────────────────────────────────────────
    for (name, df), bcolor, (c_start, c_end) in zip(blocks, banner_colors, block_col_ranges):
        n_dins = len(df) if not df.empty else 0
        banner_text = f"{name.upper()}  —  {n_dins} DIN{'s' if n_dins != 1 else ''}"
        cell = ws.cell(row=BANNER_ROW, column=c_start)
        cell.value = banner_text
        cell.fill = PatternFill(start_color=bcolor, end_color=bcolor, fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = Border(left=_med, right=_med, top=_med, bottom=_med)
        if c_end > c_start:
            ws.merge_cells(
                start_row=BANNER_ROW, start_column=c_start,
                end_row=BANNER_ROW, end_column=c_end,
            )
    ws.row_dimensions[BANNER_ROW].height = 30

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    for (name, df), (c_start, _c_end) in zip(blocks, block_col_ranges):
        cols = list(df.columns) if not df.empty else []
        for j, col_name in enumerate(cols):
            cell = ws.cell(row=HEADER_ROW, column=c_start + j)
            cell.value = col_name.replace("_", " ").title()
            cell.font = Font(bold=True, name="Calibri", size=10, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            cell.border = HEADER_BORDER
    ws.row_dimensions[HEADER_ROW].height = 26

    # ── Column widths (computed first — needed for row-height estimation) ─────
    # col_widths_map: excel_col_index → width in chars
    col_widths_map: dict[int, float] = {}
    for (name, df), (c_start, _c_end) in zip(blocks, block_col_ranges):
        for j, col_name in enumerate(df.columns if not df.empty else []):
            max_val_len = (
                df[col_name].fillna("").astype(str).str.len().max()
                if not df.empty else 0
            )
            floor = _MIN_WIDTHS.get(col_name.lower(), 0)
            width = min(max(len(str(col_name)) + 4, int(max_val_len or 0) + 3, floor), 42)
            col_widths_map[c_start + j] = width
            ws.column_dimensions[get_column_letter(c_start + j)].width = width

    # ── Rows 4+: Data (wrap_text on, row heights auto-sized) ─────────────────
    import math as _math
    LINE_HEIGHT_PT = 15.0  # Calibri 10pt ≈ 15pt per line

    for (name, df), color, (c_start, _c_end) in zip(blocks, colors, block_col_ranges):
        if df.empty:
            continue
        row_fill_a = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for r_idx, (_idx, row_series) in enumerate(df.iterrows()):
            excel_row = DATA_START + r_idx
            row_fill = row_fill_a if r_idx % 2 == 0 else WHITE_FILL
            max_lines = 1
            for j, col_name in enumerate(df.columns):
                val = _safe_cell_val(row_series[col_name])
                cell = ws.cell(row=excel_row, column=c_start + j)
                cell.value = val
                cell.fill = row_fill
                cell.border = CELL_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                status_key = str(val).lower().strip() if val is not None else ""
                if col_name == "status" and status_key in _STATUS_COLOR:
                    cell.font = Font(
                        bold=True, name="Calibri", size=10,
                        color=_STATUS_COLOR[status_key],
                    )
                else:
                    cell.font = Font(name="Calibri", size=10)
                # Estimate lines this cell needs given the column width
                col_w = col_widths_map.get(c_start + j, 20)
                text_len = len(str(val)) if val is not None else 0
                lines = _math.ceil(text_len / max(col_w, 1)) if text_len else 1
                max_lines = max(max_lines, lines)
            ws.row_dimensions[excel_row].height = max(LINE_HEIGHT_PT, max_lines * LINE_HEIGHT_PT)

    # ── Freeze top 3 rows (key + banner + header) ─────────────────────────────
    ws.freeze_panes = "A4"

    # ── Autofilter on the first block's header row ─────────────────────────────
    if block_col_ranges and blocks and not blocks[0][1].empty:
        c_start, c_end = block_col_ranges[0]
        ws.auto_filter.ref = (
            f"{get_column_letter(c_start)}{HEADER_ROW}:"
            f"{get_column_letter(c_end)}{HEADER_ROW}"
        )


def build_workbook_multiproduct(
    products: list[tuple[str, SearchResponse]],
    source_errors: Optional[dict[str, Optional[str]]] = None,
    dp_table: Optional[list[dict]] = None,
) -> "tuple[bytes, pd.DataFrame, pd.DataFrame]":
    """Build a side-by-side multi-product two-tab workbook.

    Each product in ``products`` becomes one color-coded horizontal block on
    both Tab 1 ("DPD + NOC + Patents") and Tab 2 ("Generic Submissions").  The
    same color is used on both tabs so a product is visually consistent.

    Single-product is a degenerate case (one block, no spacer) and produces
    the same underlying data as ``build_workbook``.

    Returns (xlsx_bytes, combined_sheet1_df, combined_sheet2_df).
    The combined DataFrames have a ``product`` column prepended (used by the
    dashboard JSON view) and are a flat concatenation of all product blocks.
    """
    import openpyxl

    colors = [_block_color(i) for i in range(len(products))]
    banner_colors = [_block_banner_color(i) for i in range(len(products))]

    # ── Build per-product DataFrames ──────────────────────────────────────────
    sheet1_blocks: list[tuple[str, pd.DataFrame]] = []
    sheet2_blocks: list[tuple[str, pd.DataFrame]] = []
    for name, response in products:
        s1 = build_sheet1(response, dp_table=dp_table)
        s2 = build_sheet2(response)
        sheet1_blocks.append((name, s1))
        sheet2_blocks.append((name, s2))

    # ── Assemble XLSX via openpyxl directly (not pandas ExcelWriter) ──────────
    buf = io.BytesIO()
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "DPD + NOC + Patents"
    _write_multiblock_sheet(ws1, sheet1_blocks, colors, banner_colors)

    ws2 = wb.create_sheet(title="Generic Submissions")
    _write_multiblock_sheet(ws2, sheet2_blocks, colors, banner_colors)

    if source_errors is not None:
        _build_status_sheet_multi(wb, products, source_errors)

    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    # ── Build combined flat DataFrames for the dashboard (JSON snapshot) ──────
    def _concat_with_product_col(
        named_dfs: list[tuple[str, pd.DataFrame]],
    ) -> pd.DataFrame:
        frames = []
        for name, df in named_dfs:
            if not df.empty:
                dfc = df.copy()
                dfc.insert(0, "product", name)
                frames.append(dfc)
        if not frames:
            return pd.DataFrame()
        return pd.concat(frames, ignore_index=True, sort=False)

    combined_s1 = _concat_with_product_col(sheet1_blocks)
    combined_s2 = _concat_with_product_col(sheet2_blocks)

    return xlsx_bytes, combined_s1, combined_s2


def _build_status_sheet_multi(
    wb: Any,
    products: list[tuple[str, SearchResponse]],
    source_errors: dict[str, Optional[str]],
) -> None:
    """Append a ⚠ Source Status sheet when allow_partial=True."""
    from openpyxl.styles import Font, PatternFill
    ws = wb.create_sheet(title="⚠ Source Status")
    headers = ["product", "source", "status", "record_count", "error_message", "warning"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = h
        cell.font = Font(bold=True, name="Calibri", size=10)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    row_idx = 2
    for name, response in products:
        for src in response.sources:
            ws.cell(row=row_idx, column=1).value = name
            ws.cell(row=row_idx, column=2).value = src.source
            ws.cell(row=row_idx, column=3).value = src.status
            ws.cell(row=row_idx, column=4).value = src.count
            ws.cell(row=row_idx, column=5).value = src.error_message or ""
            ws.cell(row=row_idx, column=6).value = (
                "⚠ DATA MISSING FROM THIS EXPORT" if src.status == "error" else ""
            )
            row_idx += 1


# ── CLI ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import argparse
    import asyncio
    import os

    parser = argparse.ArgumentParser(description="Build the two-tab enriched workbook.")
    parser.add_argument("--q", required=True, help="Search query")
    parser.add_argument("--field", default="ingredient",
                        help="ingredient | brand | company | din")
    parser.add_argument("--out", default=None, help="Output file path (.xlsx)")
    _args = parser.parse_args()

    async def _run() -> None:
        from app.sources.dpd import search_dpd
        from app.sources.generic_submissions import search_generic_submissions
        from app.sources.noc import search_noc
        from app.sources.patent_register import search_patent_register
        from app.normalize import normalize_query
        from app.models import SearchMetadata
        from app.enrichment.data_protection import fetch_data_protection_table
        from datetime import datetime, timezone

        canonical, extra_terms = await normalize_query(_args.q, _args.field)
        sources = await asyncio.gather(
            search_dpd(canonical, _args.field, extra_terms),
            search_generic_submissions(canonical, _args.field, extra_terms),
            search_noc(canonical, _args.field, extra_terms),
            search_patent_register(canonical, _args.field, extra_terms),
        )
        response = SearchResponse(
            metadata=SearchMetadata(
                query=_args.q,
                field=_args.field,
                timestamp=datetime.now(timezone.utc).isoformat(),
                normalized_terms=[canonical] + extra_terms,
            ),
            sources=list(sources),
        )

        # Enrich patents
        from app.enrichment.patents import enrich_patents
        all_dins = [
            r.din for s in response.sources for r in s.records
            if r.din and not _is_excluded_din(r.din)
        ]
        if all_dins:
            await enrich_patents(all_dins)

        dp_table = await fetch_data_protection_table()
        xlsx = build_workbook(response, dp_table=dp_table)
        out_path = _args.out or f"enriched_{_args.q.replace(' ', '_')}_{_args.field}.xlsx"
        with open(out_path, "wb") as fh:
            fh.write(xlsx)
        print(f"Wrote {os.path.getsize(out_path):,} bytes → {out_path}")

    asyncio.run(_run())
