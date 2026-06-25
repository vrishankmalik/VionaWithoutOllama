"""ACCEPTANCE-level suite: COMPOSITION of the screen criteria driven through
``build_filtered_workbook`` to the produced two-tab XLSX.

Complements:
  * tests/test_screen.py / test_screen_filters.py — UNIT-level compute/apply/parse.
  * tests/test_acpt_filters_each.py — each of the 8 criteria individually → workbook.

This file proves the COMBINED behavior at the workbook artifact level:
  * AND-composition of a numeric + dosage_form + no_file_date isolates one product
    on BOTH tabs.
  * An over-constrained filter yields a VALID empty workbook (sheets still
    ["Summary","Detail"], Summary header-only, Detail empty, no crash).
  * Backward-compat identity: with dosage_bases=None and date_filter=None the
    qualifying frame equals apply_criteria-only output, and the workbook Summary
    matches that frame's products.
  * IQVIA-absent guard: a value criterion without MAT columns raises ValueError
    mentioning IQVIA.
  * Realistic-large: ~4,000 synthetic products run through a moderately selective
    combined filter; the Summary survivor count matches the hand-computed number
    and the build completes.

Fully offline, hermetic, order-independent.  "Today" is pinned to date(2026,6,24).
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
import pandas as pd

from app.enrichment import screen as S
from app.enrichment.screen import (
    NoFileDateFilter,
    apply_criteria,
    build_filtered_workbook,
    build_summary_sheet,
    compute_products,
    filter_products,
    parse_criteria,
    parse_dosage_forms,
    parse_no_file_date,
)

_TODAY = date(2026, 6, 24)

_SUMMARY_HEADERS = [
    "Ingredient", "Dosage Form", "Number of Competitors",
    "Number of Filings", "Number of Approvals", "Value Sizeable ($)",
    "Quantity Sizeable (Units)", "Quantity Ext Sizeable",
]

# IQVIA latest-period column headers (mirror the real header form).
_DOLLARS = "Dollars MAT 12/2025"
_UNITS = "Units MAT 12/2025"
_EXT = "Ext Units MAT 12/2025"


# ── readback helpers ──────────────────────────────────────────────────────────

def _open(xlsx: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(xlsx))


def _assert_shell(wb: openpyxl.Workbook) -> None:
    assert wb.sheetnames == ["Summary", "Detail"]
    headers = [c.value for c in wb["Summary"][1]]
    assert headers == _SUMMARY_HEADERS


def _summary_products(wb: openpyxl.Workbook) -> set[tuple]:
    sws = wb["Summary"]
    out: set[tuple] = set()
    for r in range(2, sws.max_row + 1):
        ing = sws.cell(r, 1).value
        form = sws.cell(r, 2).value
        if ing is None and form is None:
            continue
        out.add((ing, form))
    return out


def _summary_data_rows(wb: openpyxl.Workbook) -> int:
    return len(_summary_products(wb))


def _detail_dins(wb: openpyxl.Workbook) -> set:
    dws = wb["Detail"]
    dheaders = [c.value for c in dws[1]]
    if "DIN" not in dheaders:
        return set()
    col = dheaders.index("DIN") + 1
    return {
        dws.cell(r, col).value
        for r in range(2, dws.max_row + 1)
        if dws.cell(r, col).value is not None
    }


# ── shared composition fixture ────────────────────────────────────────────────
#
#  P_A (A, TABLET)   din 1,2   competitors=2 (two marketed cos)  date 2030-01-01
#  P_B (B, TABLET)   din 3     competitors=1                     date 2020-01-01 (past, parseable)
#  P_C (C, CAPSULE)  din 4     competitors=1                     date 2030-01-01
#
# competitors>1 AND base TABLET AND date>2028  → only P_A survives.

def _mixed_sheet1() -> pd.DataFrame:
    return pd.DataFrame([
        {"din": "1", "ingredient": "A", "dosage_form": "TABLET", "company": "C1",
         "status": "marketed", "dp_6yr_no_file_date": "2030-01-01"},
        {"din": "2", "ingredient": "A", "dosage_form": "TABLET", "company": "C2",
         "status": "marketed", "dp_6yr_no_file_date": "2030-01-01"},
        {"din": "3", "ingredient": "B", "dosage_form": "TABLET", "company": "C3",
         "status": "marketed", "dp_6yr_no_file_date": "2020-01-01"},
        {"din": "4", "ingredient": "C", "dosage_form": "CAPSULE", "company": "C4",
         "status": "marketed", "dp_6yr_no_file_date": "2030-01-01"},
    ])


# ══════════════════════════════════════════════════════════════════════════════
# AND-composition: numeric + dosage_form + no_file_date → one survivor
# ══════════════════════════════════════════════════════════════════════════════

def test_and_composition_isolates_single_product_both_tabs():
    s1 = _mixed_sheet1()
    crit = parse_criteria([{"metric": "competitors", "operator": "above", "value": 1}])
    bases = parse_dosage_forms([{"metric": "dosage_form", "value": ["TABLET"]}])
    df = parse_no_file_date(
        [{"metric": "no_file_date", "operator": "greater", "value": "01/01/2028"}],
        today=_TODAY)

    xlsx, summary, detail, warnings = build_filtered_workbook(
        s1, pd.DataFrame(), crit, dosage_bases=bases, date_filter=df)
    assert warnings == []

    wb = _open(xlsx)
    _assert_shell(wb)
    # competitors>1 keeps only P_A (2 marketed); TABLET drops P_C; date>2028 drops
    # P_B (2020).  P_A is the sole survivor.
    assert _summary_products(wb) == {("A", "TABLET")}
    assert _detail_dins(wb) == {"1", "2"}
    # The returned frames agree with the file.
    assert list(summary["Ingredient"]) == ["A"]
    assert set(detail["DIN"]) == {"1", "2"}


# ══════════════════════════════════════════════════════════════════════════════
# Over-constrained → valid EMPTY workbook (no crash)
# ══════════════════════════════════════════════════════════════════════════════

def test_over_constrained_yields_valid_empty_workbook():
    s1 = _mixed_sheet1()
    # approvals above 9999 AND dosage LOTION — nothing can satisfy either.
    crit = parse_criteria([{"metric": "approvals", "operator": "above", "value": 9999}])
    bases = parse_dosage_forms([{"metric": "dosage_form", "value": ["LOTION"]}])

    xlsx, summary, detail, warnings = build_filtered_workbook(
        s1, pd.DataFrame(), crit, dosage_bases=bases)

    wb = _open(xlsx)
    _assert_shell(wb)              # sheets + header order intact
    assert _summary_data_rows(wb) == 0
    assert wb["Summary"].max_row == 1   # header row only, no data rows
    assert _detail_dins(wb) == set()
    # Returned frames are empty but carry the display headers.
    assert summary.empty and list(summary.columns) == _SUMMARY_HEADERS
    assert detail.empty


# ══════════════════════════════════════════════════════════════════════════════
# Backward-compat identity: no additive fields ⇒ apply_criteria-only
# ══════════════════════════════════════════════════════════════════════════════

def test_backward_compat_identity_frame_and_workbook():
    s1 = _mixed_sheet1()
    products, _ = compute_products(s1, pd.DataFrame())
    crit = parse_criteria([{"metric": "competitors", "operator": "above", "value": 0}])

    # 1) Frame-level: filter_products with no additive fields == apply_criteria.
    base = apply_criteria(products, crit).reset_index(drop=True)
    additive = filter_products(
        products, crit, dosage_bases=None, date_filter=None).reset_index(drop=True)
    pd.testing.assert_frame_equal(base, additive)

    # 2) Workbook-level: the Summary written equals the same frame's summary view,
    #    and reflects exactly the apply_criteria survivors.
    xlsx, summary, _, _ = build_filtered_workbook(s1, pd.DataFrame(), crit)
    expected_summary = build_summary_sheet(base).reset_index(drop=True)
    pd.testing.assert_frame_equal(summary.reset_index(drop=True), expected_summary)

    wb = _open(xlsx)
    _assert_shell(wb)
    expected_products = set(zip(base["ingredient"], base["dosage_form"]))
    assert _summary_products(wb) == expected_products
    # All three products are marketed (competitors>0): A, B, C all survive.
    assert expected_products == {("A", "TABLET"), ("B", "TABLET"), ("C", "CAPSULE")}


# ══════════════════════════════════════════════════════════════════════════════
# IQVIA-absent guard
# ══════════════════════════════════════════════════════════════════════════════

def test_value_criterion_without_iqvia_columns_raises():
    s1 = pd.DataFrame([
        {"din": "1", "ingredient": "A", "dosage_form": "TABLET",
         "company": "C1", "status": "marketed"},
    ])
    crit = parse_criteria([{"metric": "value", "operator": "above", "value": 1}])
    try:
        build_filtered_workbook(s1, pd.DataFrame(), crit)
    except ValueError as exc:
        assert "IQVIA" in str(exc)
    else:
        raise AssertionError("expected ValueError when value criterion lacks MAT columns")


# ══════════════════════════════════════════════════════════════════════════════
# Realistic-large: ~4,000 synthetic products through a combined filter
# ══════════════════════════════════════════════════════════════════════════════

# Row-count note: N_PRODUCTS products, each a single-DIN product, so Sheet 1 has
# ~4,000 rows.  Hand-computed survivor count is derived purely from the generator
# rules below — never read back from the code under test.
_N_PRODUCTS = 4000


def _large_sheet1() -> pd.DataFrame:
    """Generate N single-DIN products with deterministic, hand-derivable values.

    For product i (0..N-1):
      * ingredient  = f"DRUG{i:05d}"  (unique → one product each)
      * dosage_form = "TABLET" when i % 2 == 0 else "CAPSULE"
      * status      = "marketed"  → competitors = approvals = 1 for every product
      * dp_6yr_no_file_date = "2030-01-01" when i % 3 == 0 else "2020-01-01"

    Combined selective filter:  base TABLET  AND  date > 01/01/2028.
      survivors = products with (i even)  AND  (i % 3 == 0)  ==  i % 6 == 0.
    """
    rows = []
    for i in range(_N_PRODUCTS):
        rows.append({
            "din": f"{i:08d}",
            "ingredient": f"DRUG{i:05d}",
            "dosage_form": "TABLET" if i % 2 == 0 else "CAPSULE",
            "company": f"CO{i}",
            "status": "marketed",
            "dp_6yr_no_file_date": "2030-01-01" if i % 3 == 0 else "2020-01-01",
        })
    return pd.DataFrame(rows)


def test_large_combined_filter_survivor_count_exact():
    s1 = _large_sheet1()
    bases = parse_dosage_forms([{"metric": "dosage_form", "value": ["TABLET"]}])
    df = parse_no_file_date(
        [{"metric": "no_file_date", "operator": "greater", "value": "01/01/2028"}],
        today=_TODAY)

    xlsx, summary, detail, warnings = build_filtered_workbook(
        s1, pd.DataFrame(), [], dosage_bases=bases, date_filter=df)
    assert warnings == []

    # Hand-computed: survivors == {i : i % 6 == 0, 0 <= i < N}.
    expected = sum(1 for i in range(_N_PRODUCTS) if i % 6 == 0)
    assert len(summary) == expected

    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_data_rows(wb) == expected
    # Each survivor is single-DIN → Detail DIN count equals survivor count.
    assert len(_detail_dins(wb)) == expected
    # Spot-check: i=0 survives (TABLET, 2030); i=2 does not (TABLET but 2020);
    # i=3 does not (2030 but CAPSULE); i=6 survives.
    survivors = _summary_products(wb)
    assert ("DRUG00000", "TABLET") in survivors
    assert ("DRUG00006", "TABLET") in survivors
    assert ("DRUG00002", "TABLET") not in survivors
    assert ("DRUG00003", "CAPSULE") not in survivors
