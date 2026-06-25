"""Acceptance tests — Option 2: async export (``POST /export/start`` → SSE →
``/export/result`` / ``/export/filtered-result`` / ``/api/export-data``).

Strategy (documented choice): for the full async lifecycle and content/parity
assertions we DRIVE THE PIPELINE DIRECTLY — build a ``JobState`` via
``create_job`` (which registers it in the in-process job store so the HTTP
endpoints find it), monkeypatch the three network-touching enrichers on the
``app.export_job`` module to hermetic stubs, then ``asyncio.run(run_export_job(...))``
to completion.  This avoids the flakiness of consuming a live SSE stream while a
background ``asyncio.create_task`` runs inside TestClient, while still asserting
on the real emitted events, the real XLSX artifacts, and the real
``/api/export-data`` JSON the dashboard reads.  The conftest search mocks
(mock_dpd / mock_noc / mock_gsur / mock_patent_register) feed the Search stage
from fixtures; ``metformin`` (GLUCOPHAGE DIN 02229895) is the fixture-supported
ingredient with a DPD∩NOC overlap.

For HTTP-level request behavior (``/export/start`` dedup + 400, and the
result/export-data error-code matrix) we use ``TestClient(app)`` directly, with
``run_export_job`` patched to a no-op so no background task leaks.

Parallel-safety: the module-level ``_jobs`` dict in ``app.jobs`` is reset in a
fixture teardown so jobs created here never leak into other xdist workers.
"""
from __future__ import annotations

import asyncio
import io
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── hermetic enricher stubs (signatures match the real callees) ───────────────

async def _stub_patents(dins, on_progress=None):
    return {}


async def _stub_labeling(din_map, **kwargs):
    return {}


async def _stub_dp():
    return []


@pytest.fixture
def clean_jobs():
    """Reset the in-process job store before and after the test (xdist-safe)."""
    import app.jobs as jobs_mod
    jobs_mod._jobs.clear()
    yield jobs_mod
    jobs_mod._jobs.clear()


@pytest.fixture
def fresh_store(tmp_path):
    """Per-test enrichment SQLite store; detached on teardown."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    yield store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich_after.db"))


@pytest.fixture
def hermetic_enrichers(monkeypatch):
    """Patch the export_job network enrichers to local no-op stubs."""
    import app.export_job as ej
    monkeypatch.setattr(ej, "enrich_patents", _stub_patents)
    monkeypatch.setattr(ej, "enrich_labeling_batch_fast", _stub_labeling)
    monkeypatch.setattr(ej, "fetch_data_protection_table", _stub_dp)
    return ej


def _run_job(job, *, allow_partial=False, enable_ocr=False):
    import app.export_job as ej
    asyncio.run(ej.run_export_job(job, allow_partial, enable_ocr))


def _client():
    from fastapi.testclient import TestClient
    from app.main import app
    return TestClient(app)


# ── /export/start — dedup, order preservation, 400 ────────────────────────────

def test_export_start_dedups_case_insensitive_preserves_order(clean_jobs):
    """Duplicate + mixed-case duplicate are dropped; first-seen order preserved.

    run_export_job is patched out so no background task runs — this tests only the
    request schema / dedup logic of the endpoint.
    """
    with patch("app.main.run_export_job", new=AsyncMock(return_value=None)):
        resp = _client().post("/export/start", json={
            "queries": ["Metformin", "metformin", "  METFORMIN ", "Aspirin", "aspirin"],
        })
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert "job_id" in body
    # Case-insensitive dedup, first spelling kept, order preserved.
    assert body["queries"] == ["Metformin", "Aspirin"], body["queries"]


def test_export_start_long_list_dedups(clean_jobs):
    """An absurdly long list (60 entries, all the same word) collapses to one."""
    with patch("app.main.run_export_job", new=AsyncMock(return_value=None)):
        resp = _client().post("/export/start", json={
            "queries": [f"  metformin  " for _ in range(60)],
        })
    assert resp.status_code == 200, resp.text
    assert resp.json()["queries"] == ["metformin"]


def test_export_start_long_unique_list_preserved(clean_jobs):
    """60 distinct ingredients survive in entry order."""
    names = [f"ingredient{i:02d}" for i in range(60)]
    with patch("app.main.run_export_job", new=AsyncMock(return_value=None)):
        resp = _client().post("/export/start", json={"queries": names})
    assert resp.status_code == 200, resp.text
    assert resp.json()["queries"] == names


def test_export_start_empty_is_400(clean_jobs):
    """Whitespace-only / empty queries → 400 (no query provided)."""
    with patch("app.main.run_export_job", new=AsyncMock(return_value=None)):
        c = _client()
        assert c.post("/export/start", json={"queries": ["  ", ""]}).status_code == 400
        assert c.post("/export/start", json={"q": "   "}).status_code == 400
        assert c.post("/export/start", json={}).status_code == 400


# ── Full async lifecycle (driven directly, hermetic) ──────────────────────────

def test_async_lifecycle_completes_with_ordered_stages(
    clean_jobs, fresh_store, hermetic_enrichers,
    mock_dpd, mock_noc, mock_gsur, mock_patent_register,
):
    """Drive a single-ingredient job to completion; assert event/stage contract."""
    job = clean_jobs.create_job("life1", "metformin", "ingredient", queries=["metformin"])
    _run_job(job)

    assert job.status == "complete", job.error

    # Stage ordering: first appearance of each stage follows the pipeline order.
    stage_seq = [e["stage"] for e in job.events if "stage" in e]
    first_idx = {}
    for i, s in enumerate(stage_seq):
        first_idx.setdefault(s, i)
    expected_order = ["Search", "Patents", "Labeling", "DataProtection", "Workbook"]
    seen_in_order = [s for s in expected_order if s in first_idx]
    assert [first_idx[s] for s in seen_in_order] == sorted(first_idx[s] for s in seen_in_order), (
        f"Stages must appear in pipeline order; got {stage_seq}"
    )

    # pct in [0,1] and monotonic non-decreasing across all progress events.
    pcts = [e["pct"] for e in job.events if "pct" in e]
    assert pcts, "expected at least one progress event with pct"
    for p in pcts:
        assert 0.0 <= p <= 1.0, f"pct out of range: {p}"
    assert pcts == sorted(pcts), f"pct must be monotonic non-decreasing; got {pcts}"

    # Final event is a complete event with a download_url.
    final = job.events[-1]
    assert final.get("status") == "complete", final
    assert final.get("download_url") == "/export/result/life1"

    # Result XLSX exists, openpyxl-readable, two canonical tabs.
    assert job.result_path
    wb = openpyxl.load_workbook(job.result_path)
    assert set(wb.sheetnames) == {"DPD + NOC + Patents", "Generic Submissions"}, wb.sheetnames

    # Snapshot is populated.
    assert job.sheet1_columns and job.sheet1_records
    assert "din" in job.sheet1_columns


def test_result_endpoint_serves_finished_xlsx(
    clean_jobs, fresh_store, hermetic_enrichers,
    mock_dpd, mock_noc, mock_gsur, mock_patent_register,
):
    """/export/result/{id} → 200 xlsx after the job completes."""
    job = clean_jobs.create_job("res1", "metformin", "ingredient", queries=["metformin"])
    _run_job(job)
    assert job.status == "complete", job.error

    resp = _client().get("/export/result/res1")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA
    assert resp.content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "DPD + NOC + Patents" in wb.sheetnames


def test_export_data_parity_with_snapshot(
    clean_jobs, fresh_store, hermetic_enrichers,
    mock_dpd, mock_noc, mock_gsur, mock_patent_register,
):
    """/api/export-data/{id} returns the EXACT job snapshot (no re-scrape)."""
    job = clean_jobs.create_job("data1", "metformin", "ingredient", queries=["metformin"])
    _run_job(job)
    assert job.status == "complete", job.error

    resp = _client().get("/api/export-data/data1")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["query"] == "metformin"
    assert body["queries"] == ["metformin"]
    assert body["field"] == "ingredient"
    # Columns + records must equal the in-memory snapshot, byte-for-byte.
    assert body["sheet1"]["columns"] == job.sheet1_columns
    assert body["sheet1"]["records"] == job.sheet1_records
    assert body["sheet2"]["columns"] == job.sheet2_columns
    assert body["sheet2"]["records"] == job.sheet2_records
    # No filter_criteria → no summary section.
    assert "summary" not in body


# ── Error-path matrix ─────────────────────────────────────────────────────────

def test_result_unknown_job_404(clean_jobs):
    """/export/result/{unknown} → 404."""
    assert _client().get("/export/result/does-not-exist").status_code == 404


def test_running_job_returns_409(clean_jobs):
    """A job still in status='running' → 409 on result and export-data; 409 filtered."""
    # create_job leaves status='running' and never starts the pipeline here.
    clean_jobs.create_job("running1", "metformin", "ingredient", queries=["metformin"])
    c = _client()
    assert c.get("/export/result/running1").status_code == 409
    assert c.get("/api/export-data/running1").status_code == 409
    assert c.get("/export/filtered-result/running1").status_code == 409


def test_export_data_errored_job_422(clean_jobs, fresh_store, monkeypatch,
                                     mock_dpd, mock_noc, mock_gsur, mock_patent_register):
    """A job whose pipeline raised → status='error' → /api/export-data 422.

    Force the error by making the Patents stage raise; the other enrichers stay
    hermetic no-ops so the failure is isolated to the forced stage.
    """
    import app.export_job as ej

    async def _boom(dins, on_progress=None):
        raise RuntimeError("forced patents failure")

    monkeypatch.setattr(ej, "enrich_patents", _boom)
    monkeypatch.setattr(ej, "enrich_labeling_batch_fast", _stub_labeling)
    monkeypatch.setattr(ej, "fetch_data_protection_table", _stub_dp)

    job = clean_jobs.create_job("err1", "metformin", "ingredient", queries=["metformin"])
    _run_job(job)
    assert job.status == "error"
    assert job.error and "forced patents failure" in job.error

    c = _client()
    assert c.get("/api/export-data/err1").status_code == 422
    # An errored job is not complete → result endpoint is 409.
    assert c.get("/export/result/err1").status_code == 409


def test_filtered_result_404_when_no_criteria(
    clean_jobs, fresh_store, hermetic_enrichers,
    mock_dpd, mock_noc, mock_gsur, mock_patent_register,
):
    """A completed job with NO filter_criteria → /export/filtered-result 404."""
    job = clean_jobs.create_job("nofilt1", "metformin", "ingredient", queries=["metformin"])
    _run_job(job)
    assert job.status == "complete", job.error
    assert job.filtered_result_path is None

    resp = _client().get("/export/filtered-result/nofilt1")
    assert resp.status_code == 404, resp.text
    # And export-data carries no summary section.
    assert "summary" not in _client().get("/api/export-data/nofilt1").json()


def test_filtered_result_and_summary_when_criteria_given(
    clean_jobs, fresh_store, hermetic_enrichers,
    mock_dpd, mock_noc, mock_gsur, mock_patent_register,
):
    """filter_criteria provided → filtered Summary+Detail XLSX + a 'summary' section.

    'approvals above 0' qualifies every product (each Sheet-1 DIN has an NOC by
    construction), so the filtered workbook is non-empty.
    """
    job = clean_jobs.create_job(
        "filt1", "metformin", "ingredient", queries=["metformin"],
        filter_criteria=[{"metric": "approvals", "operator": "above", "value": 0}],
    )
    _run_job(job)
    assert job.status == "complete", job.error
    assert job.filtered_result_path

    c = _client()
    resp = c.get("/export/filtered-result/filt1")
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert wb.sheetnames == ["Summary", "Detail"], wb.sheetnames

    # The complete event advertises the filtered download URL.
    final = job.events[-1]
    assert final.get("filtered_download_url") == "/export/filtered-result/filt1"

    # export-data now includes a summary section matching the snapshot.
    data = c.get("/api/export-data/filt1").json()
    assert "summary" in data
    assert data["summary"]["columns"] == job.summary_columns
    assert data["summary"]["records"] == job.summary_records
