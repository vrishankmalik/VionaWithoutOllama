"""Offline suite for the two new filters wired through the FULL-UNIVERSE path
(option 4) and the cross-tab data agreement with the Search path (option 2).

Covers (per the locked spec):
  A0 — the universe-build wiring fix: dp_6yr_no_file_date is now POPULATED on the
       universe sheet (was 100% blank), blank only where no Register entry exists;
       cross-tab parity: the same DIN gets the same date through build_sheet1
       (Search) and build_universe_sheet1 (Universe); the read-me no longer claims
       data protection is "NOT evaluated"; the dp fetch rides the 4-h cache + reset.
  Dosage map sourcing — the bundle's base→raw map is derived from the parsed
       catalogue records, not from any filtered/sheet output.
  Option-4 job E2E — run_universe_filter_enrich_job drives the new criteria end to
       end: a matching filter selects the right survivors and enriches only those;
       an over-constrained filter yields a valid empty workbook and fetches NO PDFs.

Real data: DPD identities + dates come from the committed dp_join_products.json /
register_active_sample.html fixtures (build_dp_register_fixture.py) and the real
universe extract slice (build_universe_fixture.py).
"""
from __future__ import annotations

import asyncio
import io
import json
from pathlib import Path

import openpyxl
import pytest
from bs4 import BeautifulSoup

from app.enrichment import universe as U
from app.enrichment.data_protection import _find_active_table, _parse_data_protection_table
from app.enrichment.screen import build_dosage_form_map
from app.enrichment.universe import (
    UniverseBundle,
    build_universe_response,
    build_universe_sheet1,
)
from app.enrichment.workbook import build_sheet1
from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult

_DP_FIX = Path(__file__).parent / "fixtures" / "data_protection"
_UNI_EXTRACT = Path(__file__).parent / "fixtures" / "universe" / "extract"

pytestmark = pytest.mark.skipif(
    not (_DP_FIX / "register_active_sample.html").exists()
    or not (_DP_FIX / "dp_join_products.json").exists(),
    reason="dp fixtures missing — run python -m tests.scripts.build_dp_register_fixture",
)


def _dp_table() -> list[dict]:
    soup = BeautifulSoup((_DP_FIX / "register_active_sample.html").read_text(encoding="utf-8"),
                         "html.parser")
    return _parse_data_protection_table(_find_active_table(soup))


def _join() -> dict:
    return json.loads((_DP_FIX / "dp_join_products.json").read_text(encoding="utf-8"))


@pytest.fixture()
def clean_store(tmp_path):
    """Point the enrichment store at an empty DB so patent/labeling cols are blank."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    return store_mod


def _dpd(p: dict, drug_code: int) -> DrugRecord:
    """A DPD DrugRecord carrying a real (din, ingredient, company, dosage_form)."""
    return DrugRecord(
        source="DPD", din=p["din"], brand_name="BRAND",
        ingredient=p["ingredient"], company=p["company"],
        all_ingredients=[p["ingredient"]], strength="100 mg",
        dosage_form=p.get("dosage_form") or "TABLET",
        status=p.get("status") or "marketed",
        source_specific={"drug_code": drug_code},
    )


def _noc(p: dict) -> DrugRecord:
    return DrugRecord(
        source="NOC", din=p["din"], brand_name="BRAND",
        ingredient=p["ingredient"], company=p["company"],
        source_specific={"noc_date": "2020-01-01", "submission_type": "NDS",
                         "submission_class": "New", "reason_for_supplement": None,
                         "therapeutic_class": "Test"},
    )


def _universe_resp(records: list[DrugRecord]) -> SearchResponse:
    return build_universe_response(UniverseBundle(records, []))


# ════════════════════════════════════════════════════════════════════════════
# A0 — universe build now POPULATES dp (the fix)
# ════════════════════════════════════════════════════════════════════════════

def test_universe_sheet_populates_dp_not_blank(clean_store):
    dp = _dp_table()
    matched = _join()["matched"][:25]
    blanks = _join()["blank"][:10]
    recs = ([_dpd(p, 1000 + i) for i, p in enumerate(matched)]
            + [_dpd(p, 5000 + i) for i, p in enumerate(blanks)])
    df, _r, _l = build_universe_sheet1(_universe_resp(recs), None, dp_table=dp)

    col = df["dp_6yr_no_file_date"].fillna("").astype(str).str.strip()
    nonblank = (col != "").sum()
    assert nonblank > 0, "universe dp column is still all-blank — the fix is not wired"
    by_din = dict(zip(df["din"].astype(str), col))   # NaN already normalized to ""
    for p in matched:
        assert by_din.get(p["din"], "") == p["expected_no_file_date"], p["din"]
    for p in blanks:
        assert by_din.get(p["din"], "") == "", p["din"]


def test_universe_dp_blank_when_dp_table_none(clean_store):
    """Legacy guard: passing dp_table=None keeps the three dp columns blank."""
    recs = [_dpd(p, 1000 + i) for i, p in enumerate(_join()["matched"][:5])]
    df, _r, _l = build_universe_sheet1(_universe_resp(recs), None, dp_table=None)
    col = df["dp_6yr_no_file_date"].fillna("").astype(str).str.strip()
    assert (col != "").sum() == 0


def test_cross_tab_parity_search_vs_universe(clean_store):
    """The SAME DIN gets the SAME six-year date via build_sheet1 (Search) and
    build_universe_sheet1 (Universe) — the fix must make them agree."""
    dp = _dp_table()
    matched = {m["din"]: m for m in _join()["matched"]}
    innovator = matched["02497069"]              # alpelisib / Novartis / 2026-03-11
    generic = _join()["blank"][0]

    recs = [_dpd(innovator, 1001), _dpd(generic, 1002)]
    # Search path: build_sheet1 default needs the DIN in NOC ∩ DPD.
    search_resp = SearchResponse(
        metadata=SearchMetadata(query="alpelisib", field="ingredient", timestamp="t"),
        sources=[SourceResult(source="DPD", status="ok", records=recs),
                 SourceResult(source="NOC", status="ok", records=[_noc(innovator), _noc(generic)])],
    )
    search_df = build_sheet1(search_resp, dp_table=dp)
    universe_df, _r, _l = build_universe_sheet1(_universe_resp(recs), None, dp_table=dp)

    def _date(df, din):
        row = df[df["din"] == din]
        return "" if row.empty else str(row["dp_6yr_no_file_date"].iloc[0] or "")

    assert _date(search_df, "02497069") == "2026-03-11"
    assert _date(search_df, "02497069") == _date(universe_df, "02497069"), "cross-tab dp mismatch"


def test_readme_no_longer_says_dp_not_evaluated():
    caution = U.UNIVERSE_NOT_EVALUATED_CAUTION.lower()
    disclaimer = " ".join(U.UNIVERSE_DISCLAIMER_LINES).lower()
    # Data protection is now evaluated/populated on the universe sheet.
    assert "data protection is evaluated" in caution
    assert "are populated" in disclaimer and "data-protection" in disclaimer
    # The "NOT evaluated" caution now names only NOC + patents (not data protection).
    assert "noc and patent" in caution
    assert "data protection are not evaluated" not in caution


# ════════════════════════════════════════════════════════════════════════════
# Dosage-form map is sourced from the universe records (not sheet output)
# ════════════════════════════════════════════════════════════════════════════

def test_bundle_dosage_map_built_from_records():
    recs = U.load_dpd_universe_records(_UNI_EXTRACT)
    bundle = UniverseBundle(recs, [])
    expected = build_dosage_form_map(r.dosage_form for r in recs)
    assert bundle.dosage_form_map == expected
    assert bundle.dosage_form_map, "real extract must yield at least one base form"
    # Sourced from raw record forms — independent of any filtered sheet.
    assert "TABLET" in bundle.dosage_form_map


# ════════════════════════════════════════════════════════════════════════════
# dp fetch rides the 4-hour universe cache + reset-all-caches invalidation
# ════════════════════════════════════════════════════════════════════════════

def test_dp_fetch_rides_cache_and_reset(monkeypatch):
    recs = U.load_dpd_universe_records(_UNI_EXTRACT)
    dp = _dp_table()
    calls = {"dp": 0}

    async def _fake_fetch():
        calls["dp"] += 1
        return dp

    async def _no_gsur():
        return []

    monkeypatch.setattr(U, "_download_extract", lambda: None)
    monkeypatch.setattr(U, "load_dpd_universe_records", lambda cache_dir=U.UNIVERSE_CACHE_DIR: recs)
    monkeypatch.setattr(U, "_load_gsur_records", _no_gsur)
    monkeypatch.setattr("app.enrichment.data_protection.fetch_data_protection_table", _fake_fetch)
    U._CACHE["bundle"] = None

    b1 = asyncio.run(U.get_universe())
    assert calls["dp"] == 1 and b1.dp_table is dp
    asyncio.run(U.get_universe())                 # cached within window → no refetch
    assert calls["dp"] == 1, "dp must not be refetched while the universe is fresh"

    assert U.reset_universe_cache() == 1          # drops in-process bundle
    asyncio.run(U.get_universe())                 # fresh build → refetch
    assert calls["dp"] == 2, "reset-all-caches must force a fresh dp fetch"
    U._CACHE["bundle"] = None


# ════════════════════════════════════════════════════════════════════════════
# Option-4 job E2E — drive the new criteria through run_universe_filter_enrich_job
# ════════════════════════════════════════════════════════════════════════════

# Deterministic option-4 bundle: two TABLET products + one CAPSULE, fixed DINs.
_TABLET_DINS = {"09990001", "09990003"}
_CAPSULE_DIN = "09990002"


def _prebuilt_bundle() -> UniverseBundle:
    dp = _dp_table()
    matched = _join()["matched"]
    recs = [
        _dpd({**matched[0], "din": "09990001", "dosage_form": "TABLET"}, 9001),
        _dpd({**matched[1], "din": _CAPSULE_DIN, "dosage_form": "CAPSULE"}, 9002),
        _dpd({**_join()["blank"][0], "din": "09990003", "dosage_form": "TABLET"}, 9003),
    ]
    return UniverseBundle(recs, [], dp_table=dp)


def _run_job(filter_criteria, enrich_recorder):
    import app.universe_job as UJ
    from app.jobs import JobState

    bundle = _prebuilt_bundle()

    async def _fake_get_universe(force_refresh=False):
        return bundle

    job = JobState(job_id="t-uni-filter", query="", field="ingredient",
                   filter_criteria=filter_criteria)
    # Avoid network: stub the universe build and the survivor PDF enrichment.
    import unittest.mock as M
    with M.patch.object(UJ, "get_universe", _fake_get_universe), \
         M.patch.object(UJ, "enrich_labeling_batch_fast", enrich_recorder):
        asyncio.run(UJ.run_universe_filter_enrich_job(job, enable_ocr=False))
    return job


def test_option4_matching_filter_selects_and_enriches_survivors(clean_store):
    enriched = {"dins": None}

    async def _recorder(din_map, **kw):
        enriched["dins"] = set(din_map)

    # Dosage form = TABLET only → the CAPSULE product is excluded.
    job = _run_job([{"metric": "dosage_form", "value": ["TABLET"]}], _recorder)
    assert job.status == "complete", job.error
    # Exactly the two TABLET DINs are enriched; the CAPSULE DIN is not.
    assert enriched["dins"] == _TABLET_DINS
    assert _CAPSULE_DIN not in enriched["dins"]
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert wb.sheetnames == ["Summary", "Detail"]


def test_option4_overconstrained_filter_empty_and_no_pdf_fetch(clean_store):
    enrich_calls = {"n": 0}

    async def _recorder(din_map, **kw):
        enrich_calls["n"] += 1

    # No product is a LOTION → empty survivor set.
    job = _run_job([{"metric": "dosage_form", "value": ["LOTION"]}], _recorder)
    assert job.status == "complete", job.error
    assert enrich_calls["n"] == 0, "empty survivor set must trigger NO PDF enrichment"
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert wb.sheetnames == ["Summary", "Detail"]
    assert len(job.summary_records) == 0, "over-constrained filter must yield an empty summary"
