"""Acceptance tests for the IQVIA quarter-over-quarter compare feature.

Covers the canonical compare path end-to-end against the committed fixture pair
(tests/fixtures/iqvia/diff/) AND small in-memory synthetic pairs built with the
fixture-script helpers (_ID / _OLD_METRICS / _NEW_METRICS):

  * direct compare_iqvia signal partitioning + anchored values,
  * auto-reorder (wrong slot order) and the workbook reorder banner,
  * same-period tie handling (order respected, single warning),
  * the BOTH-floor materiality gate (abs-only / pct-only / both / zero base),
  * API E2E via TestClient (200 xlsx, 400 bad inputs, 422 pivot-only).

Hermetic: no live network, no real IQVIA files, in-memory bytes only.  Periods
are pinned by the fixtures / synthetic columns, never date.today().
"""
import csv as _csv
import io as _io
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.enrichment.iqvia_diff import compare_iqvia, build_diff_workbook
from tests.scripts.build_iqvia_diff_fixture import _ID, _OLD_METRICS, _NEW_METRICS

_DIFF_DIR = Path(__file__).parent / "fixtures" / "iqvia" / "diff"
_OLD = _DIFF_DIR / "old_extract.csv"
_NEW = _DIFF_DIR / "new_extract.xlsx"

_SHEETS = ["Summary", "New Entrants", "Exits", "Material Moves"]


# ── helpers ───────────────────────────────────────────────────────────────────

def _require_fixtures():
    if not (_OLD.exists() and _NEW.exists()):
        pytest.skip(
            "IQVIA diff fixtures missing — run tests/scripts/build_iqvia_diff_fixture.py"
        )


def _old_csv_bytes(rows: list[list]) -> bytes:
    """Serialise rows (strings) as a CSV old-extract with _OLD_METRICS header."""
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(_ID + _OLD_METRICS)
    w.writerows(rows)
    return buf.getvalue().encode("utf-8")


def _new_xlsx_bytes(rows: list[list], metrics=None) -> bytes:
    """Serialise rows as a 'data'-sheet xlsx new-extract with _NEW_METRICS header."""
    cols = _ID + (metrics if metrics is not None else _NEW_METRICS)
    df = pd.DataFrame(rows, columns=cols)
    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        df.to_excel(xw, sheet_name="data", index=False)
    return buf.getvalue()


def _id_row(molecule, product, mfr, strength, latest_dollars, latest_units, *, old: bool):
    """One identity row with a zero earlier period and a populated latest period.

    Earlier-period (3 cols) all 0; latest period = (dollars, units, units) so the
    Ext Units = Units convention from the fixtures holds.  CSV (old) needs strings.
    """
    base = ["Drugstore", molecule, strength, "30 TAB", product, "TAB ORAL",
            mfr, mfr, "TAB", "ONTARIO"]
    vals = [0, 0, 0, latest_dollars, latest_units, latest_units]
    if old:
        return base + [str(v) for v in vals]
    return base + vals


@pytest.fixture(scope="module")
def diff():
    _require_fixtures()
    return compare_iqvia(_OLD.read_bytes(), _NEW.read_bytes())


# ── Direct compare_iqvia against the committed fixture pair ────────────────────

def test_signal_membership_and_no_leakage(diff):
    assert set(diff.entrants["Product"]) == {"OZEMPIC"}
    assert set(diff.exits["Product"]) == {"ZANTAC", "PEPCID"}
    assert set(diff.moves["Product"]) == {"LIPITOR"}


def test_glucophage_below_threshold_absent_everywhere(diff):
    for df in (diff.entrants, diff.exits, diff.moves):
        assert "GLUCOPHAGE" not in set(df["Product"])


def test_anchored_entrant_value(diff):
    r = diff.entrants[diff.entrants["Product"] == "OZEMPIC"].iloc[0]
    assert r["Combined Molecule"] == "SEMAGLUTIDE"
    assert int(r["Dollars"]) == 5_000_000
    assert int(r["Units"]) == 200_000


def test_anchored_exit_values(diff):
    z = diff.exits[diff.exits["Product"] == "ZANTAC"].iloc[0]
    assert int(z["Dollars"]) == 800_000
    assert int(z["Units"]) == 60_000
    p = diff.exits[diff.exits["Product"] == "PEPCID"].iloc[0]
    assert int(p["Dollars"]) == 300_000
    assert int(p["Units"]) == 20_000


def test_anchored_move_values(diff):
    r = diff.moves[diff.moves["Product"] == "LIPITOR"].iloc[0]
    assert int(r["Dollars Old"]) == 2_000_000
    assert int(r["Dollars New"]) == 2_500_000
    assert float(r["Dollars Δ%"]) == 25.0
    assert int(r["Units Old"]) == 100_000
    assert int(r["Units New"]) == 130_000
    assert float(r["Units Δ%"]) == 30.0


def test_periods_and_no_warnings(diff):
    assert diff.old_period == (2024, 6)
    assert diff.new_period == (2024, 12)
    assert diff.warnings == []


def test_correctly_ordered_fixture_not_reordered(diff):
    assert diff.reordered is False


def test_workbook_has_exactly_four_sheets(diff):
    wb = load_workbook(_io.BytesIO(build_diff_workbook(diff)), read_only=True)
    assert wb.sheetnames == _SHEETS


# ── AUTO-REORDER: pass fixtures in the WRONG slot order ───────────────────────

def test_reorder_detected_when_slots_swapped():
    _require_fixtures()
    # new file (latest 2024/12) in the OLD slot, old file (2024/06) in the NEW slot.
    d = compare_iqvia(_NEW.read_bytes(), _OLD.read_bytes())
    assert d.reordered is True
    # Auto-ordered older → newer, so resolved periods are unchanged.
    assert d.old_period == (2024, 6)
    assert d.new_period == (2024, 12)


def test_reorder_signals_unchanged():
    _require_fixtures()
    d = compare_iqvia(_NEW.read_bytes(), _OLD.read_bytes())
    assert set(d.entrants["Product"]) == {"OZEMPIC"}
    assert set(d.exits["Product"]) == {"ZANTAC", "PEPCID"}
    assert set(d.moves["Product"]) == {"LIPITOR"}


def test_reorder_workbook_summary_has_banner():
    _require_fixtures()
    d = compare_iqvia(_NEW.read_bytes(), _OLD.read_bytes())
    wb = load_workbook(_io.BytesIO(build_diff_workbook(d)), read_only=True)
    summary_cells = [
        str(row[0].value or "")
        for row in wb["Summary"].iter_rows(min_row=1, max_col=1)
    ]
    assert any("FILES REORDERED" in c for c in summary_cells), summary_cells


# ── SAME-DATE TIE: identical latest period → respect order, one warning ────────

def test_same_period_tie_respects_order_and_warns():
    # Both files' latest MAT period is the SAME (CSV old uses 2024/06, xlsx new
    # also resolves to 2024/06). Order must be respected, not reordered, exactly
    # one warning appended naming the shared period.
    old_metrics_tie = [
        "Dollars MAT 2023/06", "Units MAT 2023/06", "Ext Units MAT 2023/06",
        "Dollars MAT 2024/06", "Units MAT 2024/06", "Ext Units MAT 2024/06",
    ]
    new_metrics_tie = [
        "Dollars MAT 06/2023", "Units MAT 06/2023", "Ext Units MAT 06/2023",
        "Dollars MAT 06/2024", "Units MAT 06/2024", "Ext Units MAT 06/2024",
    ]
    old_rows = [["Drugstore", "MOLA", "10MG", "30 TAB", "BRANDA", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 "0", "0", "0", "1000000", "50000", "50000"]]
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(_ID + old_metrics_tie)
    w.writerows(old_rows)
    old_bytes = buf.getvalue().encode("utf-8")

    new_df = pd.DataFrame(
        [["Drugstore", "MOLA", "10MG", "30 TAB", "BRANDA", "TAB ORAL",
          "ACME", "ACME", "TAB", "ONTARIO", 0, 0, 0, 2000000, 60000, 60000]],
        columns=_ID + new_metrics_tie,
    )
    nbuf = _io.BytesIO()
    with pd.ExcelWriter(nbuf, engine="openpyxl") as xw:
        new_df.to_excel(xw, sheet_name="data", index=False)

    d = compare_iqvia(old_bytes, nbuf.getvalue())
    assert d.reordered is False
    assert d.old_period == (2024, 6) and d.new_period == (2024, 6)
    assert len(d.warnings) == 1
    assert "2024/06" in d.warnings[0]


# ── MATERIALITY GATE (both floors: abs AND pct) ───────────────────────────────

def test_move_clearing_abs_but_not_pct_is_not_a_move():
    # Huge base: +$120k clears the $100k abs floor, but on a $10M base that is
    # +1.2% < 10% pct floor → NOT a move. Units held flat (also sub-threshold).
    old_bytes = _old_csv_bytes([
        _id_row("MOLB", "BRANDB", "ACME", "10MG", 10_000_000, 500_000, old=True),
    ])
    new_bytes = _new_xlsx_bytes([
        _id_row("MOLB", "BRANDB", "ACME", "10MG", 10_120_000, 500_000, old=False),
    ])
    d = compare_iqvia(old_bytes, new_bytes)
    assert d.moves.empty
    assert d.entrants.empty and d.exits.empty


def test_move_clearing_pct_but_not_abs_is_not_a_move():
    # Tiny base: +$50k on a $100k base is +50% (clears pct) but $50k < $100k abs
    # floor → NOT a move. Units +250 on 1,000 base: +25% pct but 250 < 1,000 abs.
    old_bytes = _old_csv_bytes([
        _id_row("MOLC", "BRANDC", "ACME", "10MG", 100_000, 1_000, old=True),
    ])
    new_bytes = _new_xlsx_bytes([
        _id_row("MOLC", "BRANDC", "ACME", "10MG", 150_000, 1_250, old=False),
    ])
    d = compare_iqvia(old_bytes, new_bytes)
    assert d.moves.empty


def test_move_clearing_both_floors_is_a_move():
    # +$2M on $2M base = +100% (clears $100k abs AND 10% pct) → IS a move.
    old_bytes = _old_csv_bytes([
        _id_row("MOLD", "BRANDD", "ACME", "10MG", 2_000_000, 100_000, old=True),
    ])
    new_bytes = _new_xlsx_bytes([
        _id_row("MOLD", "BRANDD", "ACME", "10MG", 4_000_000, 200_000, old=False),
    ])
    d = compare_iqvia(old_bytes, new_bytes)
    assert set(d.moves["Product"]) == {"BRANDD"}
    r = d.moves.iloc[0]
    assert int(r["Dollars Δ"]) == 2_000_000
    assert float(r["Dollars Δ%"]) == 100.0


def test_move_with_zero_base_has_blank_pct():
    # Present in old via Units only (Dollars 0), so it is a move (not entrant).
    # Dollars Δ% must be blank (NaN) — never fabricated against a 0 base.
    old_rows = [["Drugstore", "MOLE", "10MG", "30 TAB", "BRANDE", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 "0", "0", "0", "0", "5000", "5000"]]
    new_rows = [["Drugstore", "MOLE", "10MG", "30 TAB", "BRANDE", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 0, 0, 0, 500000, 9000, 9000]]
    old_bytes = _old_csv_bytes(old_rows)
    new_bytes = _new_xlsx_bytes(new_rows)
    d = compare_iqvia(old_bytes, new_bytes)
    r = d.moves[d.moves["Product"] == "BRANDE"].iloc[0]
    assert int(r["Dollars Old"]) == 0
    assert int(r["Dollars New"]) == 500_000
    assert pd.isna(r["Dollars Δ%"])
    assert float(r["Units Δ%"]) == 80.0


# ── API E2E via TestClient ────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def client():
    from fastapi.testclient import TestClient
    from app.main import app
    return TestClient(app)


def test_api_valid_pair_returns_xlsx(client):
    _require_fixtures()
    resp = client.post(
        "/api/iqvia/compare",
        files={
            "old_file": ("old_extract.csv", _OLD.read_bytes(), "text/csv"),
            "new_file": ("new_extract.xlsx", _NEW.read_bytes(),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml.sheet" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    wb = load_workbook(_io.BytesIO(resp.content), read_only=True)
    assert wb.sheetnames == _SHEETS


def test_api_wrong_content_type_is_400(client):
    _require_fixtures()
    resp = client.post(
        "/api/iqvia/compare",
        content=b"not multipart",
        headers={"content-type": "application/json"},
    )
    assert resp.status_code == 400


def test_api_missing_new_file_is_400(client):
    _require_fixtures()
    resp = client.post(
        "/api/iqvia/compare",
        files={"old_file": ("old_extract.csv", _OLD.read_bytes(), "text/csv")},
    )
    assert resp.status_code == 400


def test_api_txt_extension_is_400(client):
    resp = client.post(
        "/api/iqvia/compare",
        files={
            "old_file": ("old.txt", b"whatever", "text/plain"),
            "new_file": ("new.txt", b"whatever", "text/plain"),
        },
    )
    assert resp.status_code == 400


def test_api_pivot_only_slot_is_422(client):
    # A "data" sheet with NO metric columns → compare_iqvia raises ValueError → 422.
    pivot_df = pd.DataFrame({"Combined Molecule": ["X"], "Product": ["Y"],
                             "Manufacturer": ["Z"], "Strength": ["10MG"]})
    pbuf = _io.BytesIO()
    with pd.ExcelWriter(pbuf, engine="openpyxl") as xw:
        pivot_df.to_excel(xw, sheet_name="data", index=False)
    pivot_bytes = pbuf.getvalue()
    _require_fixtures()
    resp = client.post(
        "/api/iqvia/compare",
        files={
            "old_file": ("pivot.xlsx", pivot_bytes,
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            "new_file": ("new_extract.xlsx", _NEW.read_bytes(),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
    )
    assert resp.status_code == 422
