"""Acceptance E2E for OPTION 3 — the full no-PDF universe job.

Drives ``run_universe_full_job`` with a real-data UniverseBundle (committed extract
slice + register fixture) and asserts the contract documented in CLAUDE.md and
universe.py:

  * job completes; the workbook opens with EXACTLY the three option-3 sheets;
  * the Read-Me carries the disclaimer ("PDF data is OMITTED") and the
    NOT-evaluated caution;
  * data-protection columns are POPULATED for matched DINs while NOC / patent
    columns stay blank by design;
  * the full path NEVER fetches PM PDFs (enrich_labeling_batch_fast call count 0);
  * a terminal ``complete`` SSE event with a download_url is emitted;
  * the dashboard snapshot is capped at 2000 rows; the recon snapshot is present
    only when an IQVIA frame is supplied;
  * /api/universe/status exposes the documented status dict keys.

Hermetic: no network — get_universe is stubbed, GSUR is empty, the universe cache
is reset around every test.
"""
from __future__ import annotations

import asyncio
import io
import json
from pathlib import Path

import openpyxl
import pytest
from bs4 import BeautifulSoup

import app.universe_job as UJ
from app.enrichment import universe as U
from app.enrichment.data_protection import _find_active_table, _parse_data_protection_table
from app.enrichment.universe import UniverseBundle
from app.jobs import JobState
from app.models import DrugRecord

_DP_FIX = Path(__file__).parent / "fixtures" / "data_protection"
_UNI_EXTRACT = Path(__file__).parent / "fixtures" / "universe" / "extract"

pytestmark = pytest.mark.skipif(
    not (_DP_FIX / "register_active_sample.html").exists()
    or not (_DP_FIX / "dp_join_products.json").exists()
    or not (_UNI_EXTRACT / "drug.txt").exists(),
    reason="universe / dp fixtures missing — run the build_*_fixture.py scripts",
)

_EXPECTED_SHEETS = ["⚠ Read Me", "Full Universe (no PDF)", "Generic Submissions"]
# These columns are written verbatim (no display-name remap) by the workbook.
_DP_COL = "dp_6yr_no_file_date"
_NOC_PATENT_BLANK_COLS = (
    "noc_date", "submission_class", "noc_submission_type",
    "noc_therapeutic_class", "reason_for_supplement",
)


def _dp_table() -> list[dict]:
    soup = BeautifulSoup(
        (_DP_FIX / "register_active_sample.html").read_text(encoding="utf-8"),
        "html.parser",
    )
    return _parse_data_protection_table(_find_active_table(soup))


def _join() -> dict:
    return json.loads((_DP_FIX / "dp_join_products.json").read_text(encoding="utf-8"))


@pytest.fixture(autouse=True)
def _reset_universe_cache():
    """Never let an in-process universe bundle leak across tests / workers."""
    U._CACHE["bundle"] = None
    yield
    U._CACHE["bundle"] = None


@pytest.fixture()
def clean_store(tmp_path):
    """Point the enrichment store at an empty DB so patent/labeling cols are blank."""
    import app.enrichment.store as store_mod

    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    return store_mod


def _real_bundle() -> UniverseBundle:
    """A bundle built from the committed real extract slice + register fixture."""
    recs = U.load_dpd_universe_records(_UNI_EXTRACT)
    return UniverseBundle(recs, [], dp_table=_dp_table())


def _dpd_from_join(p: dict, drug_code: int) -> DrugRecord:
    """A DPD DrugRecord carrying a real (din, ingredient, company) that the active
    Register matches by ingredient + manufacturer (the dp join fixture identities)."""
    return DrugRecord(
        source="DPD", din=p["din"], brand_name="BRAND",
        ingredient=p["ingredient"], company=p["company"],
        all_ingredients=[p["ingredient"]], strength="100 mg",
        dosage_form=p.get("dosage_form") or "TABLET",
        status=p.get("status") or "MARKETED",
        source_specific={"drug_code": drug_code},
    )


def _matched_bundle() -> UniverseBundle:
    """Bundle whose DPD identities are the register-matched products (so dp populates)
    plus some unmatched generics (so blanks are exercised)."""
    join = _join()
    matched = join["matched"][:20]
    blanks = join["blank"][:8]
    recs = ([_dpd_from_join(p, 1000 + i) for i, p in enumerate(matched)]
            + [_dpd_from_join(p, 5000 + i) for i, p in enumerate(blanks)])
    return UniverseBundle(recs, [], dp_table=_dp_table())


def _run_full(job: JobState, recorder, iqvia_df=None, bundle=None):
    """Drive run_universe_full_job with a stubbed universe + a PDF-enrich recorder.

    No IQVIA by default (so no recon sheet); pass iqvia_df to exercise that branch.
    """
    import unittest.mock as M

    the_bundle = bundle if bundle is not None else _real_bundle()

    async def _fake_get_universe(force_refresh=False):
        return the_bundle

    with M.patch.object(UJ, "get_universe", _fake_get_universe), \
         M.patch.object(UJ, "enrich_labeling_batch_fast", recorder), \
         M.patch.object(UJ, "_resolve_iqvia", lambda job: iqvia_df):
        asyncio.run(UJ.run_universe_full_job(job))
    return the_bundle


def _enrich_recorder():
    calls = {"n": 0}

    async def _rec(*a, **kw):  # pragma: no cover - asserted never called
        calls["n"] += 1

    return calls, _rec


def _sheet_rows(job: JobState, sheet: str):
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return wb, headers, rows


# ════════════════════════════════════════════════════════════════════════════
# Full universe E2E: sheets, disclaimer, dp populated, NOC/patent blank
# ════════════════════════════════════════════════════════════════════════════

def test_full_job_completes_with_exact_three_sheets(clean_store):
    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-sheets", query="", field="ingredient")
    _run_full(job, rec)

    assert job.status == "complete", job.error
    assert job.result_path and Path(job.result_path).exists()
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert wb.sheetnames == _EXPECTED_SHEETS


def test_full_job_readme_has_disclaimer_and_caution(clean_store):
    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-readme", query="", field="ingredient")
    _run_full(job, rec)

    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    ws = wb["⚠ Read Me"]
    text = "\n".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value is not None
    )
    # Disclaimer body (rendered verbatim from UNIVERSE_DISCLAIMER_LINES).
    assert "PDF) data is OMITTED" in text
    assert "are NOT included here" in text or "NOT included here" in text
    # The bold NOT-evaluated caution string is rendered verbatim.
    assert U.UNIVERSE_NOT_EVALUATED_CAUTION in text
    assert "NOC and patent data are NOT evaluated" in text


def test_full_job_dp_populated_but_noc_patent_blank(clean_store):
    """Bundle of register-matched products → dp columns populate; NOC/patent stay
    blank by design on the no-PDF universe."""
    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-dp", query="", field="ingredient")
    _run_full(job, rec, bundle=_matched_bundle())

    _wb, headers, rows = _sheet_rows(job, "Full Universe (no PDF)")

    def col(name: str) -> int:
        assert name in headers, f"{name!r} not in sheet1 headers: {headers}"
        return headers.index(name)

    din_i = col("DIN")
    dp_i = col(_DP_COL)

    dp_by_din = {}
    for r in rows[1:]:
        din = str(r[din_i]).strip() if r[din_i] is not None else ""
        val = r[dp_i]
        dp_by_din[din] = "" if val is None else str(val).strip()

    populated = sum(1 for v in dp_by_din.values() if v)
    assert populated > 0, "data-protection column is all-blank — dp not wired into option 3"

    # Each register-matched DIN carries its expected six-year no-file date; each
    # generic (blank) DIN stays blank.
    join = _join()
    checked = 0
    for m in join["matched"][:20]:
        assert dp_by_din.get(m["din"], "") == m["expected_no_file_date"], m["din"]
        checked += 1
    for p in join["blank"][:8]:
        assert dp_by_din.get(p["din"], "") == "", p["din"]
    assert checked > 0

    # NOC columns appear but are BLANK by design on the no-PDF universe.
    for name in _NOC_PATENT_BLANK_COLS:
        if name not in headers:
            continue
        i = headers.index(name)
        nonblank = [r[i] for r in rows[1:] if r[i] not in (None, "")]
        assert not nonblank, f"column {name!r} should be blank by design, got {nonblank[:3]}"
    # Patent DETAIL columns (number / filing / grant / expiry) that survived pruning
    # must be blank by design.  patent_count legitimately holds 0 (a count, not a
    # detail), so it is excluded from the blank check.
    for i, h in enumerate(headers):
        hl = h.lower()
        if "patent" not in hl or hl == "patent_count":
            continue
        nonblank = [r[i] for r in rows[1:] if r[i] not in (None, "")]
        assert not nonblank, f"patent column {h!r} should be blank, got {nonblank[:3]}"


def test_full_job_never_fetches_pdfs(clean_store):
    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-nopdf", query="", field="ingredient")
    _run_full(job, rec)
    assert job.status == "complete", job.error
    assert calls["n"] == 0, "the full no-PDF path must NEVER call enrich_labeling_batch_fast"


def test_full_job_emits_complete_event_with_download_url(clean_store):
    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-evt", query="", field="ingredient")
    _run_full(job, rec)

    completes = [e for e in job.events if e.get("status") == "complete"]
    assert len(completes) == 1, "exactly one terminal complete event expected"
    assert completes[-1]["download_url"] == f"/export/result/{job.job_id}"


# ════════════════════════════════════════════════════════════════════════════
# Dashboard snapshot cap + recon snapshot presence
# ════════════════════════════════════════════════════════════════════════════

def test_dashboard_snapshot_capped_at_2000(clean_store):
    """Even a large universe must cap the in-page dashboard preview at 2000 rows."""
    base = U.load_dpd_universe_records(_UNI_EXTRACT)
    template = base[0]
    big: list[DrugRecord] = []
    for i in range(2100):
        big.append(DrugRecord(
            source="DPD",
            din=f"90{i:06d}",
            brand_name=template.brand_name or "BRAND",
            ingredient=template.ingredient or "TESTOL",
            company=template.company or "ACME",
            all_ingredients=template.all_ingredients or ["TESTOL"],
            strength=template.strength or "100 mg",
            dosage_form=template.dosage_form or "TABLET",
            status="MARKETED",
            source_specific={"drug_code": 800000 + i},
        ))
    bundle = UniverseBundle(big, [], dp_table=_dp_table())

    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-cap", query="", field="ingredient")
    _run_full(job, rec, bundle=bundle)

    assert job.status == "complete", job.error
    assert len(job.sheet1_records) == 2000, len(job.sheet1_records)
    # The downloaded file still carries every row.
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    data_rows = wb["Full Universe (no PDF)"].max_row - 1  # minus header
    assert data_rows >= 2100, data_rows


def test_recon_snapshot_absent_without_iqvia(clean_store):
    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-norecon", query="", field="ingredient")
    _run_full(job, rec)
    # No IQVIA frame was resolved → no recon snapshot, three sheets only.
    assert job.recon_records == []
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert wb.sheetnames == _EXPECTED_SHEETS


def test_recon_sheet_present_with_iqvia(clean_store):
    """An IQVIA frame matching a universe DIN adds a reconciliation sheet + snapshot."""
    import pandas as pd

    recs = U.load_dpd_universe_records(_UNI_EXTRACT)
    # Pick a real product and key a collapsed-IQVIA-shaped row to it so the matcher
    # stamps at least one row → non-empty reconciliation frame.
    target = next(r for r in recs if r.din and r.ingredient and r.company)
    molecule = str(target.ingredient).split(" ")[0]
    iq = pd.DataFrame([{
        "Combined Molecule": molecule,
        "Product": target.brand_name or "BRAND",
        "Manufacturer": target.company,
        "Strength": "100 mg",
        "Dollars MAT 12/2025": 1_000_000,
        "Units MAT 12/2025": 5_000,
        "Ext Units MAT 12/2025": 5_000,
    }])

    calls, rec = _enrich_recorder()
    job = JobState(job_id="t-opt3-recon", query="", field="ingredient")
    _run_full(job, rec, iqvia_df=iq)

    assert job.status == "complete", job.error
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    recon_sheets = [s for s in wb.sheetnames if "recon" in s.lower()]
    assert recon_sheets, f"reconciliation sheet expected, got {wb.sheetnames}"
    assert job.recon_records, "recon snapshot should be populated when IQVIA supplied"


# ════════════════════════════════════════════════════════════════════════════
# /api/universe/status documented contract
# ════════════════════════════════════════════════════════════════════════════

def test_status_endpoint_keys_after_build(clean_store):
    """After a cached build, /api/universe/status returns the documented keys."""
    from fastapi.testclient import TestClient
    from app.main import app

    U._CACHE["bundle"] = None
    with TestClient(app) as client:
        empty = client.get("/api/universe/status").json()
        assert empty["cached"] is False
        assert empty["ttl_seconds"] == U.UNIVERSE_TTL

        U._CACHE["bundle"] = _real_bundle()
        full = client.get("/api/universe/status").json()

    for key in (
        "cached", "ttl_seconds", "age_seconds", "expires_in_seconds",
        "fresh", "dpd_records", "gsur_records",
    ):
        assert key in full, f"missing status key {key!r}: {full}"
    assert full["cached"] is True
    assert full["fresh"] is True
    assert full["dpd_records"] > 0
