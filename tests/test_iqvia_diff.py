"""Regression test for the IQVIA quarter-over-quarter comparison.

Anchored to the committed fixture pair in tests/fixtures/iqvia/diff/ (regenerate
with tests/scripts/build_iqvia_diff_fixture.py).  The pair is fully synthetic and
deliberately mismatched across the two files — CSV vs xlsx, YYYY/MM vs MM/YYYY
date order, a decoy Pivot sheet — so this exercises the whole canonical path:
generalized parse → collapse → latest-MAT resolution → normalised-identity diff.

Every expected number below is hand-computed from the fixture rows; see the
builder script's docstring for the per-product rationale.
"""
from pathlib import Path

import pytest
import pandas as pd

from app.enrichment.iqvia_diff import compare_iqvia, build_diff_workbook

_DIFF_DIR = Path(__file__).parent / "fixtures" / "iqvia" / "diff"
_OLD = _DIFF_DIR / "old_extract.csv"
_NEW = _DIFF_DIR / "new_extract.xlsx"


@pytest.fixture(scope="module")
def diff():
    if not (_OLD.exists() and _NEW.exists()):
        pytest.skip(
            "IQVIA diff fixtures missing — run tests/scripts/build_iqvia_diff_fixture.py"
        )
    return compare_iqvia(_OLD.read_bytes(), _NEW.read_bytes())


def _by_product(df: pd.DataFrame, name: str) -> pd.Series:
    rows = df[df["Product"] == name]
    assert len(rows) == 1, f"expected exactly one {name!r} row, got {len(rows)}"
    return rows.iloc[0]


# ── Latest-MAT resolution (per file, independent) ─────────────────────────────

def test_latest_periods_resolved_per_file(diff):
    # Old CSV latest = 2024/06 (YYYY/MM); new xlsx latest = 12/2024 (MM/YYYY).
    assert diff.old_period == (2024, 6)
    assert diff.new_period == (2024, 12)
    # New is strictly later → no swap/equal-period warnings.
    assert diff.warnings == []


# ── Signal partitioning: exact membership, no leakage ─────────────────────────

def test_entrants_exits_moves_membership(diff):
    assert set(diff.entrants["Product"]) == {"OZEMPIC"}
    assert set(diff.exits["Product"]) == {"ZANTAC", "PEPCID"}
    assert set(diff.moves["Product"]) == {"LIPITOR"}


def test_below_threshold_row_absent_everywhere(diff):
    # GLUCOPHAGE moved +$40k (<$100k) and +500u (<1,000): below both floors.
    for df in (diff.entrants, diff.exits, diff.moves):
        assert "GLUCOPHAGE" not in set(df["Product"]), "below-threshold row leaked into output"


# ── ENTRANT ───────────────────────────────────────────────────────────────────

def test_entrant_values(diff):
    r = _by_product(diff.entrants, "OZEMPIC")
    assert r["Combined Molecule"] == "SEMAGLUTIDE"
    assert int(r["Dollars"]) == 5_000_000
    assert int(r["Units"]) == 200_000
    assert int(r["Ext Units"]) == 200_000


# ── EXITS ──────────────────────────────────────────────────────────────────────

def test_exit_absent_from_new(diff):
    r = _by_product(diff.exits, "ZANTAC")           # gone from the new file entirely
    assert int(r["Dollars"]) == 800_000
    assert int(r["Units"]) == 60_000


def test_exit_zero_latest_in_new(diff):
    # PEPCID's row exists in new but its latest-MAT cells are "-" (0) → still an exit.
    r = _by_product(diff.exits, "PEPCID")
    assert int(r["Dollars"]) == 300_000           # the OLD value is reported


def test_exits_sorted_by_dollars_desc(diff):
    assert list(diff.exits["Product"]) == ["ZANTAC", "PEPCID"]


def test_signal_rows_group_a_products_strengths_together():
    # A multi-strength exit must keep all of its strengths ADJACENT (and never drop
    # the small one) — a plain Dollars-desc sort scatters them, making the small
    # strength read as "missing".  Build a pair where one product (BIGDRUG) exits at
    # two very different strengths ($90M @ 80MG, $0.1M @ 160MG) alongside a mid-size
    # exit (MIDDRUG $5M): the product groups rank largest-first, but BIGDRUG's two
    # strengths stay side by side rather than landing ~1 row apart around MIDDRUG.
    from tests.scripts.build_iqvia_diff_fixture import _ID, _OLD_METRICS, _NEW_METRICS
    import csv as _csv, io as _io

    # _ID order: Channel, Combined Molecule, Strength, Pack, Product, Form 3,
    # Manufacturer, Corporation, Product Form, Province. Metric cols carry TWO
    # periods (6 values): earlier period 0, latest period = the test value.
    def _row(brand, mfr, strength, dollars):
        return ["Drugstore", brand, strength, "30 TAB", brand, "TAB ORAL",
                mfr, mfr, "TAB", "ONTARIO",
                "0", "0", "0", str(dollars), str(dollars // 10), str(dollars // 10)]

    old_rows = [
        _row("BIGDRUG", "ACME", "80MG", 90_000_000),
        _row("BIGDRUG", "ACME", "160MG", 100_000),
        _row("MIDDRUG", "BETA", "50MG", 5_000_000),
    ]
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(_ID + _OLD_METRICS)
    w.writerows(old_rows)
    old_bytes = buf.getvalue().encode("utf-8")
    # New file: everything in old exits — one inert, unrelated row so the latest-MAT
    # period still resolves to a later quarter than old (6 metric values).
    new_df = pd.DataFrame(
        [["Drugstore", "ZZOTHER", "1MG", "30 TAB", "ZZOTHER", "TAB ORAL",
          "GAMMA", "GAMMA", "TAB", "ONTARIO", 0, 0, 0, 1, 1, 1]],
        columns=_ID + _NEW_METRICS,
    )
    nbuf = _io.BytesIO()
    with pd.ExcelWriter(nbuf, engine="openpyxl") as xw:
        new_df.to_excel(xw, sheet_name="data", index=False)
    d = compare_iqvia(old_bytes, nbuf.getvalue())

    products = list(d.exits["Product"])
    # Both BIGDRUG strengths present (small one not dropped) ...
    assert products.count("BIGDRUG") == 2
    # ... and adjacent (group held together), with the larger product group first.
    first = products.index("BIGDRUG")
    assert products[first:first + 2] == ["BIGDRUG", "BIGDRUG"]
    assert products[0] == "BIGDRUG"  # $90.1M group ranks ahead of MIDDRUG's $5M
    big = d.exits[d.exits["Product"] == "BIGDRUG"]
    assert list(big["Strength"]) == ["80MG", "160MG"]  # within group: Dollars desc


# ── MATERIAL MOVE (+ identity normalisation) ──────────────────────────────────

def test_material_move_values(diff):
    r = _by_product(diff.moves, "LIPITOR")
    # Old split across two channel rows (1.2M + 0.8M) must have collapsed to 2.0M.
    assert int(r["Dollars Old"]) == 2_000_000
    assert int(r["Dollars New"]) == 2_500_000
    assert int(r["Dollars Δ"]) == 500_000
    assert float(r["Dollars Δ%"]) == 25.0
    assert int(r["Units Old"]) == 100_000
    assert int(r["Units New"]) == 130_000
    assert int(r["Units Δ"]) == 30_000
    assert float(r["Units Δ%"]) == 30.0


def test_material_move_survives_formatting_jitter(diff):
    # Old "PFIZER CANADA ULC"/"LIPITOR 20MG TABLETS" vs new "PFIZER"/"LIPITOR" must
    # fold to ONE identity — never a phantom exit + entrant.
    assert "LIPITOR" not in set(diff.entrants["Product"])
    assert "LIPITOR" not in set(diff.exits["Product"])


# ── Manufacturer identity normalisation (merge vs split) ──────────────────────

def test_norm_company_identity_merges_legal_and_geographic_jitter():
    # Legal-suffix and geographic ("Canada", "Health", "Consumer Healthcare",
    # "Division") jitter for the SAME firm must fold to one identity key.
    from app.enrichment.iqvia_diff import _norm_company_identity as n
    for a, b in [
        ("PFIZER CANADA ULC", "PFIZER"),
        ("Bayer Inc", "Bayer Incorporated"),
        ("Sandoz Canada Inc", "Sandoz"),
        ("GSK Consumer Healthcare", "GSK"),
        ("Merck Canada Inc", "Merck"),
        ("Teva Canada Limited", "Teva Canada Ltd"),
        ("Novartis Pharmaceuticals Canada Inc", "Novartis Pharmaceuticals Inc"),
    ]:
        assert n(a) == n(b) != "", f"{a!r} should equal {b!r}: {n(a)!r} vs {n(b)!r}"


def test_norm_company_identity_keeps_firm_distinguishing_words():
    # A descriptive word that distinguishes two REAL firms (Pharma vs Therapeutics,
    # Pharma vs Labs) must be KEPT — the diff has no fuzzy floor, so stripping it
    # would silently merge two companies and erase a real exit + entrant.
    from app.enrichment.iqvia_diff import _norm_company_identity as n
    for a, b in [
        ("Acme Pharma Inc", "Acme Therapeutics Inc"),
        ("Acme Pharmaceuticals", "Acme Therapeutics"),
        ("Sun Pharma", "Sun Labs"),
        ("Beigene Switz GmbH", "Beone Med I GmbH"),
    ]:
        assert n(a) != n(b), f"{a!r} must differ from {b!r}: both -> {n(a)!r}"


def test_distinct_firms_are_exit_plus_entrant_not_silent_merge():
    # End-to-end: same molecule/brand/strength sold by two DIFFERENT manufacturers
    # that differ only by a firm-distinguishing word must surface as a visible
    # exit + entrant pair — never collapse into one (possibly dropped) move.
    from tests.scripts.build_iqvia_diff_fixture import _ID, _OLD_METRICS, _NEW_METRICS
    import csv as _csv, io as _io

    def _row(mfr, dollars, metrics_cols):
        base = ["Drugstore", "MOLZ", "50MG", "30 TAB", "BRANDZ", "TAB ORAL",
                mfr, mfr, "TAB", "ONTARIO"]
        if metrics_cols is _OLD_METRICS:
            return base + ["0", "0", "0", str(dollars), str(dollars // 2), str(dollars // 2)]
        return base + [0, 0, 0, dollars, dollars // 2, dollars // 2]

    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(_ID + _OLD_METRICS)
    w.writerow(_row("Acme Pharma Inc", 100_000_000, _OLD_METRICS))
    old_bytes = buf.getvalue().encode("utf-8")
    new_df = pd.DataFrame([_row("Acme Therapeutics Inc", 103_000_000, _NEW_METRICS)],
                          columns=_ID + _NEW_METRICS)
    nbuf = _io.BytesIO()
    with pd.ExcelWriter(nbuf, engine="openpyxl") as xw:
        new_df.to_excel(xw, sheet_name="data", index=False)
    d = compare_iqvia(old_bytes, nbuf.getvalue())

    assert list(d.exits["Manufacturer"]) == ["Acme Pharma Inc"]
    assert list(d.entrants["Manufacturer"]) == ["Acme Therapeutics Inc"]
    assert d.moves.empty


# ── Workbook shape ─────────────────────────────────────────────────────────────


def test_workbook_metric_headers_carry_mat_period(diff):
    # The diff reports a single latest-MAT period per file; the metric headers must
    # name it so a figure is never mistaken for an all-time total. Entrants carry
    # the new-file period, exits the old-file period, moves carry both; Δ/Δ% stay
    # bare (they span both periods).
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(build_diff_workbook(diff)), read_only=True)

    def _headers(sheet):
        return [str(c.value) for c in next(wb[sheet].iter_rows(min_row=1, max_row=1))]

    old_lbl = f"{diff.old_period[0]}/{diff.old_period[1]:02d}"   # 2024/06
    new_lbl = f"{diff.new_period[0]}/{diff.new_period[1]:02d}"   # 2024/12

    ent, ex, mv = _headers("New Entrants"), _headers("Exits"), _headers("Material Moves")
    for metric in ("Dollars", "Units", "Ext Units"):
        assert f"{metric} (MAT {new_lbl})" in ent
        assert f"{metric} (MAT {old_lbl})" in ex
        assert f"{metric} Old (MAT {old_lbl})" in mv
        assert f"{metric} New (MAT {new_lbl})" in mv
        assert f"{metric} Δ" in mv and f"{metric} Δ%" in mv   # deltas stay bare
    # Identity columns are never relabelled.
    assert ent[:4] == ["Combined Molecule", "Product", "Manufacturer", "Strength"]

def test_workbook_has_four_sheets(diff):
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(build_diff_workbook(diff)), read_only=True)
    assert wb.sheetnames == ["Summary", "New Entrants", "Exits", "Material Moves"]


def test_summary_has_no_extract_date_fields(diff):
    # The per-file MAT periods are reporting periods, not extract/pull dates, so
    # they were removed from the Summary sheet. Assert they cannot silently return:
    # no Summary "Metric" cell may reference an extract date / MAT period.
    import io
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(build_diff_workbook(diff)), read_only=True)
    metrics = [
        str(row[0].value or "")
        for row in wb["Summary"].iter_rows(min_row=2, max_col=1)
    ]
    assert not any("extract" in m.lower() or "mat period" in m.lower() for m in metrics), metrics


def test_no_invented_values_pct_blank_when_no_base():
    # A product present in old with zero Dollars but positive Units (so it is a
    # move, not an entrant) must leave Dollars Δ% blank rather than fabricate a %.
    # Build a minimal in-memory pair to assert the contract directly.
    from tests.scripts.build_iqvia_diff_fixture import _ID, _OLD_METRICS, _NEW_METRICS
    import csv as _csv, io as _io

    old_rows = [["Drugstore", "DRUGX", "10MG", "30 TAB", "BRANDX", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 "0", "5,000", "5,000", "0", "5,000", "5,000"]]
    new_rows = [["Drugstore", "DRUGX", "10MG", "30 TAB", "BRANDX", "TAB ORAL",
                 "ACME", "ACME", "TAB", "ONTARIO",
                 0, 5000, 5000, 500000, 9000, 9000]]
    buf = _io.StringIO()
    w = _csv.writer(buf)
    w.writerow(_ID + _OLD_METRICS)
    w.writerows(old_rows)
    old_bytes = buf.getvalue().encode("utf-8")
    new_df = pd.DataFrame(new_rows, columns=_ID + _NEW_METRICS)
    nbuf = _io.BytesIO()
    with pd.ExcelWriter(nbuf, engine="openpyxl") as xw:
        new_df.to_excel(xw, sheet_name="data", index=False)
    d = compare_iqvia(old_bytes, nbuf.getvalue())
    r = d.moves[d.moves["Product"] == "BRANDX"].iloc[0]
    assert int(r["Dollars Old"]) == 0
    assert int(r["Dollars New"]) == 500_000
    assert pd.isna(r["Dollars Δ%"])              # 0 base → blank, never faked
    assert float(r["Units Δ%"]) == 80.0          # 5,000 → 9,000
