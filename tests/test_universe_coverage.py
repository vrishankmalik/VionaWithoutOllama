"""Comprehensive offline regression suite for the Full-universe feature (options 3 & 4).

This complements the build's own synthetic tests in ``tests/test_universe.py`` by
anchoring against REAL Health Canada + IQVIA rows, trimmed into committed fixtures
so the whole suite runs in the default offline ``make test`` pass:

    tests/fixtures/universe/extract/*.txt   — a curated DIN cohort sliced verbatim
                                              from the live allfiles.zip
    tests/fixtures/universe/iqvia_slice.xlsx — the real IQVIA province/channel rows
                                              for the cohort molecules

Regenerate with ``python -m tests.scripts.build_universe_fixture`` (needs network).
Every numeric anchor below is reproduced by that generator's MANIFEST and matches
the hand-verified IQVIA anchors already pinned in ``tests/test_iqvia.py`` (SANIS
PROGESTERONE 100MG = 218591 units / $21,215,081; AURO = 233159 / $13,005,865; the
PRO DOC generic-label aggregation 37942 + 48 = 37990).

Design decision recorded with the maintainer: the assembled universe is a DPD-only
sheet — ``build_universe_response`` carries NO live NOC/patent/data-protection join
(no NOC bulk extract exists; a per-DIN join would cost the hours the no-PDF universe
exists to avoid).  So at the universe level NOC/patent/DP columns are uniformly
blank (cite-or-blank), and the "NOC-anchored populated vs grandfathered blank"
distinction is exercised at the ``build_sheet1(include_dpd_only=True)`` primitive
(section B2), which is the real seam the universe reuses.
"""
from __future__ import annotations

import asyncio
import io
import zipfile
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from app.enrichment import screen
from app.enrichment import universe as U
from app.enrichment.iqvia import collapse_iqvia, match_iqvia_to_sheet1, parse_iqvia
from app.enrichment.universe import (
    build_universe_response,
    build_universe_sheet1,
    build_universe_sheet2,
    UniverseBundle,
)
from app.enrichment.workbook import build_sheet1
from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult

_FIX = Path(__file__).parent / "fixtures" / "universe"
_EXTRACT = _FIX / "extract"
_IQVIA_SLICE = _FIX / "iqvia_slice.xlsx"

pytestmark = pytest.mark.skipif(
    not (_EXTRACT / "drug.txt").exists() or not _IQVIA_SLICE.exists(),
    reason="universe fixtures missing — run python -m tests.scripts.build_universe_fixture",
)

# ── Real cohort anchors (reproduced by the fixture generator MANIFEST) ─────────
_EXPECT_RECORD_COUNT = 9
_DOLLARS = "Dollars MAT 12/2025"
_UNITS = "Units MAT 12/2025"
_EXT = "Ext Units MAT 12/2025"

# DIN → (brand, company, latest-MAT dollars, latest-MAT units, confidence)
_ANCHORS = {
    "02516187": ("PROGESTERONE", "SANIS HEALTH INC", 21215081, 218591, "exact"),
    "02493578": ("AURO-PROGESTERONE", "AURO PHARMA INC", 13005865, 233159, "exact"),
    "02314908": ("PRO-METFORMIN", "PRO DOC LIMITEE", 475347, 37990, "exact"),
    "02380196": ("JAMP METFORMIN", "JAMP PHARMA CORPORATION", 5603513, 507908, "high"),
    "02284065": ("PMS-AMLODIPINE", "PHARMASCIENCE INC", 1544348, 33489, "exact"),
    "02522519": ("PRZ-AMLODIPINE", "PHARMARIS CANADA INC", 55080, 760, "exact"),
    "02099233": ("GLUCOPHAGE", "SANOFI-AVENTIS CANADA INC", 1744425, 37795, "exact"),
}
_GRANDFATHERED_DIN = "00015741"  # TAPAZOLE — no IQVIA presence → confidence 'none'


# ── Offline loaders (neutralise the live allfiles download) ────────────────────
def _load_real_records() -> list[DrugRecord]:
    """Parse the trimmed real extract directly.

    load_dpd_universe_records() is a pure, no-network parser that honours its
    cache_dir argument (download is a separate step done by get_universe), so we read
    the fixture dir with no stubbing — see test_load_records_pure_parse_honors_cache_dir.
    """
    return U.load_dpd_universe_records(_EXTRACT)


@pytest.fixture(scope="module")
def records() -> list[DrugRecord]:
    return _load_real_records()


@pytest.fixture(scope="module")
def iqvia_collapsed() -> pd.DataFrame:
    return collapse_iqvia(parse_iqvia(_IQVIA_SLICE.read_bytes()))


@pytest.fixture()
def bundle(records) -> UniverseBundle:
    return UniverseBundle(list(records), [])


@pytest.fixture()
def universe_with_iqvia(bundle, iqvia_collapsed):
    resp = build_universe_response(bundle)
    df, recon, low = build_universe_sheet1(resp, iqvia_collapsed)
    return df, recon, low


@pytest.fixture()
def clean_store(tmp_path):
    """Point the enrichment store at an empty DB so patent/labeling cols are blank."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    return store_mod


def _by_din(df: pd.DataFrame) -> dict:
    return {str(r["din"]): r for _, r in df.iterrows()}


# ════════════════════════════════════════════════════════════════════════════
# A. Frozen-behaviour guarantees
# ════════════════════════════════════════════════════════════════════════════

_CANONICAL_SHEET1_COLUMNS = [
    "din", "ingredient_name", "dosage_form", "ingredient", "brand_name",
    "company", "strength", "route", "status", "_drug_code",
    "noc_date", "reason_for_supplement", "submission_class",
    "noc_submission_type", "noc_therapeutic_class",
    "patent_count", "patent_number", "patent_grant_date", "patent_expiry_date",
    "active_ingredient", "nonmedicinal_ingredients", "pack_size", "pack_style",
    "color", "shape", "size_mm", "weight", "ph",
    "dp_6yr_no_file_date", "pediatric_extension", "data_protection_ends",
]


def _frozen_response() -> SearchResponse:
    dpd = [
        DrugRecord(source="DPD", din="11111111", ingredient="MOLX 100 MG",
                   brand_name="BRANDX", company="ALPHA INC", strength="100 MG",
                   status="Marketed", dosage_form="Tablet", route="Oral",
                   all_ingredients=["MOLX"], source_specific={"drug_code": 1}),
        DrugRecord(source="DPD", din="22222222", ingredient="ASA 81 MG",
                   brand_name="ASPIRIN", company="BAYER INC", strength="81 MG",
                   status="Marketed", dosage_form="Tablet", route="Oral",
                   all_ingredients=["ASA"], source_specific={"drug_code": 2}),
    ]
    noc = [DrugRecord(source="NOC", din="11111111", source_specific={
        "submission_type": "NDS", "noc_date": "2020-01-01", "submission_class": "NDS",
        "reason_for_supplement": "", "therapeutic_class": "X"})]
    return SearchResponse(
        metadata=SearchMetadata(query="q", field="ingredient", timestamp="t"),
        sources=[SourceResult(source="DPD", status="ok", records=dpd),
                 SourceResult(source="NOC", status="ok", records=noc)],
    )


def test_A_default_flag_byte_identical_and_intersection(clean_store):
    """build_sheet1 default == include_dpd_only=False, and still yields DPD ∩ NOC."""
    resp = _frozen_response()
    default = build_sheet1(resp)
    explicit_false = build_sheet1(resp, include_dpd_only=False)
    pd.testing.assert_frame_equal(default, explicit_false)
    assert set(default["din"]) == {"11111111"}, "default must drop the DPD-only DIN"


def test_A_canonical_column_order_pinned(clean_store):
    """The frozen Sheet-1 schema/column order is unchanged (anchor on a clean store)."""
    resp = _frozen_response()
    df = build_sheet1(resp)
    assert list(df.columns) == _CANONICAL_SHEET1_COLUMNS


def test_A_universe_preserves_canonical_prefix_and_extends(clean_store, bundle):
    """include_dpd_only=True keeps the canonical column order and only ADDS columns
    (iqvia_match_confidence); it never reorders or drops the frozen schema."""
    df, _recon, _low = build_universe_sheet1(build_universe_response(bundle), None)
    for col in _CANONICAL_SHEET1_COLUMNS:
        assert col in df.columns, f"universe dropped frozen column {col!r}"
    # Frozen columns keep their relative order.
    present = [c for c in df.columns if c in _CANONICAL_SHEET1_COLUMNS]
    assert present == _CANONICAL_SHEET1_COLUMNS
    assert "iqvia_match_confidence" in df.columns


def test_A_tab_buttons_wire_same_endpoints_and_params():
    """Tab A buttons hit the SAME endpoints/params as today — assert the wiring."""
    from app.main import _HTML_UI as html
    # Frozen search/export wiring intact.
    for needle in ("doExport(false)", "doExport(true)", "/export/start",
                   "queries, field", "filter_criteria"):
        assert needle in html, f"frozen wiring missing: {needle!r}"
    # Universe buttons post to /universe/start with {mode, enable_ocr, filter_criteria}
    # and reuse the SAME SSE + result plumbing as the export path.
    for needle in ("doUniverse('full')", "doUniverse('filter_enrich')",
                   "/universe/start", "mode, enable_ocr",
                   "body.filter_criteria = criteria",
                   "/export/stream/${jobId}"):
        assert needle in html, f"universe wiring missing: {needle!r}"
    # Reset still clears all caches (incl. universe) via the one endpoint.
    assert "/api/reset-all-caches" in html


def test_A_shared_modules_unchanged_for_existing_flows(clean_store):
    """screen + match_iqvia behave exactly as before on the legacy DPD∩NOC sheet."""
    # The IQVIA matcher is reused verbatim — exact-brand priority anchor from
    # tests/test_iqvia.py still holds (PRO DOC generic-label aggregation 37942+48).
    sheet1 = pd.DataFrame([
        {"din": "02314908", "ingredient": "METFORMIN HYDROCHLORIDE",
         "brand_name": "PRO-METFORMIN", "company": "PRO DOC LIMITEE",
         "strength": "500 MG", "status": "Marketed"},
        {"din": "02246834", "ingredient": "METFORMIN HYDROCHLORIDE",
         "brand_name": "SANDOZ METFORMIN", "company": "SANDOZ CANADA INC",
         "strength": "500 MG", "status": "Marketed"},
    ])
    iqvia = pd.DataFrame([
        {"Combined Molecule": "METFORMIN", "Product": "METFORMIN",
         "Manufacturer": "PRO DOC", "Strength": "500MG",
         "Units MAT 12/2025": 37942, "Dollars MAT 12/2025": 379420},
        {"Combined Molecule": "METFORMIN", "Product": "PRO-METFORMIN",
         "Manufacturer": "PRO DOC", "Strength": "0.5GM",
         "Units MAT 12/2025": 48, "Dollars MAT 12/2025": 480},
    ])
    enriched, _ = match_iqvia_to_sheet1(sheet1, iqvia)
    row = enriched[enriched["din"] == "02314908"]
    assert int(row["Units MAT 12/2025"].iloc[0]) == 37942 + 48


# ════════════════════════════════════════════════════════════════════════════
# B. Universe build (Option 3) — real extract, DPD-only, no PDF
# ════════════════════════════════════════════════════════════════════════════

def test_B_record_count_matches_independent_parse(records):
    """Universe product count == independent count of drug.txt rows with a DIN."""
    import csv
    indep = 0
    with open(_EXTRACT / "drug.txt", encoding="latin-1", newline="") as fh:
        for row in csv.reader(fh):
            if len(row) > 3 and row[3].strip() and row[0].strip():
                indep += 1
    assert len(records) == indep == _EXPECT_RECORD_COUNT


def test_B_column_verify_runs_on_real_extract():
    """_verify_columns passes against the GLUCOPHAGE 02099233 anchor in the fixture."""
    U._verify_columns(_EXTRACT)  # raises AssertionError on schema drift


def test_load_records_pure_parse_honors_cache_dir(monkeypatch):
    """load_dpd_universe_records honours the cache_dir arg and NEVER downloads.

    Regression for the param that used to be silently dropped (download was hard-wired
    to UNIVERSE_CACHE_DIR).  We point _download_extract at a tripwire and prove the
    parser reads the fixture dir without ever calling it.
    """
    def _boom():
        raise AssertionError("the parser must not download — download is get_universe's job")

    monkeypatch.setattr(U, "_download_extract", _boom)
    recs = U.load_dpd_universe_records(_EXTRACT)
    assert len(recs) == _EXPECT_RECORD_COUNT
    assert {r.din for r in recs} >= set(_ANCHORS)


def test_load_records_runs_verify_on_parsed_cache_dir(monkeypatch):
    """The schema check runs against the cache_dir actually parsed, so offline fixture
    parsing keeps its guard (not only the production UNIVERSE_CACHE_DIR)."""
    seen = {}
    real_verify = U._verify_columns

    def _spy(cache_dir):
        seen["dir"] = cache_dir
        return real_verify(cache_dir)

    monkeypatch.setattr(U, "_verify_columns", _spy)
    U.load_dpd_universe_records(_EXTRACT)
    assert seen.get("dir") == _EXTRACT, "verify must run on the parsed cache_dir"


def test_B_dpd_identity_fully_populated(records):
    by_din = {r.din: r for r in records}
    glu = by_din["02099233"]
    assert glu.brand_name == "GLUCOPHAGE"
    assert glu.company == "SANOFI-AVENTIS CANADA INC"
    assert glu.dosage_form == "TABLET"
    assert glu.status == "MARKETED"
    assert "METFORMIN" in glu.ingredient.upper()
    assert glu.source_specific.get("drug_code") == 17104


def test_B_current_status_collapses_history(records):
    """A DIN with many status-history rows resolves to its current (flag Y) status."""
    # TAPAZOLE has 10 status rows (mixed APPROVED/MARKETED); current flag-Y = MARKETED.
    by_din = {r.din: r for r in records}
    assert by_din[_GRANDFATHERED_DIN].status == "MARKETED"


def test_B_universe_noc_patent_dp_uniformly_blank(clean_store, bundle):
    """Cite-or-blank: with no NOC/patent/DP join, those columns are blank for EVERY row."""
    df, _recon, _low = build_universe_sheet1(build_universe_response(bundle), None)
    assert not df.empty
    for col in U._NOC_COLS:
        assert (df[col].fillna("") == "").all(), f"{col} must be blank for all universe rows"
    for col in ("patent_number", "patent_grant_date", "patent_expiry_date",
                "data_protection_ends", "dp_6yr_no_file_date"):
        if col in df.columns:
            assert df[col].isna().all() or (df[col].fillna("") == "").all(), (
                f"{col} must be blank in the no-join universe"
            )


def test_B_primitive_noc_anchored_populated_vs_grandfathered_blank(clean_store):
    """B at the real seam: build_sheet1(include_dpd_only=True) populates NOC for a
    NOC-anchored DIN and leaves a grandfathered (DPD-only) DIN's NOC blank."""
    dpd = [
        DrugRecord(source="DPD", din="11111111", ingredient="MOLX 100 MG",
                   brand_name="BX", company="A INC", strength="100 MG",
                   status="Marketed", dosage_form="Tablet",
                   source_specific={"drug_code": 1}),
        DrugRecord(source="DPD", din="22222222", ingredient="ASA 81 MG",
                   brand_name="ASPIRIN", company="BAYER INC", strength="81 MG",
                   status="Marketed", dosage_form="Tablet",
                   source_specific={"drug_code": 2}),
    ]
    noc = [DrugRecord(source="NOC", din="11111111", source_specific={
        "submission_type": "NDS", "noc_date": "2020-01-01", "submission_class": "NDS",
        "reason_for_supplement": "", "therapeutic_class": "ONC"})]
    resp = SearchResponse(
        metadata=SearchMetadata(query="q", field="ingredient", timestamp="t"),
        sources=[SourceResult(source="DPD", status="ok", records=dpd),
                 SourceResult(source="NOC", status="ok", records=noc)])
    df = build_sheet1(resp, include_dpd_only=True)
    by = _by_din(df)
    assert set(by) == {"11111111", "22222222"}
    assert by["11111111"]["noc_date"] == "2020-01-01"           # NOC-anchored → populated
    assert by["11111111"]["noc_submission_type"] == "NDS"
    # Grandfathered: build_sheet1 emits the 'No NOC record' sentinel; the universe
    # blanks it. Here (raw primitive) the sentinel is present and distinct.
    assert by["22222222"]["noc_date"] == U._NOC_SENTINEL


def test_B_disclaimer_and_schema_in_workbook(clean_store, bundle):
    df, recon, low = build_universe_sheet1(build_universe_response(bundle), None)
    s2 = build_universe_sheet2(bundle)
    xlsx = U.build_universe_workbook(df, s2, recon, low)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames[0].startswith("⚠"), f"first sheet must be the Read Me; got {wb.sheetnames}"
    assert "Full Universe (no PDF)" in wb.sheetnames
    assert "Generic Submissions" in wb.sheetnames
    # The PDF-omitted disclaimer text is present.
    disc = wb[wb.sheetnames[0]]
    joined = "\n".join(str(c.value) for row in disc.iter_rows() for c in row if c.value)
    assert "Product-Monograph (PDF) data is OMITTED" in joined
    assert "Filter & enrich" in joined


def test_B_disclaimer_warns_patent_noc_dp_not_evaluated(clean_store, bundle):
    """Safety: the Read-Me must say NOC/patent/DP are NOT evaluated, that a blank cell
    is 'not assessed' not 'none' / 'free to launch', and point ONLY to the standard
    per-product export (the path that actually runs those joins — NOT 'Filter & enrich',
    which only adds PDF labeling)."""
    df, recon, low = build_universe_sheet1(build_universe_response(bundle), None)
    s2 = build_universe_sheet2(bundle)
    xlsx = U.build_universe_workbook(df, s2, recon, low)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    disc = wb[wb.sheetnames[0]]
    cells = [c for row in disc.iter_rows() for c in row if c.value]
    joined = " ".join(str(c.value) for c in cells)
    assert "NOT evaluated" in joined
    assert "not assessed" in joined
    assert "free to launch" in joined
    assert "standard per-product export" in joined
    # The caution must be rendered as a distinct bold-red cell (matches the IQVIA
    # low-confidence caution styling) so it cannot be missed.
    bold_red = [
        c for c in cells
        if "NOT evaluated" in str(c.value) and c.font and c.font.bold
        and getattr(c.font.color, "rgb", None) and "9B1C1C" in str(c.font.color.rgb)
    ]
    assert bold_red, "the not-evaluated warning must be a bold red caution line"
    # And the remedy must NOT misdirect to 'Filter & enrich' for these three joins:
    # that phrase may appear only in the (correct) PDF paragraph, never in the caution.
    assert "Filter & enrich" not in str(bold_red[0].value)


def test_B_option3_does_no_pdf_fetch(monkeypatch, bundle):
    """Spy the labeling path: Option 3 must fetch ZERO Product Monographs."""
    from app import universe_job
    from app.jobs import create_job

    calls = {"n": 0}

    async def _boom(*a, **k):
        calls["n"] += 1
        return {}

    async def _fake_universe(force_refresh=False):
        return bundle

    monkeypatch.setattr(universe_job, "get_universe", _fake_universe)
    monkeypatch.setattr(universe_job, "enrich_labeling_batch_fast", _boom)
    monkeypatch.setattr(universe_job, "_resolve_iqvia", lambda job: None)
    import app.enrichment.labeling as labeling
    monkeypatch.setattr(labeling, "enrich_labeling_batch_fast", _boom)

    job = create_job("u3", "Full universe", "ingredient")
    asyncio.run(universe_job.run_universe_full_job(job))
    assert job.status == "complete", job.error
    assert calls["n"] == 0, "Option 3 must never fetch a PM PDF"
    assert job.result_path and Path(job.result_path).exists()


# ════════════════════════════════════════════════════════════════════════════
# C. IQVIA on the universe — DPD-native match, no PDF
# ════════════════════════════════════════════════════════════════════════════

def test_C_iqvia_sizing_matches_real_anchors(universe_with_iqvia):
    df, _recon, _low = universe_with_iqvia
    by = _by_din(df)
    for din, (_brand, _co, dollars, units, _conf) in _ANCHORS.items():
        assert int(by[din][_DOLLARS]) == dollars, f"{din} dollars"
        assert int(by[din][_UNITS]) == units, f"{din} units"


def test_C_generic_label_aggregation_on_universe(universe_with_iqvia):
    """PRO-METFORMIN DIN reflects BOTH the generic METFORMIN group and the exact
    PRO-METFORMIN group (37942 + 48 = 37990) — the real aggregation, on the universe."""
    df, _recon, _low = universe_with_iqvia
    assert int(_by_din(df)["02314908"][_UNITS]) == 37990


def test_C_unmatched_din_has_no_sizing_no_fabrication(universe_with_iqvia):
    """TAPAZOLE has no IQVIA presence → metric cells None/NaN, never 0."""
    df, _recon, _low = universe_with_iqvia
    row = _by_din(df)[_GRANDFATHERED_DIN]
    for col in (_DOLLARS, _UNITS, _EXT):
        val = row[col]
        assert val is None or pd.isna(val), f"{col}={val!r} must be blank, not 0"


def test_C_match_logic_identical_to_existing_matcher(bundle, iqvia_collapsed):
    """The universe must REUSE the matcher unchanged: its IQVIA columns equal a
    direct match_iqvia_to_sheet1() call on the same blanked sheet."""
    resp = build_universe_response(bundle)
    base = U._blank_noc_sentinels(build_sheet1(resp, include_dpd_only=True))
    direct, _recon = match_iqvia_to_sheet1(base, iqvia_collapsed)
    universe_df, _r, _l = build_universe_sheet1(resp, iqvia_collapsed)
    direct_by, uni_by = _by_din(direct), _by_din(universe_df)
    for din in _ANCHORS:
        for col in (_DOLLARS, _UNITS, _EXT):
            a, b = direct_by[din][col], uni_by[din][col]
            assert (a == b) or (pd.isna(a) and pd.isna(b)), f"{din}/{col} diverged"


def test_C_match_confidence_column_real(universe_with_iqvia):
    df, _recon, _low = universe_with_iqvia
    by = _by_din(df)
    for din, (_b, _c, _d, _u, conf) in _ANCHORS.items():
        assert by[din]["iqvia_match_confidence"] == conf, f"{din} confidence"
    assert by[_GRANDFATHERED_DIN]["iqvia_match_confidence"] == "none"


def test_C_low_confidence_kpi_counts_house_brand_fuzzy():
    """The low-confidence KPI counts a matched-but-fuzzy (score<85) house-brand match.

    In real data exact-brand priority resolves PRO DOC / JAMP / Pharmascience to
    'exact', so the 'low' band specifically catches the residual fuzzy matches; this
    pins that a matched score in [65,85) with no exact-brand note counts as 'low'.
    """
    sheet1 = pd.DataFrame([
        {"din": "30000001", "brand_name": "JAMP-METFORMIN"},   # fuzzy, weak
        {"din": "30000002", "brand_name": "PRO-METFORMIN"},    # exact-brand
    ])
    recon = pd.DataFrame([
        {"din": "30000001", "status": "matched", "top_score": 70.0, "notes": "score=70"},
        {"din": "30000002", "status": "matched", "top_score": 100.0,
         "notes": "exact-brand match; score=100"},
    ])
    df, low = U.attach_match_confidence(sheet1, recon)
    conf = {r["din"]: r["iqvia_match_confidence"] for _, r in df.iterrows()}
    assert conf["30000001"] == "low"
    assert conf["30000002"] == "exact"
    assert low == 1


def test_C_real_cohort_low_count_is_zero(universe_with_iqvia):
    """Lock the real-data observation: the cohort's house brands all resolve exact."""
    _df, _recon, low = universe_with_iqvia
    assert low == 0


def test_C_match_has_no_pdf_dependency(monkeypatch, bundle, iqvia_collapsed):
    """The IQVIA match runs with an empty labeling store and no PM fetch."""
    import app.enrichment.labeling as labeling

    async def _boom(*a, **k):
        raise AssertionError("IQVIA match must not fetch a PM PDF")

    monkeypatch.setattr(labeling, "enrich_labeling_batch_fast", _boom)
    df, _recon, _low = build_universe_sheet1(build_universe_response(bundle), iqvia_collapsed)
    assert int(_by_din(df)["02516187"][_UNITS]) == 218591


# ════════════════════════════════════════════════════════════════════════════
# D. Filter stage (Option 4, pre-enrich) — six criteria on the no-PDF layer
# ════════════════════════════════════════════════════════════════════════════

def _products(universe_with_iqvia, bundle):
    df, _recon, _low = universe_with_iqvia
    s2 = build_universe_sheet2(bundle)
    products, _w = screen.compute_products(df, s2)
    return products


def _by_form(products: pd.DataFrame, ing_sub: str, form: str):
    m = products[(products["ingredient"].str.contains(ing_sub, case=False)) &
                 (products["dosage_form"] == form)]
    assert len(m) == 1, f"expected one {ing_sub}/{form} product, got {len(m)}"
    return m.iloc[0]


def test_D_all_six_criteria_compute_on_no_pdf_layer(universe_with_iqvia, bundle):
    """All six criteria (incl. the three IQVIA sizing ones) compute with no PDF."""
    products = _products(universe_with_iqvia, bundle)
    met = _by_form(products, "METFORMIN HYDROCHLORIDE", "TABLET")
    assert met["competitors"] == 3
    assert met["approvals"] == 3
    assert met["value_sizeable"] == 7823285
    assert met["quantity_sizeable"] == 583693
    assert met["quantity_ext_sizeable"] == 247519340
    assert met["filings"] == 0  # no GSUR rows in the universe sheet2


@pytest.mark.parametrize("metric,operator,value,should_pass_ing,should_fail_ing", [
    # competitors: METFORMIN(3) vs AMLODIPINE(2) at boundary 'above 2'
    ("competitors", "above", 2, "METFORMIN HYDROCHLORIDE", "AMLODIPINE (AMLODIPINE BESYLATE)"),
    # approvals: same shape
    ("approvals", "above", 2, "METFORMIN HYDROCHLORIDE", "AMLODIPINE (AMLODIPINE BESYLATE)"),
    # value: PROGESTERONE($34.2M) passes 'above 1,000,000'; CADUET ($356,929) fails
    ("value", "above", 1_000_000, "PROGESTERONE",
     "AMLODIPINE (AMLODIPINE BESYLATE) + ATORVASTATIN (ATORVASTATIN CALCIUM)"),
    # quantity: METFORMIN(583,693) passes 'above 100,000'; AMLODIPINE(34,249) fails
    ("quantity", "above", 100_000, "METFORMIN HYDROCHLORIDE", "AMLODIPINE (AMLODIPINE BESYLATE)"),
    # quantity_ext: METFORMIN(247.5M) passes 'above 50,000,000'; PROGESTERONE(37.1M) fails
    ("quantity_ext", "above", 50_000_000, "METFORMIN HYDROCHLORIDE", "PROGESTERONE"),
])
def test_D_criterion_boundary_pass_and_fail(universe_with_iqvia, bundle,
                                            metric, operator, value,
                                            should_pass_ing, should_fail_ing):
    products = _products(universe_with_iqvia, bundle)
    criteria = screen.parse_criteria([{"metric": metric, "operator": operator, "value": value}])
    qualifying = screen.apply_criteria(products, criteria)
    quals = set(qualifying["ingredient"])
    assert any(should_pass_ing in q for q in quals), f"{should_pass_ing} should pass {metric} {operator} {value}"
    assert not any(should_fail_ing == q for q in quals), f"{should_fail_ing} should fail"


def test_D_exactly_operator_zero_value(universe_with_iqvia, bundle):
    """METHIMAZOLE (TAPAZOLE, no IQVIA) has value exactly 0 → passes 'value exactly 0'."""
    products = _products(universe_with_iqvia, bundle)
    criteria = screen.parse_criteria([{"metric": "value", "operator": "exactly", "value": 0}])
    quals = set(screen.apply_criteria(products, criteria)["ingredient"])
    assert any("METHIMAZOLE" in q for q in quals)
    assert not any("PROGESTERONE" == q for q in quals)


def test_D_filings_criterion_with_gsur(records):
    """Filings (GSUR) criterion: a metformin filing makes the metformin product pass."""
    bundle = UniverseBundle(list(records), [DrugRecord(
        source="GenericSubmissions", ingredient="metformin hydrochloride",
        company="GenCo", all_ingredients=["metformin"], status="Under Review",
        source_specific={"therapeutic_area": "X", "date_accepted": "2025/01"})])
    df, _r, _l = build_universe_sheet1(build_universe_response(bundle), None)
    s2 = build_universe_sheet2(bundle)
    products, _w = screen.compute_products(df, s2)
    criteria = screen.parse_criteria([{"metric": "filings", "operator": "above", "value": 0}])
    quals = set(screen.apply_criteria(products, criteria)["ingredient"])
    assert any("METFORMIN" in q for q in quals), "metformin product should have a filing"
    assert not any("PROGESTERONE" == q for q in quals), "progesterone has no filing"


def test_D_iqvia_absent_disables_sizing_criteria_gracefully(clean_store, bundle):
    """With no IQVIA loaded, criteria 4-6 raise a clear error (no crash); 1-3 still work."""
    df, _r, _l = build_universe_sheet1(build_universe_response(bundle), None)
    s2 = build_universe_sheet2(bundle)
    # No IQVIA → no Dollars/Units MAT columns at all.
    assert not any("MAT" in c for c in df.columns)
    # Criteria 1-3 compute fine.
    products, _w = screen.compute_products(df, s2)
    c123 = screen.parse_criteria([{"metric": "competitors", "operator": "above", "value": 2}])
    assert not screen.apply_criteria(products, c123).empty
    # A value criterion must raise (graceful guard), not silently pass on zeros.
    crit = screen.parse_criteria([{"metric": "value", "operator": "above", "value": 1}])
    with pytest.raises(ValueError):
        screen.build_filtered_workbook(df, s2, crit)


def test_D_filter_stage_no_pdf_fetch(monkeypatch, universe_with_iqvia, bundle):
    """compute_products + apply_criteria + build_filtered_workbook fetch no PM PDF."""
    import app.enrichment.labeling as labeling

    async def _boom(*a, **k):
        raise AssertionError("filter stage must not fetch a PM PDF")

    monkeypatch.setattr(labeling, "enrich_labeling_batch_fast", _boom)
    df, _recon, _low = universe_with_iqvia
    s2 = build_universe_sheet2(bundle)
    criteria = screen.parse_criteria([{"metric": "value", "operator": "above", "value": 1_000_000}])
    xlsx, summary, detail, _w = screen.build_filtered_workbook(df, s2, criteria)
    assert xlsx[:2] == b"PK" and len(summary) >= 1


# ════════════════════════════════════════════════════════════════════════════
# E. Enrich-survivors stage (Option 4)
# ════════════════════════════════════════════════════════════════════════════

def _run_option4(monkeypatch, bundle, criteria, enrich_fn, iqvia=None):
    from app import universe_job
    from app.jobs import create_job

    async def _fake_universe(force_refresh=False):
        return bundle

    monkeypatch.setattr(universe_job, "get_universe", _fake_universe)
    monkeypatch.setattr(universe_job, "enrich_labeling_batch_fast", enrich_fn)
    monkeypatch.setattr(universe_job, "_resolve_iqvia", lambda job: iqvia)
    job = create_job("u4", "Full universe", "ingredient", filter_criteria=criteria)
    asyncio.run(universe_job.run_universe_filter_enrich_job(job, enable_ocr=False))
    return job


def test_E_only_survivors_enriched(monkeypatch, bundle):
    """Survivors (METFORMIN Tablets: 3 marketed companies) are fetched; everything
    filtered out — progesterone, amlodipine, methimazole — is NEVER fetched."""
    seen = {}

    async def _fake(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        seen.update(din_map)
        return {}

    job = _run_option4(monkeypatch, bundle,
                       [{"metric": "competitors", "operator": "above", "value": 2}], _fake)
    assert job.status == "complete", job.error
    # competitors>2 → only the metformin Tablet product (3 marketed companies).
    assert set(seen) == {"02099233", "02314908", "02380196"}, set(seen)
    for excluded in ("02516187", "02493578", "02284065", "02522519", _GRANDFATHERED_DIN):
        assert excluded not in seen, f"filtered-out DIN {excluded} must not be fetched"


def test_E_each_survivor_fetched_once(monkeypatch, bundle):
    counts = {}

    async def _fake(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        for d in din_map:
            counts[d] = counts.get(d, 0) + 1
        return {}

    _run_option4(monkeypatch, bundle,
                 [{"metric": "competitors", "operator": "above", "value": 2}], _fake)
    assert counts == {"02099233": 1, "02314908": 1, "02380196": 1}


def test_E_survivor_labeling_cite_or_blank(clean_store, bundle):
    """patch_labeling_for_dins applies cited values for survivors and leaves
    non-survivors blank; absent fields stay blank (never fabricated)."""
    store = clean_store
    import time
    # Seed a cited color for one survivor; a 'Not stated' ph (cite-or-blank).
    store.upsert_labeling("02314908", {
        "color": "white", "color_page": 6, "ph": "Not stated",
        "needs_ocr": 0, "has_unverified": 0, "drug_code": 80043, "fetched_at": time.time(),
    })
    df, _r, _l = build_universe_sheet1(build_universe_response(bundle), None)
    patched = U.patch_labeling_for_dins(df, {"02314908"})
    by = _by_din(patched)
    assert by["02314908"]["color"] == "white"          # cited → applied
    # A non-survivor keeps its blank labeling.
    non = by["02516187"]["color"]
    assert non is None or pd.isna(non) or non == "", "non-survivor labeling must stay blank"


def test_E_empty_survivor_set_builds_valid_empty_workbook(monkeypatch, bundle):
    """A filter nothing passes → a valid (empty-enrichment) workbook, no crash, no fetch."""
    fetched = {"n": 0}

    async def _fake(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        fetched["n"] += len(din_map)
        return {}

    job = _run_option4(monkeypatch, bundle,
                       [{"metric": "competitors", "operator": "above", "value": 999}], _fake)
    assert job.status == "complete", job.error
    assert fetched["n"] == 0, "no survivors → no PM fetch"
    assert job.result_path and Path(job.result_path).exists()
    wb = openpyxl.load_workbook(job.result_path)
    assert "Summary" in wb.sheetnames and "Detail" in wb.sheetnames


def test_E_all_pass_enriches_every_din(monkeypatch, bundle):
    """A trivially-true filter (competitors above 0) enriches every product's DINs."""
    seen = set()

    async def _fake(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        seen.update(din_map)
        return {}

    job = _run_option4(monkeypatch, bundle,
                       [{"metric": "competitors", "operator": "above", "value": 0}], _fake)
    assert job.status == "complete", job.error
    # Every cohort DIN is in some product with ≥1 marketed company → all enriched.
    assert seen == {r.din for r in bundle.dpd_records}


def test_E_final_excel_has_universe_plus_survivor_schema(monkeypatch, clean_store, bundle):
    """The Option-4 workbook is the Summary + Detail schema with survivor rows."""
    import time
    clean_store.upsert_labeling("02380196", {
        "color": "white", "needs_ocr": 0, "has_unverified": 0,
        "drug_code": 86602, "fetched_at": time.time()})

    async def _fake(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        return {}

    job = _run_option4(monkeypatch, bundle,
                       [{"metric": "competitors", "operator": "above", "value": 2}], _fake)
    assert job.status == "complete", job.error
    wb = openpyxl.load_workbook(job.result_path)
    assert wb.sheetnames[:2] == ["Summary", "Detail"]
    # Detail carries the surviving DINs.
    detail = wb["Detail"]
    headers = [c.value for c in detail[1]]
    assert "DIN" in headers


# ════════════════════════════════════════════════════════════════════════════
# F. 4-hour cache + reset
# ════════════════════════════════════════════════════════════════════════════

def test_F_one_build_within_window(monkeypatch):
    """Option 3 then Option 4 within the window trigger ONE build, not two."""
    builds = {"n": 0}

    def _fake_load(cache_dir=U.UNIVERSE_CACHE_DIR):
        builds["n"] += 1
        return [DrugRecord(source="DPD", din="10000001", ingredient="MOLX 100 MG",
                           brand_name="B", company="C INC", strength="100 MG",
                           status="Marketed", dosage_form="Tablet",
                           source_specific={"drug_code": 1})]

    downloads = {"n": 0}

    def _fake_download():
        downloads["n"] += 1

    async def _fake_gsur():
        return []

    monkeypatch.setattr(U, "load_dpd_universe_records", _fake_load)
    monkeypatch.setattr(U, "_download_extract", _fake_download)
    monkeypatch.setattr(U, "_load_gsur_records", _fake_gsur)
    U._CACHE["bundle"] = None

    asyncio.run(U.get_universe())
    asyncio.run(U.get_universe())
    assert builds["n"] == 1, "fresh cache reused — no double parse"


def test_F_rebuild_after_ttl(monkeypatch):
    builds = {"n": 0}

    def _fake_load(cache_dir=U.UNIVERSE_CACHE_DIR):
        builds["n"] += 1
        return []

    async def _fake_gsur():
        return []

    monkeypatch.setattr(U, "load_dpd_universe_records", _fake_load)
    monkeypatch.setattr(U, "_load_gsur_records", _fake_gsur)
    monkeypatch.setattr(U, "_download_extract", lambda: None)  # keep offline
    U._CACHE["bundle"] = None

    asyncio.run(U.get_universe())
    assert builds["n"] == 1
    U._CACHE["bundle"].built_at -= (U.UNIVERSE_TTL + 1)
    asyncio.run(U.get_universe())
    assert builds["n"] == 2, "stale cache must rebuild after the 4h window"


def test_F_reset_clears_inprocess_and_ondisk(monkeypatch, tmp_path):
    """reset_universe_cache() drops the in-process bundle AND removes the on-disk
    extract dir so the next request must re-pull allfiles.zip."""
    # Point the on-disk extract dir at a temp location with a sentinel file.
    fake_dir = tmp_path / "universe"
    fake_dir.mkdir()
    (fake_dir / "drug.txt").write_text("sentinel", encoding="latin-1")
    monkeypatch.setattr(U, "UNIVERSE_CACHE_DIR", fake_dir)
    U._CACHE["bundle"] = UniverseBundle([], [])

    had = U.reset_universe_cache()
    assert had == 1
    assert U._CACHE["bundle"] is None, "in-process cache must be cleared"
    assert not fake_dir.exists(), "on-disk extract must be removed so the next pull is fresh"


def test_F_reset_then_request_rebuilds(monkeypatch):
    builds = {"n": 0}

    def _fake_load(cache_dir=U.UNIVERSE_CACHE_DIR):
        builds["n"] += 1
        return []

    async def _fake_gsur():
        return []

    monkeypatch.setattr(U, "load_dpd_universe_records", _fake_load)
    monkeypatch.setattr(U, "_load_gsur_records", _fake_gsur)
    monkeypatch.setattr(U, "_download_extract", lambda: None)  # keep offline
    monkeypatch.setattr(U, "UNIVERSE_CACHE_DIR", Path("/nonexistent_universe_dir"))
    U._CACHE["bundle"] = None

    asyncio.run(U.get_universe())
    U.reset_universe_cache()
    assert U._CACHE["bundle"] is None
    asyncio.run(U.get_universe())
    assert builds["n"] == 2, "reset must force a fresh build"


def test_F_reset_all_caches_endpoint_invokes_universe_reset(monkeypatch):
    """/api/reset-all-caches wires reset_universe_cache (returns universe_cleared)."""
    from fastapi.testclient import TestClient
    import app.main as main

    called = {"n": 0}
    monkeypatch.setattr(main, "cache_clear_all", lambda: 0)
    monkeypatch.setattr(main, "reset_patents_table", lambda: 0)
    monkeypatch.setattr(main, "reset_labeling_table", lambda: 0)
    U._CACHE["bundle"] = UniverseBundle([], [])

    client = TestClient(main.app)
    resp = client.post("/api/reset-all-caches")
    assert resp.status_code == 200
    body = resp.json()
    assert "universe_cleared" in body
    assert U._CACHE["bundle"] is None, "endpoint must clear the universe cache"


# ════════════════════════════════════════════════════════════════════════════
# G. End-to-end through the real endpoints
# ════════════════════════════════════════════════════════════════════════════

async def _drive_endpoint(monkeypatch, bundle, body, iqvia=None):
    from httpx import ASGITransport, AsyncClient
    import app.universe_job as uj
    from app.main import app

    async def _fake_universe(force_refresh=False):
        return bundle

    async def _fake_enrich(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        return {}

    monkeypatch.setattr(uj, "get_universe", _fake_universe)
    monkeypatch.setattr(uj, "enrich_labeling_batch_fast", _fake_enrich)
    monkeypatch.setattr(uj, "_resolve_iqvia", lambda job: iqvia)

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://t") as client:
        start = await client.post("/universe/start", json=body)
        assert start.status_code == 200, start.text
        job_id = start.json()["job_id"]
        # Drain the SSE stream until the job reports terminal status.
        import json as _json
        async with client.stream("GET", f"/export/stream/{job_id}") as s:
            async for line in s.aiter_lines():
                if line.startswith("data:"):
                    evt = _json.loads(line[5:].strip())
                    if evt.get("status") in ("complete", "error"):
                        assert evt["status"] == "complete", evt
                        break
        res = await client.get(f"/export/result/{job_id}")
        assert res.status_code == 200, res.text
        return res.content


async def test_G_option3_endpoint_returns_valid_xlsx(monkeypatch, bundle):
    content = await _drive_endpoint(monkeypatch, bundle, {"mode": "full", "enable_ocr": False})
    assert content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames[0].startswith("⚠")
    assert "Full Universe (no PDF)" in wb.sheetnames


async def test_G_option4_endpoint_returns_valid_xlsx(monkeypatch, bundle, iqvia_collapsed):
    body = {"mode": "filter_enrich", "enable_ocr": False,
            "filter_criteria": [{"metric": "competitors", "operator": "above", "value": 1}]}
    content = await _drive_endpoint(monkeypatch, bundle, body, iqvia=iqvia_collapsed)
    assert content[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(content))
    assert wb.sheetnames[:2] == ["Summary", "Detail"]


def test_G_endpoint_rejects_bad_mode_and_missing_criteria():
    from fastapi.testclient import TestClient
    from app.main import app
    client = TestClient(app)
    assert client.post("/universe/start", json={"mode": "bogus"}).status_code == 400
    assert client.post("/universe/start",
                       json={"mode": "filter_enrich", "filter_criteria": []}).status_code == 400


def test_G_bounded_async_completes_at_realistic_survivor_count(monkeypatch, records):
    """Many survivors → the bounded labeling path completes without deadlock or
    unbounded fan-out (peak concurrency never exceeds the configured semaphore)."""
    # Build a realistic universe of 60 marketed metformin DINs (real-shaped rows).
    big = []
    for i in range(60):
        din = f"9{i:07d}"
        big.append(DrugRecord(
            source="DPD", din=din, ingredient="METFORMIN HYDROCHLORIDE 500 MG",
            brand_name=f"GENERIC METFORMIN {i}", company=f"COMPANY {i} INC",
            strength="500 MG", status="Marketed", dosage_form="Tablet",
            route="Oral", all_ingredients=["METFORMIN"],
            source_specific={"drug_code": 900000 + i}))
    bundle = UniverseBundle(big, [])

    import app.universe_job as uj
    peak = {"cur": 0, "max": 0}
    processed = {"n": 0}
    sem_size = uj._LABEL_SEM_SIZE

    async def _bounded(din_map, enable_ocr=None, concurrency=8, on_progress=None):
        # Emulate the real semaphore-bounded fan-out to assert no unbounded spawn.
        sem = asyncio.Semaphore(concurrency)

        async def _one(d):
            async with sem:
                peak["cur"] += 1
                peak["max"] = max(peak["max"], peak["cur"])
                await asyncio.sleep(0)
                processed["n"] += 1
                peak["cur"] -= 1

        await asyncio.gather(*[_one(d) for d in din_map])
        return {}

    job = _run_option4(monkeypatch, bundle,
                       [{"metric": "competitors", "operator": "above", "value": 1}], _bounded)
    assert job.status == "complete", job.error
    assert processed["n"] == 60, "every survivor processed (no deadlock)"
    assert peak["max"] <= sem_size, f"fan-out exceeded semaphore: {peak['max']} > {sem_size}"
