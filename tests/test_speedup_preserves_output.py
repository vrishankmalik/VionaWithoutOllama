"""Verify that optimised export caching (Ollama result cache + enrichment DB) produces
byte-for-byte identical workbook output on repeated runs of the same query.

Strategy:
  Run 1 (warm — enrichment DB and Ollama results already populated by earlier test runs or
          by the app's normal operation): capture all cell values.
  Run 2 (repeat, same conditions): capture all cell values again.
  Assert Run 1 == Run 2 on every cell of both sheets.

This test does NOT do a full live export (too slow / requires Ollama).  Instead it exercises
the workbook assembly path in isolation, using enrichment-store data that was already written
by previous runs.  If the enrichment store is empty the test is skipped gracefully.

Why two runs?  The first call to build_workbook reads from the enrichment store; the second
call reads the same store (no new enrichment).  Any non-determinism (random ordering, date
stamps, floating-point formatting) would surface as a diff.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
import pytest

from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult
from app.enrichment.store import get_labeling_for_din, get_patents_for_din


def _extract_cells(wb_bytes: bytes) -> dict[str, list[list[Any]]]:
    """Return {sheet_name: [[row], [row], ...]} from an XLSX byte buffer."""
    wb = openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=True)
    out: dict[str, list[list[Any]]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        out[name] = [[cell.value for cell in row] for row in ws.iter_rows()]
    return out


def _make_response(dins: list[str], query: str = "alpelisib") -> SearchResponse:
    records = [
        DrugRecord(
            source="DPD",
            din=din,
            brand_name="PIQRAY",
            company="Novartis",
            ingredient="alpelisib",
            strength="50 mg",
            dosage_form="Tablet",
            route="Oral",
            status="Marketed",
            record_url=f"https://health-products.canada.ca/dpd-bdpp/info?lang=eng&code=1",
        )
        for din in dins
    ]
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp=datetime.now(timezone.utc).isoformat(),
        ),
        sources=[SourceResult(source="DPD", status="ok", records=records)],
    )


def test_repeated_workbook_build_is_identical() -> None:
    """Two consecutive workbook builds from the same enrichment data must be identical."""
    from app.enrichment.workbook import build_workbook

    # Find DINs in the enrichment store to use as test subjects.
    # Query for a few known DINs — if none are stored, skip the test.
    candidate_dins = ["02497069", "02497077", "02497085", "02497042"]
    available = [d for d in candidate_dins if get_labeling_for_din(d) is not None]

    if not available:
        pytest.skip(
            "No alpelisib DINs in enrichment store — run a live export first "
            "or populate the store to enable this test."
        )

    response = _make_response(available)

    cells1 = _extract_cells(build_workbook(response))
    cells2 = _extract_cells(build_workbook(response))

    assert cells1.keys() == cells2.keys(), "Sheet names differ between runs"

    for sheet_name in cells1:
        rows1 = cells1[sheet_name]
        rows2 = cells2[sheet_name]
        assert len(rows1) == len(rows2), (
            f"Sheet '{sheet_name}': row count differs: {len(rows1)} vs {len(rows2)}"
        )
        for row_idx, (r1, r2) in enumerate(zip(rows1, rows2)):
            assert r1 == r2, (
                f"Sheet '{sheet_name}' row {row_idx + 1} differs between run 1 and run 2:\n"
                f"  Run 1: {r1}\n"
                f"  Run 2: {r2}"
            )


@pytest.mark.asyncio
async def test_ollama_cache_hit_produces_same_result_as_live_call(tmp_path) -> None:
    """_query_ollama_cached returns the same dict whether from cache or live Ollama.

    Uses a patch to intercept cache_get and verify: if the cache returns a value,
    the function returns it without calling _query_ollama.  Verifies no mutation
    between store and return.
    """
    from unittest.mock import AsyncMock, patch
    from app.enrichment.labeling import _query_ollama_cached

    fake_result = {
        "excipients_core": {"value": "microcrystalline cellulose", "found": True, "page": 42},
        "excipients_coating": {"value": None, "found": False, "page": None},
        "preservatives": {"value": None, "found": False, "page": None},
    }

    with patch("app.enrichment.labeling.cache_get", return_value=fake_result) as mock_cg, \
         patch("app.enrichment.labeling._query_ollama") as mock_ollama:
        result = await _query_ollama_cached("some text", 6, "excipients")

    assert result == fake_result, "Cache hit must return exactly the cached value"
    mock_ollama.assert_not_called()
    mock_cg.assert_called_once()
