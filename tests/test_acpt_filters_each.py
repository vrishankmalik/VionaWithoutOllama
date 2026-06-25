"""ACCEPTANCE-level suite: each of the 8 screen criteria driven through
``build_filtered_workbook`` individually.

Distinct from the UNIT suites (tests/test_screen.py, tests/test_screen_filters.py):
those exercise compute/apply/parse in isolation; here we drive the full workbook
assembly path end-to-end and assert on the PRODUCED XLSX bytes — that it is
openpyxl-readable, has sheets exactly ["Summary", "Detail"], the Summary header
row is the canonical display headers, the qualifying product set is EXACT, and the
Detail DIN set is exactly the survivors' DINs.

Each test runs ONE criterion only.  Synthetic Sheet-1 frames carry KNOWN values;
the IQVIA value/quantity/quantity_ext tests reuse the human-verified golden MAT
12/2025 anchors:
    02516187  SANIS / PROGESTERONE       Dollars 21,215,081  Units 218,591
    02493578  AURO  / AURO-PROGESTERONE  Dollars 13,005,865  Units 233,159
    product (PROGESTERONE, Capsule) aggregate = $34,220,946 / 451,750 units / 300 ext

Fully offline, hermetic, order-independent.  "Today" is pinned via today=date(2026,6,24).
"""
from __future__ import annotations

import io
from datetime import date

import openpyxl
import pandas as pd

from app.enrichment.screen import (
    build_filtered_workbook,
    parse_criteria,
    parse_dosage_forms,
    parse_no_file_date,
)

_TODAY = date(2026, 6, 24)

# ── canonical Summary display headers (verbatim from spec) ────────────────────
_SUMMARY_HEADERS = [
    "Ingredient",
    "Dosage Form",
    "Number of Competitors",
    "Number of Filings",
    "Number of Approvals",
    "Value Sizeable ($)",
    "Quantity Sizeable (Units)",
    "Quantity Ext Sizeable",
]

# ── golden IQVIA per-DIN anchors (din -> (dollars, units, ext)) ───────────────
_G_SANIS = (21_215_081, 218_591, 100)
_G_AURO = (13_005_865, 233_159, 200)
_GOLD_VALUE = _G_SANIS[0] + _G_AURO[0]   # 34,220,946
_GOLD_UNITS = _G_SANIS[1] + _G_AURO[1]   # 451,750
_GOLD_EXT = _G_SANIS[2] + _G_AURO[2]     # 300

# IQVIA latest-period column headers (mirror the real header form).
_DOLLARS = "Dollars MAT 12/2025"
_UNITS = "Units MAT 12/2025"
_EXT = "Ext Units MAT 12/2025"


# ── shared synthetic builders ─────────────────────────────────────────────────

def _sheet1_iqvia() -> pd.DataFrame:
    """PROGESTERONE 'Capsule' product (3 DINs, 2 marketed) + a verbatim-distinct
    'Capsule (extended-release)' product (1 DIN) with tiny IQVIA values."""
    return pd.DataFrame([
        {"din": "02516187", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "SANIS HEALTH INC", "status": "marketed",
         _DOLLARS: _G_SANIS[0], _UNITS: _G_SANIS[1], _EXT: _G_SANIS[2]},
        {"din": "02493578", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "AURO PHARMA INC", "status": "marketed",
         _DOLLARS: _G_AURO[0], _UNITS: _G_AURO[1], _EXT: _G_AURO[2]},
        # cancelled sibling: no IQVIA match (0 to sums), not a competitor, but a
        # distinct company holding an approval.
        {"din": "09999999", "ingredient": "PROGESTERONE", "dosage_form": "Capsule",
         "company": "OLD PHARMA LTD", "status": "cancelled post market",
         _DOLLARS: None, _UNITS: None, _EXT: None},
        # verbatim-distinct dosage form → its OWN product, never merged.
        {"din": "02505223", "ingredient": "PROGESTERONE",
         "dosage_form": "Capsule (extended-release)",
         "company": "BIJUVA CO", "status": "marketed",
         _DOLLARS: 500, _UNITS: 5, _EXT: 1},
    ])


def _sheet2_progesterone() -> pd.DataFrame:
    """GSUR filings: 3 progesterone submissions + 1 unrelated (metformin)."""
    return pd.DataFrame([
        {"medicinal_ingredient": "progesterone", "company": "GenA"},
        {"medicinal_ingredient": "progesterone", "company": "GenB"},
        {"medicinal_ingredient": "progesterone", "company": "Not available"},
        {"medicinal_ingredient": "metformin", "company": "GenC"},
    ])


# ── workbook-level assertion helpers (reused by every test) ───────────────────

def _open(xlsx: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(io.BytesIO(xlsx))


def _assert_shell(wb: openpyxl.Workbook) -> list[str]:
    """Every produced workbook must have exactly these two sheets with the
    canonical Summary header row.  Returns the Summary header list."""
    assert wb.sheetnames == ["Summary", "Detail"]
    headers = [c.value for c in wb["Summary"][1]]
    assert headers == _SUMMARY_HEADERS
    return headers


def _summary_products(wb: openpyxl.Workbook) -> set[tuple]:
    """Exact set of (Ingredient, Dosage Form) qualifying rows in the Summary tab."""
    sws = wb["Summary"]
    out = set()
    for r in range(2, sws.max_row + 1):
        ing = sws.cell(r, 1).value
        form = sws.cell(r, 2).value
        if ing is None and form is None:
            continue
        out.add((ing, form))
    return out


def _summary_value(wb: openpyxl.Workbook, header: str, row: int = 2):
    headers = [c.value for c in wb["Summary"][1]]
    return wb["Summary"].cell(row, headers.index(header) + 1).value


def _detail_dins(wb: openpyxl.Workbook) -> set[str]:
    dws = wb["Detail"]
    dheaders = [c.value for c in dws[1]]
    assert "DIN" in dheaders, dheaders
    col = dheaders.index("DIN") + 1
    return {
        dws.cell(r, col).value
        for r in range(2, dws.max_row + 1)
        if dws.cell(r, col).value is not None
    }


# ══════════════════════════════════════════════════════════════════════════════
# 1. competitors  (above / below / exactly)
# ══════════════════════════════════════════════════════════════════════════════

def test_competitors_above():
    crit = parse_criteria([{"metric": "competitors", "operator": "above", "value": 1}])
    xlsx, summary, detail, warnings = build_filtered_workbook(
        _sheet1_iqvia(), _sheet2_progesterone(), crit)
    assert warnings == []
    wb = _open(xlsx)
    _assert_shell(wb)
    # Capsule has 2 marketed competitors (>1); ER form has 1 (not >1).
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule")}
    assert _summary_value(wb, "Number of Competitors") == 2
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999"}


def test_competitors_below():
    crit = parse_criteria([{"metric": "competitors", "operator": "below", "value": 2}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    # only the ER form has competitors (1) below 2.
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule (extended-release)")}
    assert _detail_dins(wb) == {"02505223"}


def test_competitors_exactly():
    crit = parse_criteria([{"metric": "competitors", "operator": "exactly", "value": 2}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule")}
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999"}


# ══════════════════════════════════════════════════════════════════════════════
# 2. filings  (GSUR / Sheet-2 driven — proves both tabs feed the screen)
# ══════════════════════════════════════════════════════════════════════════════

def test_filings_above_uses_sheet2():
    # Capsule + ER both share ingredient PROGESTERONE → both get filings == 3.
    crit = parse_criteria([{"metric": "filings", "operator": "above", "value": 2}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {
        ("PROGESTERONE", "Capsule"),
        ("PROGESTERONE", "Capsule (extended-release)"),
    }
    # filings come from Sheet 2 (3 progesterone GSUR rows; metformin excluded).
    assert _summary_value(wb, "Number of Filings") == 3
    # Detail spans only the Sheet-1 survivor DINs of BOTH qualifying products.
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999", "02505223"}


def test_filings_exactly_zero_when_no_sheet2():
    # With an empty Sheet 2, every product has filings == 0.
    crit = parse_criteria([{"metric": "filings", "operator": "exactly", "value": 0}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), pd.DataFrame(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {
        ("PROGESTERONE", "Capsule"),
        ("PROGESTERONE", "Capsule (extended-release)"),
    }
    assert _summary_value(wb, "Number of Filings") == 0


# ══════════════════════════════════════════════════════════════════════════════
# 3. approvals  (distinct companies across ALL DINs)
# ══════════════════════════════════════════════════════════════════════════════

def test_approvals_above():
    # Capsule has 3 distinct companies (incl. cancelled); ER has 1.
    crit = parse_criteria([{"metric": "approvals", "operator": "above", "value": 2}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule")}
    assert _summary_value(wb, "Number of Approvals") == 3
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999"}


# ══════════════════════════════════════════════════════════════════════════════
# 4-6. value / quantity / quantity_ext  (golden IQVIA anchored sums)
# ══════════════════════════════════════════════════════════════════════════════

def test_value_above_golden_sum():
    # Capsule aggregates to the golden $34,220,946; ER is $500.
    crit = parse_criteria([{"metric": "value", "operator": "above", "value": 30_000_000}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule")}
    assert _summary_value(wb, "Value Sizeable ($)") == _GOLD_VALUE
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999"}


def test_quantity_above_golden_sum():
    crit = parse_criteria([{"metric": "quantity", "operator": "above", "value": 400_000}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule")}
    assert _summary_value(wb, "Quantity Sizeable (Units)") == _GOLD_UNITS
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999"}


def test_quantity_ext_above_golden_sum():
    crit = parse_criteria([{"metric": "quantity_ext", "operator": "above", "value": 100}])
    xlsx, *_ = build_filtered_workbook(_sheet1_iqvia(), _sheet2_progesterone(), crit)
    wb = _open(xlsx)
    _assert_shell(wb)
    # Capsule ext = 300 (>100); ER ext = 1 (not >100).
    assert _summary_products(wb) == {("PROGESTERONE", "Capsule")}
    assert _summary_value(wb, "Quantity Ext Sizeable") == _GOLD_EXT
    assert _detail_dins(wb) == {"02516187", "02493578", "09999999"}


# ══════════════════════════════════════════════════════════════════════════════
# 7. dosage_form  (base form selects all raw variants incl. a multi-form cell)
# ══════════════════════════════════════════════════════════════════════════════

def _sheet1_dosage() -> pd.DataFrame:
    """One product per distinct raw form so the filter's reach is unambiguous,
    plus a multi-form 'KIT; TABLET' cell that must match the TABLET base."""
    return pd.DataFrame([
        {"din": "1", "ingredient": "DRUGA", "dosage_form": "TABLET",
         "company": "C1", "status": "marketed"},
        {"din": "2", "ingredient": "DRUGB", "dosage_form": "TABLET (EXTENDED-RELEASE)",
         "company": "C2", "status": "marketed"},
        {"din": "3", "ingredient": "DRUGC", "dosage_form": "TABLET (CHEWABLE)",
         "company": "C3", "status": "marketed"},
        {"din": "4", "ingredient": "DRUGD", "dosage_form": "CAPSULE",
         "company": "C4", "status": "marketed"},
        {"din": "5", "ingredient": "DRUGE", "dosage_form": "KIT; TABLET",
         "company": "C5", "status": "marketed"},
    ])


def test_dosage_form_base_selects_all_raw_variants():
    bases = parse_dosage_forms([{"metric": "dosage_form", "value": ["TABLET"]}])
    xlsx, *_ = build_filtered_workbook(
        _sheet1_dosage(), pd.DataFrame(), [], dosage_bases=bases)
    wb = _open(xlsx)
    _assert_shell(wb)
    # plain + extended-release + chewable + the multi-form KIT;TABLET — all TABLET.
    assert _summary_products(wb) == {
        ("DRUGA", "TABLET"),
        ("DRUGB", "TABLET (EXTENDED-RELEASE)"),
        ("DRUGC", "TABLET (CHEWABLE)"),
        ("DRUGE", "KIT; TABLET"),
    }
    # CAPSULE (din 4) must not leak into Detail; KIT;TABLET (din 5) is included.
    assert _detail_dins(wb) == {"1", "2", "3", "5"}


def test_dosage_form_nonmatching_base_is_empty():
    bases = parse_dosage_forms([{"metric": "dosage_form", "value": ["LOTION"]}])
    xlsx, summary, detail, _ = build_filtered_workbook(
        _sheet1_dosage(), pd.DataFrame(), [], dosage_bases=bases)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == set()
    assert wb["Summary"].max_row == 1   # header only
    assert _detail_dins(wb) == set()


# ══════════════════════════════════════════════════════════════════════════════
# 8. no_file_date  (greater excludes blanks; less includes blanks)
# ══════════════════════════════════════════════════════════════════════════════

def _sheet1_dates() -> pd.DataFrame:
    """Four single-DIN products: 2027, 2030, N/A (blank), "" (blank)."""
    return pd.DataFrame([
        {"din": "1", "ingredient": "A", "dosage_form": "TABLET", "company": "C1",
         "status": "marketed", "dp_6yr_no_file_date": "2027-01-01"},
        {"din": "2", "ingredient": "B", "dosage_form": "TABLET", "company": "C2",
         "status": "marketed", "dp_6yr_no_file_date": "2030-01-01"},
        {"din": "3", "ingredient": "C", "dosage_form": "TABLET", "company": "C3",
         "status": "marketed", "dp_6yr_no_file_date": "N/A"},
        {"din": "4", "ingredient": "D", "dosage_form": "TABLET", "company": "C4",
         "status": "marketed", "dp_6yr_no_file_date": ""},
    ])


def test_no_file_date_greater_excludes_blanks():
    # greater than 01/01/2028 → only the 2030 product; blanks (C, D) excluded.
    date_filter = parse_no_file_date(
        [{"metric": "no_file_date", "operator": "greater", "value": "01/01/2028"}],
        today=_TODAY)
    xlsx, *_ = build_filtered_workbook(
        _sheet1_dates(), pd.DataFrame(), [], date_filter=date_filter)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {("B", "TABLET")}
    assert _detail_dins(wb) == {"2"}


def test_no_file_date_less_includes_blanks():
    # less than 01/01/2028 → 2027 (A) AND both blanks (C, D); 2030 (B) excluded.
    date_filter = parse_no_file_date(
        [{"metric": "no_file_date", "operator": "less", "value": "01/01/2028"}],
        today=_TODAY)
    xlsx, *_ = build_filtered_workbook(
        _sheet1_dates(), pd.DataFrame(), [], date_filter=date_filter)
    wb = _open(xlsx)
    _assert_shell(wb)
    assert _summary_products(wb) == {("A", "TABLET"), ("C", "TABLET"), ("D", "TABLET")}
    assert _detail_dins(wb) == {"1", "3", "4"}
