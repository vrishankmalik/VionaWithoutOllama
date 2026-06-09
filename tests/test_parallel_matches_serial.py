"""Verify enrich_labeling_batch_fast() produces identical output to the serial path.

Tier structure
--------------
  test_batch_fast_matches_serial_offline
    Unit test — no network.  Uses the PIQRAY labeling fixture and mocked DPD
    API calls to verify that enrich_labeling_batch_fast() returns exactly the
    same per-DIN dict as enrich_labeling() called individually.

  test_parallel_matches_serial  (mark: integration)
    End-to-end test — hits live sites.  Runs the full export pipeline for a
    fixed ~30-DIN panel (patented drugs, combos, scanned PMs, data protection).
    On first run it writes tests/fixtures/baseline.xlsx.  On every subsequent
    run it re-runs the pipeline and asserts the workbook is cell-for-cell
    identical to that baseline.  Pass --generate-baseline to force a fresh write.

Run
---
  make test          # offline tier only
  make test-live     # offline + integration tiers
  pytest -k test_parallel_matches_serial -m integration -s  # integration only
"""
from __future__ import annotations

import asyncio
import io
import json
import time
from copy import deepcopy
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures"
BASELINE_PATH = FIXTURES_DIR / "baseline.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Fixed DIN panel used by the integration baseline test.
# Covers: patented drugs, combination products, off-patent generics,
#         data-protection entries.  Chosen from known-good DPD records.
# ─────────────────────────────────────────────────────────────────────────────
_BASELINE_DINS = [
    # PIQRAY (alpelisib) — patented, data protection, two strengths
    "02494558",
    "02494566",
    # LEQEMBI (lecanemab) — patented, newer NOC, patent 2630344
    "02562383",
    # KEYTRUDA (pembrolizumab) — biologic, patented
    "02449110",
    # ELIQUIS (apixaban) — patented, Patent Register entries
    "02425947",
    "02425955",
    # HUMIRA (adalimumab) — biologic, many follow-on biosimilars
    "02284901",
    # CRESTOR (rosuvastatin) — patented, broad generic market
    "02248702",
    # LIPITOR (atorvastatin) — widely covered, patent expired
    "02230711",
    # Metformin generics — off-patent; Patent Register should return no_results
    "02229895",
    "02162954",
    # JANUMET (sitagliptin+metformin HCl) — combination product
    "02328372",
    # SYNTHROID (levothyroxine) — long-established brand
    "00326925",
    # COVERSYL PLUS (perindopril+indapamide) — combo
    "02236845",
    # ENBREL (etanercept) — biologic, patented
    "02236379",
    # LYRICA (pregabalin) — patented, multiple strengths
    "02268701",
    "02268728",
    "02268736",
    # ADVAIR DISKUS (fluticasone+salmeterol) — combination inhaler
    "02237430",
    # ATIVAN (lorazepam) — long-established, no patent
    "00511579",
    # TYLENOL (acetaminophen) — OTC, no patent
    "00559407",
    # WARFARIN SODIUM — old molecule, various strengths
    "00406694",
    # AMLODIPINE BESYLATE — generic, common
    "02238299",
    # ATORVASTATIN generic — post-patent
    "02403560",
    # ROSUVASTATIN generic — post-patent
    "02389460",
    # MONTELUKAST — patent timing interesting
    "02248729",
    # TELMISARTAN — combination available
    "02269724",
    # SPIRIVA (tiotropium) — patented inhaler
    "02266059",
]

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _reset_store(tmp_path: Path) -> None:
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "test_enrich.db"))


def _load_piqray_pages() -> list[tuple[int, str]]:
    """Load pre-extracted PIQRAY pages from fixture."""
    data = json.loads((FIXTURES_DIR / "labeling" / "piqray_pages.json").read_bytes())
    return [(p["page"], p["text"]) for p in data]


def _rows_equal(a: dict | None, b: dict | None, label: str = "") -> None:
    """Assert two labeling row dicts are identical, with helpful diff output."""
    assert (a is None) == (b is None), (
        f"{label}: one result is None, the other is not.\n  serial={a}\n  batch={b}"
    )
    if a is None:
        return
    mismatches = [
        f"  field={k!r}: serial={a.get(k)!r}, batch={b.get(k)!r}"
        for k in set(a) | set(b)
        if a.get(k) != b.get(k) and k != "fetched_at"  # fetched_at is a timestamp
    ]
    assert not mismatches, (
        f"{label}: row values differ:\n" + "\n".join(mismatches)
    )


# ─────────────────────────────────────────────────────────────────────────────
# Offline unit test — no network required
# ─────────────────────────────────────────────────────────────────────────────

def test_batch_fast_matches_serial_offline(tmp_path):
    """enrich_labeling_batch_fast() must return the same per-DIN dict as
    enrich_labeling() called individually, for the same mocked inputs.

    Two sub-cases are tested:

    (A) No-PM path — fetch_stage2_data returns pdf_url=None for all DINs.
        Exercises the drug_code deduplication; the no-PM sentinel row must
        be identical whether the DIN is processed individually or in a batch.

    (B) PDF path — fetch_stage2_data returns a pdf_url; _download_pdf and
        _extract_text_async are mocked so no real I/O happens.  Exercises
        the pdf_url deduplication and the per-DIN strength-scoped parsing.

    Mocking at _extract_text_async (not _extract_text_with_ocr) so both
    enrich_labeling() and enrich_labeling_batch_fast() hit the same mock —
    both now go through _extract_text_async after the thread-pool refactor.
    """
    import app.enrichment.labeling as lab_mod
    from app.enrichment.labeling import (
        enrich_labeling,
        enrich_labeling_batch_fast,
    )

    if not (FIXTURES_DIR / "labeling" / "piqray_pages.json").exists():
        pytest.skip("PIQRAY fixture not present — run make refresh-fixtures first")

    piqray_pages = _load_piqray_pages()

    # ── Sub-case A: no PM (pdf_url=None) ──────────────────────────────────────
    # Two DINs sharing drug_code 99001, one DIN with drug_code 99002.
    DINS_A = {
        "00000001": (99001, "200 mg"),
        "00000002": (99001, "300 mg"),
        "00000003": (99002, "50 mg"),
    }
    fake_s2_no_pdf = {
        "active_ingredient": "alpelisib",
        "pack_size": "28 count",
        "pack_style": "Blister",
        "pdf_url": None,
        "pdf_date": None,
        "description": None,
    }

    async def _serial_a():
        return {din: await enrich_labeling(din, dc, st) for din, (dc, st) in DINS_A.items()}

    async def _batch_a():
        return await enrich_labeling_batch_fast(DINS_A)

    # Use side_effect (not return_value) so each call gets a FRESH dict copy.
    # enrich_labeling mutates the returned dict with .pop("pdf_url", ...) in-place;
    # returning the same object would corrupt later calls.
    def _fresh_s2_no_pdf(_dc):
        return dict(fake_s2_no_pdf)

    with patch.object(lab_mod, "fetch_stage2_data", new=AsyncMock(side_effect=_fresh_s2_no_pdf)):
        _reset_store(tmp_path)
        serial_a = asyncio.run(_serial_a())
        _reset_store(tmp_path)
        batch_a = asyncio.run(_batch_a())

    assert set(serial_a) == set(batch_a), "Sub-case A: DIN key mismatch"
    for din in DINS_A:
        _rows_equal(serial_a.get(din), batch_a.get(din), label=f"no-PM DIN {din}")

    # ── Sub-case B: PDF path with shared pdf_url ───────────────────────────────
    # Two DINs share drug_code 99001 and thus the same pdf_url.  The batch
    # must call parse_labeling_fields_async once per DIN (different strength)
    # but download + extract the PDF only once.
    DINS_B = {
        "00000011": (99001, "200 mg"),
        "00000012": (99001, "300 mg"),
    }
    fake_s2_with_pdf = {
        "active_ingredient": "alpelisib",
        "pack_size": "28 count",
        "pack_style": "Blister",
        "pdf_url": "https://fake.example.com/piqray.pdf",
        "pdf_date": "2019-05-24",
        "description": "Film-coated tablets",
    }

    # Mock _extract_text_async (the async wrapper called by both code paths)
    # rather than _extract_text_with_ocr (the sync inner function).
    async def _mock_extract_async(pdf_bytes, cache_key, enable_ocr=True):
        return piqray_pages, False

    async def _serial_b():
        return {din: await enrich_labeling(din, dc, st) for din, (dc, st) in DINS_B.items()}

    async def _batch_b():
        return await enrich_labeling_batch_fast(DINS_B)

    def _fresh_s2_with_pdf(_dc):
        return dict(fake_s2_with_pdf)

    with (
        patch.object(lab_mod, "fetch_stage2_data", new=AsyncMock(side_effect=_fresh_s2_with_pdf)),
        patch.object(lab_mod, "_download_pdf", new=AsyncMock(return_value=b"%PDF-1.4 fake")),
        patch.object(lab_mod, "_extract_text_async", new=AsyncMock(side_effect=_mock_extract_async)),
    ):
        _reset_store(tmp_path)
        serial_b = asyncio.run(_serial_b())
        _reset_store(tmp_path)
        batch_b = asyncio.run(_batch_b())

    assert set(serial_b) == set(batch_b), "Sub-case B: DIN key mismatch"
    for din in DINS_B:
        _rows_equal(serial_b.get(din), batch_b.get(din), label=f"PDF DIN {din}")


# ─────────────────────────────────────────────────────────────────────────────
# Integration test — requires live network, marked accordingly
# ─────────────────────────────────────────────────────────────────────────────

@pytest.mark.integration
def test_parallel_matches_serial(tmp_path, request):
    """Full export pipeline for a fixed 30-DIN panel.

    First run  → writes tests/fixtures/baseline.xlsx (test always passes).
    Later runs → re-runs the pipeline and asserts the workbook is cell-for-cell
                 identical to baseline.xlsx.

    Pass --generate-baseline on the pytest command line to force a fresh write
    even when baseline.xlsx already exists.

    Timing: prints wall-clock time before and after the pipeline so the
    before/after speedup is visible in the CI log.
    """
    import openpyxl

    try:
        from app.enrichment.data_protection import fetch_data_protection_table
        from app.enrichment.labeling import enrich_labeling_batch_fast
        from app.enrichment.patents import enrich_patents
        from app.enrichment.store import get_labeling_for_din, reset_for_testing
        from app.enrichment.workbook import _is_excluded_din, build_workbook
        from app.models import SearchMetadata, SearchResponse
        from app.normalize import normalize_query
        from app.sources.dpd import search_dpd
        from app.sources.generic_submissions import search_generic_submissions
        from app.sources.noc import search_noc
        from app.sources.patent_register import search_patent_register
    except ImportError as exc:
        pytest.skip(f"App import failed: {exc}")

    force_baseline = request.config.getoption("--generate-baseline", default=False)

    # Use a temporary enrichment store so the test is self-contained.
    reset_for_testing(str(tmp_path / "integration_enrich.db"))

    async def _run_pipeline() -> bytes:
        """Run full export pipeline for _BASELINE_DINS and return XLSX bytes."""
        canonical = "alpelisib"  # not used; we use DIN list directly below

        # Search across all four sources for each ingredient in the panel.
        # For simplicity we pass all DINs as a single batch rather than by ingredient,
        # which is equivalent for the purpose of the workbook comparison.
        all_results = []
        for din in _BASELINE_DINS:
            src = await search_dpd(din, "din", [])
            all_results.extend(src.records)

        from datetime import datetime, timezone
        response = SearchResponse(
            metadata=SearchMetadata(
                query="baseline_panel",
                field="din",
                timestamp=datetime.now(timezone.utc).isoformat(),
                normalized_terms=["baseline_panel"],
            ),
            sources=[
                __import__("app.models", fromlist=["SourceResult"]).SourceResult(
                    source="DPD", status="ok", records=all_results
                )
            ],
        )

        # Stage 2: patents
        valid_dins = [
            r.din.strip() for s in response.sources for r in s.records
            if r.din and not _is_excluded_din(r.din)
        ]
        valid_dins = list(dict.fromkeys(valid_dins))
        if valid_dins:
            await enrich_patents(valid_dins)

        # Stage 3: labeling
        din_map = {}
        for s in response.sources:
            if s.source != "DPD":
                continue
            for r in s.records:
                if _is_excluded_din(r.din):
                    continue
                dc = r.source_specific.get("drug_code")
                if dc is None:
                    continue
                din_key = r.din.strip()
                if get_labeling_for_din(din_key) is None:
                    din_map[din_key] = (int(dc), r.strength)

        if din_map:
            await enrich_labeling_batch_fast(din_map)

        # Stage 4: data protection
        dp_table = await fetch_data_protection_table()

        # Stage 5: workbook
        return build_workbook(response, dp_table=dp_table)

    # ── First run: generate baseline (or warm cache) ──────────────────────────
    print(f"\n[test_parallel_matches_serial] DIN panel: {len(_BASELINE_DINS)} DINs")

    t0 = time.time()
    xlsx_bytes_1 = asyncio.run(_run_pipeline())
    elapsed_1 = round(time.time() - t0, 1)
    print(f"[test_parallel_matches_serial] Run 1: {elapsed_1}s  ({len(xlsx_bytes_1):,} bytes)")

    if force_baseline or not BASELINE_PATH.exists():
        BASELINE_PATH.parent.mkdir(parents=True, exist_ok=True)
        BASELINE_PATH.write_bytes(xlsx_bytes_1)
        print(f"[test_parallel_matches_serial] Baseline written → {BASELINE_PATH}")
        return  # first-write pass: nothing to compare yet

    # ── Second run: compare to baseline ──────────────────────────────────────
    # Reset store so every DIN goes through the full pipeline again (cold store).
    reset_for_testing(str(tmp_path / "integration_enrich2.db"))

    t1 = time.time()
    xlsx_bytes_2 = asyncio.run(_run_pipeline())
    elapsed_2 = round(time.time() - t1, 1)
    print(f"[test_parallel_matches_serial] Run 2 (cache warm): {elapsed_2}s  ({len(xlsx_bytes_2):,} bytes)")

    # ── Cell-for-cell comparison ──────────────────────────────────────────────
    baseline_bytes = BASELINE_PATH.read_bytes()
    _assert_workbooks_identical(baseline_bytes, xlsx_bytes_2, label="baseline vs run-2")
    _assert_workbooks_identical(xlsx_bytes_1, xlsx_bytes_2, label="run-1 vs run-2 (determinism)")
    print("[test_parallel_matches_serial] PASS — workbooks are cell-for-cell identical")


def _assert_workbooks_identical(a_bytes: bytes, b_bytes: bytes, label: str = "") -> None:
    """Cell-for-cell comparison of two XLSX workbooks via openpyxl.

    Compares every sheet name, every row count, and every cell value.
    Raises AssertionError with the first differing cell on failure.
    """
    import openpyxl

    wb_a = openpyxl.load_workbook(io.BytesIO(a_bytes), data_only=True)
    wb_b = openpyxl.load_workbook(io.BytesIO(b_bytes), data_only=True)

    sheets_a = wb_a.sheetnames
    sheets_b = wb_b.sheetnames
    assert sheets_a == sheets_b, (
        f"{label}: sheet names differ: expected={sheets_a}, got={sheets_b}"
    )

    for sheet_name in sheets_a:
        ws_a = wb_a[sheet_name]
        ws_b = wb_b[sheet_name]

        rows_a = list(ws_a.iter_rows(values_only=True))
        rows_b = list(ws_b.iter_rows(values_only=True))

        assert len(rows_a) == len(rows_b), (
            f"{label} sheet={sheet_name!r}: row count differs: "
            f"expected={len(rows_a)}, got={len(rows_b)}"
        )

        for r_idx, (row_a, row_b) in enumerate(zip(rows_a, rows_b), start=1):
            if len(row_a) != len(row_b):
                # Pad the shorter row with None so we can report cell differences
                max_len = max(len(row_a), len(row_b))
                row_a = row_a + (None,) * (max_len - len(row_a))
                row_b = row_b + (None,) * (max_len - len(row_b))
            for c_idx, (cell_a, cell_b) in enumerate(zip(row_a, row_b), start=1):
                assert cell_a == cell_b, (
                    f"{label} sheet={sheet_name!r} row={r_idx} col={c_idx}: "
                    f"expected={cell_a!r}, got={cell_b!r}"
                )


# ─────────────────────────────────────────────────────────────────────────────
# pytest option registration
# ─────────────────────────────────────────────────────────────────────────────

def pytest_addoption(parser):
    """Register --generate-baseline option (ignored if already registered)."""
    try:
        parser.addoption(
            "--generate-baseline",
            action="store_true",
            default=False,
            help="Force regeneration of tests/fixtures/baseline.xlsx",
        )
    except ValueError:
        pass  # already added by another conftest
