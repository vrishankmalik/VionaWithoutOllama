"""Acceptance tests for OPTION 1 — synchronous export path.

Covers the synchronous single-product `/api/export` endpoint plus the
orchestration-level multi-product workbook build (`build_workbook_multiproduct`)
that the async job reuses.

Hermetic: respx-mocked sources (metformin/GLUCOPHAGE golden DIN 02229895, code
99999), enrichment store redirected to tmp_path, heavy enrichment monkeypatched
to no-op async stubs.  No live network, no shared on-disk state.

Mirrors the passing E2E setup in tests/test_build_workbook.py
(`test_export_allow_partial_builds_with_warning`):
  - fixtures: mock_noc, mock_dpd, mock_gsur, mock_patent_register
  - query:    q=metformin&field=ingredient
  - patches:  app.main.enrich_patents / enrich_labeling_batch as AsyncMock no-ops
"""
from __future__ import annotations

import io
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest

TWO_SHEETS = {"DPD + NOC + Patents", "Generic Submissions"}
XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── shared enrichment-store reset (tmp_path, parallel-safe) ───────────────────

@pytest.fixture
def fresh_store(tmp_path):
    """Redirect the enrichment SQLite store at a per-test tmp db and reset after."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    yield store_mod
    # Teardown: detach from the tmp db so the next test/worker starts clean.
    store_mod.reset_for_testing(str(tmp_path / "enrich_after.db"))


# ── strategy A: HTTP-level (TestClient) ──────────────────────────────────────

def _client():
    from fastapi.testclient import TestClient
    from app.main import app
    return TestClient(app)


def test_export_happy_path_returns_xlsx(
    fresh_store, mock_noc, mock_dpd, mock_gsur, mock_patent_register
):
    """/api/export happy path → 200, xlsx content-type, attachment, two canonical tabs."""
    with patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        resp = _client().get("/api/export?q=metformin&field=ingredient")

    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == XLSX_MEDIA
    cd = resp.headers.get("content-disposition", "")
    assert "attachment" in cd and ".xlsx" in cd

    # openpyxl-readable + exactly the two canonical sheet names.
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert set(wb.sheetnames) == TWO_SHEETS, wb.sheetnames


def test_export_sheet1_din_sorted_ascending(
    fresh_store, mock_noc, mock_dpd, mock_gsur, mock_patent_register
):
    """Sheet 1 DINs must be sorted ascending in the rendered XLSX."""
    with patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        resp = _client().get("/api/export?q=metformin&field=ingredient")
    assert resp.status_code == 200, resp.text

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["DPD + NOC + Patents"]
    headers = [c.value for c in ws[1]]
    assert "DIN" in headers, headers
    din_col = headers.index("DIN")
    dins = [
        str(row[din_col].value)
        for row in ws.iter_rows(min_row=2)
        if row[din_col].value not in (None, "")
    ]
    assert dins == sorted(dins), f"Sheet 1 DINs not ascending: {dins}"


def test_export_forced_source_error_409_then_allow_partial_200(
    fresh_store, mock_noc, mock_dpd, mock_gsur, mock_patent_register
):
    """A forced NOC error → 409 without allow_partial; 200 with allow_partial=true."""
    from app.models import SourceResult
    forced = SourceResult(source="NOC", status="error", error_message="forced test error")

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = _client()

        r409 = client.get("/api/export?q=metformin&field=ingredient")
        assert r409.status_code == 409, r409.text
        detail = r409.json().get("detail", "")
        assert "NOC" in detail and "allow_partial" in detail

        r200 = client.get("/api/export?q=metformin&field=ingredient&allow_partial=true")
        assert r200.status_code == 200, r200.text
        assert r200.headers["content-type"] == XLSX_MEDIA
        wb = openpyxl.load_workbook(io.BytesIO(r200.content))
        # allow_partial surfaces a Source Status warning sheet alongside the two tabs.
        assert any("Status" in n for n in wb.sheetnames), wb.sheetnames


def test_export_missing_q_is_422(
    fresh_store, mock_noc, mock_dpd, mock_gsur, mock_patent_register
):
    """q is a required query param → FastAPI validation 422 when omitted."""
    resp = _client().get("/api/export?field=ingredient")
    assert resp.status_code == 422, resp.text


# ── strategy B: orchestration-level (build_workbook_multiproduct) ────────────
#
# The sync /api/export path uses build_workbook; the async job uses
# build_workbook_multiproduct.  Both share build_sheet1/build_sheet2.  Here we
# drive the multiproduct builder directly with synthetic responses to assert the
# single- and multi-product structural contracts deterministically.

from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult


def _dpd(din: str, ingredient: str, brand: str = "BRAND") -> DrugRecord:
    return DrugRecord(
        source="DPD", din=din, brand_name=brand, company="Acme",
        ingredient=ingredient, strength="500 mg", dosage_form="Tablet",
        all_ingredients=[ingredient],
    )


def _noc(din: str, ingredient: str) -> DrugRecord:
    return DrugRecord(
        source="NOC", din=din, brand_name="BRAND", company="Acme",
        ingredient=ingredient,
        source_specific={
            "noc_date": "2019-01-01", "submission_type": "NDS",
            "submission_class": "New", "reason_for_supplement": None,
            "therapeutic_class": "Test",
        },
    )


def _response(ingredient: str, dins: list[str]) -> SearchResponse:
    dpd = [_dpd(d, ingredient) for d in dins]
    noc = [_noc(d, ingredient) for d in dins]
    return SearchResponse(
        metadata=SearchMetadata(
            query=ingredient, field="ingredient",
            timestamp="2026-01-01T00:00:00+00:00",
        ),
        sources=[
            SourceResult(source="DPD", status="ok", records=dpd),
            SourceResult(source="NOC", status="ok", records=noc),
            SourceResult(source="GenericSubmissions", status="no_results"),
        ],
    )


def test_multiproduct_single_block_two_tabs(fresh_store):
    """Single-product build → readable XLSX with exactly the two canonical tabs."""
    from app.enrichment.workbook import build_workbook_multiproduct

    resp = _response("metformin", ["02229895", "02229896"])
    xlsx, sheet1_df, sheet2_df, _recon = build_workbook_multiproduct([("metformin", resp)])

    assert xlsx[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert set(wb.sheetnames) == TWO_SHEETS, wb.sheetnames
    # Sheet 1 holds DPD∩NOC rows for both DINs.
    dins = set(sheet1_df["din"].astype(str))
    assert {"02229895", "02229896"} <= dins, dins


def test_multiproduct_side_by_side_blocks(fresh_store):
    """queries=[a,b] → both products present as stacked blocks (vertical layout).

    The combined Sheet 1 distinguishes products by the leading ``ingredient_name``
    column (and row fill color), NOT a separate 'product' column — products are
    vertically concatenated with ``ingredient_name`` always first.
    """
    from app.enrichment.workbook import build_workbook_multiproduct

    resp_a = _response("metformin", ["02229895"])
    resp_b = _response("alpelisib", ["02498014"])
    xlsx, sheet1_df, sheet2_df, _recon = build_workbook_multiproduct(
        [("metformin", resp_a), ("alpelisib", resp_b)]
    )

    assert xlsx[:2] == b"PK"
    # The ingredient_name column is the per-row product discriminator.
    assert "ingredient_name" in sheet1_df.columns, sheet1_df.columns.tolist()
    products = set(sheet1_df["ingredient_name"].astype(str).str.lower())
    assert {"metformin", "alpelisib"} <= products, products
    # Each product's DIN is present.
    dins = set(sheet1_df["din"].astype(str))
    assert "02229895" in dins and "02498014" in dins, dins


def test_multiproduct_excludes_na_din_rows(fresh_store):
    """NOC rows whose DIN is blank/'N/A'/'Not Applicable' never reach Sheet 1."""
    from app.enrichment.workbook import build_workbook_multiproduct

    resp = _response("metformin", ["02229895"])
    # Inject junk NOC rows that must be excluded.
    for bad in ("", "N/A", "Not Applicable"):
        resp.sources[1].records.append(_noc(bad, "metformin"))

    _xlsx, sheet1_df, _s2, _recon = build_workbook_multiproduct([("metformin", resp)])
    dins = set(sheet1_df["din"].astype(str).str.strip())
    assert "" not in dins and "N/A" not in dins and "Not Applicable" not in dins
    assert "02229895" in dins
