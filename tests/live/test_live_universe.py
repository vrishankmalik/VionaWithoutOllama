"""Part 1 — Live: full-universe build (option 3), dosage-forms endpoint, option 4.

  * Option 3: the real allfiles.zip builds the whole catalogue (~13.5k products)
    and the no-PDF workbook assembles with its three documented sheets.
  * /api/dosage-forms returns ~55 live base forms (band-checked for churn).
  * Option 4: filter→enrich runs end-to-end over the live universe and emits a
    valid enriched workbook.  A deliberately narrow (far-future no-file-date)
    filter keeps the live survivor PM-PDF fetch tiny so the orchestration is
    verified without a multi-minute fetch (bulk PDF volume is a Part-2 concern).
"""
from __future__ import annotations

import asyncio
import io
from pathlib import Path

import openpyxl
import pytest

from app.enrichment import universe as U
from app.enrichment.universe import (
    build_universe_response,
    build_universe_sheet1,
    build_universe_sheet2,
    build_universe_workbook,
)
from app.jobs import JobState

pytestmark = pytest.mark.integration


# ── Option 3: full no-PDF universe ────────────────────────────────────────────

def test_option3_full_catalogue_builds(live_universe):
    recs = live_universe.dpd_records
    assert len(recs) > 10000, f"only {len(recs)} universe products — parse regression?"
    dins = {r.din for r in recs}
    assert len(dins) == len(recs), "every universe record must carry a unique DIN"


def test_option3_workbook_assembles_three_sheets(live_universe):
    resp = build_universe_response(live_universe)
    s1, recon, low = build_universe_sheet1(resp, None, dp_table=live_universe.dp_table)
    s2 = build_universe_sheet2(live_universe)
    xlsx = build_universe_workbook(s1, s2, recon, low)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == ["⚠ Read Me", "Full Universe (no PDF)", "Generic Submissions"], wb.sheetnames
    # The data sheet carries the full catalogue.
    assert wb["Full Universe (no PDF)"].max_row - 1 > 10000


# ── /api/dosage-forms ─────────────────────────────────────────────────────────

def test_dosage_forms_endpoint_returns_live_bases(live_universe):
    """The endpoint serves ~55 base forms off the live build (band-checked)."""
    from fastapi.testclient import TestClient
    from app.main import app

    # live_universe is already cached in U._CACHE, so the endpoint reuses it.
    with TestClient(app) as client:
        data = client.get("/api/dosage-forms").json()
    bases = data["base_forms"]
    assert isinstance(bases, list) and bases == sorted(bases)
    assert 40 <= len(bases) <= 80, f"expected ~55 base forms, got {len(bases)}: {bases[:10]}…"


# ── Option 4: filter → enrich survivors → enriched xlsx ───────────────────────

def test_option4_filter_enrich_produces_valid_xlsx(live_universe):
    """End-to-end option-4 over the live universe with a narrow filter.

    A far-future six-year-no-file-date filter yields a tiny survivor set, so the
    live PM-PDF enrichment stays small while the full orchestration + workbook
    assembly are exercised against real data.
    """
    from app.universe_job import run_universe_filter_enrich_job

    # get_universe() inside the job will reuse the cached live bundle.
    job = JobState(job_id="live-opt4", query="Full universe", field="ingredient")
    job.filter_criteria = [
        {"metric": "no_file_date", "operator": "greater_or_equal", "value": "01/01/2099"}
    ]
    asyncio.run(run_universe_filter_enrich_job(job, enable_ocr=False))

    assert job.status == "complete", f"option-4 job failed: {job.error}"
    assert job.result_path and Path(job.result_path).exists()
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert len(wb.sheetnames) >= 2, f"enriched workbook should have ≥2 sheets: {wb.sheetnames}"
    # The Summary snapshot is always populated for a filtered job (even if 0 survivors).
    assert job.summary_columns, "filtered job must expose a Summary snapshot"
    # A terminal complete event with a download URL was emitted.
    completes = [e for e in job.events if e.get("status") == "complete"]
    assert len(completes) == 1 and completes[0]["download_url"] == f"/export/result/{job.job_id}"
