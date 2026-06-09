"""Tests for app/enrichment/workbook.py.

Three structural tests per the build spec:
  1. NOC rows with DIN = "Not Applicable" / blank are excluded from Sheet 1.
  2. Sheet 1 rows are sorted by DIN ascending.
  3. Sheet 2 (Generic Submissions) is standalone — never joined to Sheet 1.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import pytest

from app.models import DrugRecord, SearchMetadata, SearchResponse, SourceResult


# ── shared fixture builder ────────────────────────────────────────────────────

def _make_response(
    dpd_records: list[DrugRecord] | None = None,
    noc_records: list[DrugRecord] | None = None,
    gsur_records: list[DrugRecord] | None = None,
    query: str = "alpelisib",
) -> SearchResponse:
    sources = []
    if dpd_records is not None:
        sources.append(SourceResult(source="DPD", status="ok", records=dpd_records))
    if noc_records is not None:
        sources.append(SourceResult(source="NOC", status="ok", records=noc_records))
    if gsur_records is not None:
        sources.append(SourceResult(source="GenericSubmissions", status="ok", records=gsur_records))
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp=datetime.now(timezone.utc).isoformat(),
        ),
        sources=sources,
    )


def _dpd(din: str, brand: str = "BRAND", strength: str = "50 mg") -> DrugRecord:
    return DrugRecord(
        source="DPD",
        din=din,
        brand_name=brand,
        company="Novartis",
        ingredient="alpelisib",
        strength=strength,
        all_ingredients=["alpelisib"],
    )


def _noc(din: str, brand: str = "BRAND", submission_type: str = "NDS") -> DrugRecord:
    return DrugRecord(
        source="NOC",
        din=din,
        brand_name=brand,
        company="Novartis",
        ingredient="alpelisib",
        source_specific={"noc_date": "2019-05-24", "submission_type": submission_type},
    )


def _gsur(ingredient: str = "alpelisib") -> DrugRecord:
    return DrugRecord(
        source="GenericSubmissions",
        ingredient=ingredient,
        company="GenericCo",
        source_specific={"therapeutic_area": "Oncology", "date_accepted": "2022/06"},
    )


# ── Test 1: NOC "Not Applicable" DINs excluded from Sheet 1 ──────────────────

def test_noc_not_applicable_din_excluded():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[
            _noc("02498014"),          # valid — should appear
            _noc("Not Applicable"),    # must be excluded
            _noc(""),                  # blank — must be excluded
            _noc("N/A"),               # N/A — must be excluded
        ],
    )
    df = build_sheet1(response)

    assert not df.empty, "Sheet 1 should have rows"
    dins_in_sheet = set(df["din"].astype(str).str.strip())
    # "Not Applicable", "", "N/A" must not appear
    assert "Not Applicable" not in dins_in_sheet
    assert "" not in dins_in_sheet
    assert "N/A" not in dins_in_sheet
    # Valid DIN from both NOC and DPD should be present
    assert "02498014" in dins_in_sheet


def test_noc_not_applicable_only_gives_empty_sheet1():
    """If all NOC rows have N/A DINs and there are no DPD records, Sheet 1 is empty."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        noc_records=[_noc("Not Applicable"), _noc("N/A")],
    )
    df = build_sheet1(response)
    assert df.empty or len(df) == 0


# ── Test 2: Sheet 1 rows sorted by DIN ascending ─────────────────────────────

def test_sheet1_din_sorted_ascending():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[
            _dpd("02498022"),
            _dpd("00012345"),
            _dpd("02498014"),
        ],
    )
    df = build_sheet1(response)
    dins = list(df["din"].astype(str))
    assert dins == sorted(dins), f"DINs should be sorted ascending, got: {dins}"


def test_sheet1_single_row_still_sorted():
    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    assert len(df) == 1
    assert df.iloc[0]["din"] == "02498014"


# ── Test 3: Sheet 2 is standalone — no DIN column, only GSUR data ────────────

def test_sheet2_is_standalone_no_din_column():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        gsur_records=[_gsur("alpelisib"), _gsur("alpelisib hydrochloride")],
    )
    df = build_sheet2(response)
    assert "din" not in df.columns, "Sheet 2 must not have a DIN column"
    assert "medicinal_ingredient" in df.columns


def test_sheet2_filtered_to_queried_ingredient():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(
        gsur_records=[
            _gsur("alpelisib"),
            _gsur("alpelisib hydrochloride"),
            _gsur("metformin"),          # unrelated — must NOT appear
            _gsur("metformin hydrochloride"),  # unrelated — must NOT appear
        ],
        query="alpelisib",
    )
    df = build_sheet2(response)
    ings = list(df["medicinal_ingredient"].str.lower())
    assert all("alpelisib" in i for i in ings), (
        f"Sheet 2 should only contain alpelisib submissions, got: {ings}"
    )
    assert not any("metformin" in i for i in ings), "metformin rows must be filtered out"


def test_sheet2_empty_when_no_gsur_source():
    from app.enrichment.workbook import build_sheet2

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet2(response)
    assert len(df) == 0


# ── Test 4: build_workbook produces valid XLSX bytes ─────────────────────────

def test_build_workbook_returns_xlsx():
    from app.enrichment.workbook import build_workbook

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],
        gsur_records=[_gsur()],
    )
    xlsx = build_workbook(response)

    # Should be a valid ZIP/XLSX (starts with PK magic bytes)
    assert xlsx[:2] == b"PK", "XLSX should start with PK (ZIP magic bytes)"

    # Verify both sheets exist
    import zipfile
    with zipfile.ZipFile(io.BytesIO(xlsx)) as zf:
        names = zf.namelist()
    # openpyxl encodes sheet names in xl/worksheets/
    assert any("sheet" in n.lower() for n in names), "No worksheet files found in XLSX"


def test_build_workbook_sheet1_has_patent_columns(tmp_path):
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Add a patent so the patent_1_* group isn't dropped by the all-empty cleanup
    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    # Wide patent columns must be present (patent data was added)
    assert "patent_count" in df.columns, "patent_count missing"
    assert "patent_1_number" in df.columns, "patent_1_number missing"
    assert "patent_1_filing_date" in df.columns, "patent_1_filing_date missing"
    assert "patent_1_expiry_date" in df.columns, "patent_1_expiry_date missing"
    # Old merged-string columns must be absent
    assert "patent_numbers" not in df.columns, "stale patent_numbers column present"
    assert "all_patents_detail" not in df.columns, "stale all_patents_detail column present"


def test_build_workbook_sheet1_has_labeling_columns(tmp_path):
    import time
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Add minimal labeling data so labeling columns are not dropped by the
    # all-empty cleanup (Change 2).  These sentinels are real values ("Not in PM"
    # / "No PM available"), so the columns must be kept.
    store_mod.upsert_labeling("02498014", {
        "colour": "pink", "shape": "round", "ph": "Not stated",
        "needs_ocr": 0, "has_unverified": 0, "drug_code": 99001,
        "fetched_at": time.time(),
    })

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014")])
    df = build_sheet1(response)
    for col in ("colour", "shape", "ph"):
        assert col in df.columns, f"Expected labeling column '{col}' in Sheet 1"


# ── Test 5: patent block aggregation ─────────────────────────────────────────

def test_patent_aggregation(tmp_path):
    """_aggregate_patents_wide correctly produces wide columns for multiple patents."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")
    store_mod.upsert_patent("02498014", "2900000", "2015-01-01", "2020-03-01", "2035-01-01")

    from app.enrichment.workbook import _aggregate_patents_wide
    agg = _aggregate_patents_wide("02498014", 2)

    assert agg["patent_count"] == 2
    # Both patent numbers must appear across the two groups (order is expiry-desc from store)
    patent_numbers = {agg["patent_1_number"], agg["patent_2_number"]}
    assert "2709025" in patent_numbers
    assert "2900000" in patent_numbers
    # Dates must be attached to the correct group
    for i in (1, 2):
        if agg[f"patent_{i}_number"] == "2709025":
            assert agg[f"patent_{i}_filing_date"] == "2008-12-10"
        elif agg[f"patent_{i}_number"] == "2900000":
            assert agg[f"patent_{i}_filing_date"] == "2015-01-01"


# ── Test 6: fail-loud guard — error-vs-no_results distinction ─────────────────

def _make_response_with_noc_error(query: str = "metformin") -> "SearchResponse":
    """Response where NOC has status=error, DPD has results."""
    from app.models import SourceResult
    sources = [
        SourceResult(source="DPD", status="ok", records=[_dpd("02242974")]),
        SourceResult(source="NOC", status="error", error_message="test forced error"),
        SourceResult(source="GenericSubmissions", status="no_results"),
        SourceResult(source="PatentRegister", status="no_results"),
    ]
    return SearchResponse(
        metadata=SearchMetadata(
            query=query,
            field="ingredient",
            timestamp="2026-01-01T00:00:00+00:00",
        ),
        sources=sources,
    )


def test_build_workbook_with_source_errors_adds_warning_sheet():
    """allow_partial=True path: source_errors causes a '⚠ Source Status' sheet to appear."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response_with_noc_error()
    xlsx = build_workbook(response, source_errors={"NOC": "test forced error"})

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    sheet_names = wb.sheetnames
    assert any("Status" in name for name in sheet_names), (
        f"Expected a Source Status warning sheet; got sheets: {sheet_names}"
    )


def test_build_workbook_no_source_errors_no_warning_sheet():
    """Default path (source_errors=None): no warning sheet is added."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")], noc_records=[_noc("02498014")])
    xlsx = build_workbook(response, source_errors=None)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 2, (
        f"Expected exactly 2 sheets with no errors; got: {wb.sheetnames}"
    )


def test_build_workbook_no_results_source_does_not_block():
    """A no_results source (Patent, GSUR) must not add a warning sheet."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(dpd_records=[_dpd("02498014")])
    # no_results is NOT an error — source_errors dict should be empty → no sheet
    xlsx = build_workbook(response, source_errors=None)

    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    assert len(wb.sheetnames) == 2


async def test_export_refuses_with_409_on_source_error(mock_noc, mock_dpd, mock_gsur, mock_patent_register):
    """HTTP 409 when a source is in error and allow_partial is not set."""
    from fastapi.testclient import TestClient
    from app.main import app
    from unittest.mock import AsyncMock, patch
    from app.models import SourceResult

    forced_error = SourceResult(
        source="NOC", status="error", error_message="forced test error"
    )

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced_error)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = TestClient(app)
        resp = client.get("/api/export?q=metformin&field=ingredient")
        assert resp.status_code == 409, f"Expected 409, got {resp.status_code}: {resp.text}"
        body = resp.json()
        assert "NOC" in body.get("detail", "")
        assert "allow_partial" in body.get("detail", "")


async def test_export_allow_partial_builds_with_warning(mock_noc, mock_dpd, mock_gsur, mock_patent_register):
    """allow_partial=true: builds a workbook with the source-status warning sheet."""
    from fastapi.testclient import TestClient
    import openpyxl
    from app.main import app
    from unittest.mock import AsyncMock, patch
    from app.models import SourceResult

    forced_error = SourceResult(
        source="NOC", status="error", error_message="forced test error"
    )

    with patch("app.main.search_noc", new=AsyncMock(return_value=forced_error)), \
         patch("app.main.enrich_patents", new=AsyncMock(return_value={})), \
         patch("app.main.enrich_labeling_batch", new=AsyncMock(return_value={})):
        client = TestClient(app)
        resp = client.get("/api/export?q=metformin&field=ingredient&allow_partial=true")
        assert resp.status_code == 200, f"Expected 200 with allow_partial; got {resp.status_code}"
        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert any("Status" in name for name in wb.sheetnames), (
            f"Expected Source Status sheet; got: {wb.sheetnames}"
        )


# ── Test 7: SNDS/SANDS filtering ─────────────────────────────────────────────

def test_snds_rows_excluded_from_sheet1():
    """NOC records with SNDS or SANDS submission types must be dropped from Sheet 1."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[
            _noc("02498014", submission_type="NDS"),                             # keep
            _noc("02498022", submission_type="Supplement to a New Drug Submission (SNDS)"),  # drop
        ],
    )
    df = build_sheet1(response)

    # Both DPD DINs must appear (they're DPD products)
    dins = set(df["din"].astype(str))
    assert "02498014" in dins
    assert "02498022" in dins

    # 02498014 should have real NOC data; 02498022 should have "No NOC record"
    row_nds = df[df["din"] == "02498014"].iloc[0]
    row_snds = df[df["din"] == "02498022"].iloc[0]
    assert row_nds["noc_submission_type"] == "NDS"
    assert row_snds["noc_submission_type"] == "No NOC record", (
        f"Expected 'No NOC record' for filtered SNDS row, got: {row_snds['noc_submission_type']!r}"
    )


def test_sands_rows_excluded_from_sheet1():
    """SANDS submissions must also be dropped."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498022")],
        noc_records=[
            _noc("02498022", submission_type="Supplement to an Abbreviated New Drug Submission (SANDS)"),
        ],
    )
    df = build_sheet1(response)
    row = df[df["din"] == "02498022"].iloc[0]
    assert row["noc_submission_type"] == "No NOC record"


def test_no_noc_record_din_gets_sentinel():
    """A DPD DIN with no matching NOC record gets 'No NOC record' in all noc_* columns."""
    from app.enrichment.workbook import build_sheet1

    # DPD record for 02498022 but no NOC record for it
    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022")],
        noc_records=[_noc("02498014")],  # only 02498014 has NOC
    )
    df = build_sheet1(response)

    row_with_noc = df[df["din"] == "02498014"].iloc[0]
    row_no_noc = df[df["din"] == "02498022"].iloc[0]

    # Row with NOC match should have real data
    assert row_with_noc["noc_submission_type"] == "NDS"
    assert row_with_noc["noc_date"] == "2019-05-24"

    # Row without NOC match should have sentinel in every noc_* column
    for col in ("noc_brand_name", "noc_company", "noc_date",
                "noc_submission_type", "noc_therapeutic_class"):
        assert row_no_noc[col] == "No NOC record", (
            f"Expected 'No NOC record' in column '{col}', got: {row_no_noc[col]!r}"
        )


def test_noc_only_column_values_after_filtering():
    """After SNDS/SANDS filtering, only NDS, ANDS, and 'No NOC record' appear in noc_submission_type."""
    from app.enrichment.workbook import build_sheet1

    response = _make_response(
        dpd_records=[_dpd("02498014"), _dpd("02498022"), _dpd("02498030")],
        noc_records=[
            _noc("02498014", submission_type="NDS"),
            _noc("02498022", submission_type="ANDS"),
            _noc("02498030", submission_type="Supplement to a New Drug Submission (SNDS)"),
        ],
    )
    df = build_sheet1(response)

    allowed = {"NDS", "ANDS", "No NOC record"}
    for val in df["noc_submission_type"].dropna():
        assert str(val) in allowed, (
            f"Unexpected noc_submission_type in Sheet 1: {val!r}"
        )


# ── Test 8: No PM available sentinel ─────────────────────────────────────────

def test_no_pm_available_constant_exported():
    """NO_PM_AVAILABLE constant must be importable and distinct from NOT_IN_PM."""
    from app.enrichment.labeling import NO_PM_AVAILABLE, NOT_IN_PM
    assert NO_PM_AVAILABLE != NOT_IN_PM
    assert NO_PM_AVAILABLE == "No PM available"


# ── Test 9: Change 2 — all-empty column cleanup ──────────────────────────────

def test_empty_patent_groups_dropped(tmp_path):
    """When no DIN has patents, all patent_N_* groups must be dropped.

    patent_count stays (0 is a real integer value, not empty); the four-column
    patent_1_* group is dropped because every cell is None.
    """
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    # Do NOT add any patents → all patent_N_* cols will be None

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    # The four-column patent group must have been dropped
    assert "patent_1_number" not in df.columns, (
        "patent_1_number must be dropped when no DINs have patents"
    )
    assert "patent_1_filing_date" not in df.columns
    assert "patent_1_grant_date" not in df.columns
    assert "patent_1_expiry_date" not in df.columns
    # patent_count is 0 (integer) → not empty → must stay
    assert "patent_count" in df.columns, "patent_count must stay even when all counts are 0"


def test_patent_group_kept_when_one_din_has_patent(tmp_path):
    """When even one DIN has a patent, the patent_N_* group must be kept
    for ALL rows (trailing DINs get None in those cols, which is correct)."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    store_mod.upsert_patent("02498014", "2709025", "2008-12-10", "2014-08-26", "2028-12-10")
    # 02498022 has no patents but the group must still exist because 02498014 does

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "patent_1_number" in df.columns, "patent_1_number must be kept (02498014 has a patent)"
    row_with = df[df["din"] == "02498014"].iloc[0]
    row_without = df[df["din"] == "02498022"].iloc[0]
    assert row_with["patent_1_number"] == "2709025"
    assert row_without["patent_1_number"] is None or str(row_without["patent_1_number"]) in (
        "None", "nan", ""
    )


def test_sentinel_values_prevent_column_drop(tmp_path):
    """Columns containing only sentinel strings ('No NOC record', 'No PM available')
    must NOT be dropped — sentinels convey real information."""
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    from app.enrichment.workbook import build_sheet1

    # DPD record only → noc_* columns get "No NOC record" sentinel
    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    # All rows have "No NOC record" in noc_brand_name — must NOT be dropped
    assert "noc_brand_name" in df.columns, (
        "noc_brand_name must be kept even when all values are the 'No NOC record' sentinel"
    )
    assert "noc_submission_type" in df.columns


def test_single_nonempty_row_prevents_column_drop(tmp_path):
    """A column with even one non-empty value must not be dropped,
    even if all other rows are None."""
    import time
    import app.enrichment.store as store_mod
    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))

    # Give one DIN a labeling colour; other DIN has nothing
    store_mod.upsert_labeling("02498014", {
        "colour": "blue", "needs_ocr": 0, "has_unverified": 0,
        "drug_code": 99001, "fetched_at": time.time(),
    })
    # 02498022 has no labeling → colour=None for that row

    from app.enrichment.workbook import build_sheet1

    response = _make_response(dpd_records=[_dpd("02498014"), _dpd("02498022")])
    df = build_sheet1(response)

    assert "colour" in df.columns, (
        "colour must be kept: 02498014 has a non-empty value even though 02498022 is None"
    )


def test_export_never_produces_old_sheet_names():
    """The /api/export endpoint must never produce the old multi-sheet format."""
    import openpyxl
    from app.enrichment.workbook import build_workbook

    response = _make_response(
        dpd_records=[_dpd("02498014")],
        noc_records=[_noc("02498014")],
        gsur_records=[_gsur()],
    )
    xlsx = build_workbook(response)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    old_names = {
        "Drug Product Database (DPD)",
        "Generic Submissions Under Review",
        "Notice of Compliance (NOC)",
        "Patent Register",
        "Combined",
        "By Combination",
        "Search Metadata",
    }
    for sheet_name in wb.sheetnames:
        assert sheet_name not in old_names, (
            f"Old sheet name '{sheet_name}' must not appear in the new two-sheet workbook"
        )
    assert set(wb.sheetnames) == {"DPD + NOC + Patents", "Generic Submissions"}, (
        f"Expected exactly the two new sheet names; got: {wb.sheetnames}"
    )
