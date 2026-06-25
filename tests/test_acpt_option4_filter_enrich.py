"""Acceptance E2E for OPTION 4 — filter the universe, then enrich ONLY survivors.

Drives ``run_universe_filter_enrich_job`` with a stubbed universe + a recorder that
captures the survivor DIN work-list handed to ``enrich_labeling_batch_fast`` (the
survivor-only PM PDF stage).  Net-new coverage relative to test_universe_filters.py
(which already covers the dosage-form TABLET/CAPSULE case):

  * a NUMERIC criterion (``approvals above N``) selects exactly the right survivors;
  * a ``no_file_date`` criterion selects exactly the right survivors;
  * the Summary + Detail workbook is written and ``job.summary_records`` populated;
  * an over-constrained numeric filter (``approvals above 9999``) ⇒ ZERO PDF enrich
    calls + a valid empty Summary (len 0) + a 2-sheet workbook, job still complete;
  * a ``value`` criterion with NO IQVIA loaded ⇒ job errors with a message naming
    IQVIA (criteria 4-6 require IQVIA).

Hermetic + parallel-safe: get_universe stubbed, IQVIA store untouched, the module
universe cache reset around every test, clean enrichment store.
"""
from __future__ import annotations

import asyncio
import io
import json
from pathlib import Path

import openpyxl
import pytest
from bs4 import BeautifulSoup

import app.universe_job as UJ
from app.enrichment import universe as U
from app.enrichment.data_protection import _find_active_table, _parse_data_protection_table
from app.enrichment.universe import UniverseBundle
from app.jobs import JobState
from app.models import DrugRecord

_DP_FIX = Path(__file__).parent / "fixtures" / "data_protection"

pytestmark = pytest.mark.skipif(
    not (_DP_FIX / "register_active_sample.html").exists()
    or not (_DP_FIX / "dp_join_products.json").exists(),
    reason="dp fixtures missing — run python -m tests.scripts.build_dp_register_fixture",
)


def _dp_table() -> list[dict]:
    soup = BeautifulSoup(
        (_DP_FIX / "register_active_sample.html").read_text(encoding="utf-8"),
        "html.parser",
    )
    return _parse_data_protection_table(_find_active_table(soup))


def _join() -> dict:
    return json.loads((_DP_FIX / "dp_join_products.json").read_text(encoding="utf-8"))


@pytest.fixture(autouse=True)
def _reset_universe_cache():
    U._CACHE["bundle"] = None
    yield
    U._CACHE["bundle"] = None


@pytest.fixture()
def clean_store(tmp_path):
    import app.enrichment.store as store_mod

    store_mod.reset_for_testing(str(tmp_path / "enrich.db"))
    return store_mod


# ── deterministic option-4 bundle ─────────────────────────────────────────────
#
# Three MARKETED products that share ONE ingredient + dosage form so they form a
# single product (so "approvals" = distinct companies = 3) but with distinct
# register identities for the no-file-date filter:
#   * two register-matched innovators (carry a six-year no-file date),
#   * one generic blank (no register date).
def _dpd(din, ingredient, company, drug_code, dosage_form="TABLET", status="MARKETED"):
    return DrugRecord(
        source="DPD", din=din, brand_name="BRAND",
        ingredient=ingredient, company=company,
        all_ingredients=[ingredient.split(" ")[0]], strength="100 mg",
        dosage_form=dosage_form, status=status,
        source_specific={"drug_code": drug_code},
    )


# Two distinct products, used for the numeric-approvals selection.
#   PRODUCT A (MOLX, TABLET): 3 companies → approvals == 3
#   PRODUCT B (MOLY, CAPSULE): 1 company  → approvals == 1
_A_DINS = {"09990001", "09990002", "09990003"}
_B_DINS = {"09990009"}


def _two_product_bundle() -> UniverseBundle:
    recs = [
        _dpd("09990001", "MOLX 100 MG", "ALPHA INC", 9001),
        _dpd("09990002", "MOLX 100 MG", "BETA INC", 9002),
        _dpd("09990003", "MOLX 100 MG", "GAMMA INC", 9003),
        _dpd("09990009", "MOLY 50 MG", "DELTA INC", 9009, dosage_form="CAPSULE"),
    ]
    return UniverseBundle(recs, [], dp_table=_dp_table())


def _dpd_from_join(p, drug_code, dosage_form="TABLET"):
    return _dpd(p["din"], p["ingredient"], p["company"], drug_code, dosage_form=dosage_form)


def _date_bundle():
    """Three distinct-ingredient products: two register-matched (dated) DINs + one
    blank generic. Each join entry carries its own ingredient, so these are THREE
    products (not one) — used to exercise the no_file_date filter's inclusion rules
    across dated and blank products."""
    join = _join()
    matched = join["matched"]
    blank = join["blank"][0]
    # Pick two matched products with distinct expected dates.
    m0, m1 = matched[0], matched[1]
    recs = [
        _dpd_from_join(m0, 9101),
        _dpd_from_join(m1, 9102),
        _dpd_from_join(blank, 9103),
    ]
    expected_latest = max(m0["expected_no_file_date"], m1["expected_no_file_date"])
    matched_dins = {m0["din"], m1["din"], blank["din"]}
    return UniverseBundle(recs, [], dp_table=_dp_table()), expected_latest, matched_dins


def _run(filter_criteria, recorder, bundle):
    import unittest.mock as M

    async def _fake_get_universe(force_refresh=False):
        return bundle

    job = JobState(job_id="t-opt4", query="", field="ingredient",
                   filter_criteria=filter_criteria)
    with M.patch.object(UJ, "get_universe", _fake_get_universe), \
         M.patch.object(UJ, "enrich_labeling_batch_fast", recorder), \
         M.patch.object(UJ, "_resolve_iqvia", lambda job: None):
        asyncio.run(UJ.run_universe_filter_enrich_job(job, enable_ocr=False))
    return job


# ════════════════════════════════════════════════════════════════════════════
# NUMERIC criterion selects the right survivors
# ════════════════════════════════════════════════════════════════════════════

def test_numeric_approvals_filter_selects_and_enriches_survivors(clean_store):
    captured = {"dins": None}

    async def _recorder(din_map, **kw):
        captured["dins"] = set(din_map)

    # approvals above 2 → only PRODUCT A (3 companies) qualifies; PRODUCT B (1) drops.
    job = _run([{"metric": "approvals", "operator": "above", "value": 2}],
               _recorder, _two_product_bundle())

    assert job.status == "complete", job.error
    assert captured["dins"] == _A_DINS, captured["dins"]
    assert not (captured["dins"] & _B_DINS), "non-qualifying product DIN was enriched"

    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert wb.sheetnames == ["Summary", "Detail"]
    assert len(job.summary_records) == 1, job.summary_records


# ════════════════════════════════════════════════════════════════════════════
# no_file_date criterion selects the right survivors
# ════════════════════════════════════════════════════════════════════════════

def test_no_file_date_filter_selects_and_enriches_survivors(clean_store):
    bundle, expected_latest, matched_dins = _date_bundle()
    captured = {"dins": None}

    async def _recorder(din_map, **kw):
        captured["dins"] = set(din_map)

    # "greater than 2000-01-01" (a past->we need future MM/DD/YYYY) — use a threshold
    # safely after the product's representative date so the product qualifies via
    # "less" (date < threshold).  Far-future threshold → the dated product passes.
    job = _run([{"metric": "no_file_date", "operator": "less", "value": "01/01/2099"}],
               _recorder, bundle)

    assert job.status == "complete", job.error
    # Under "less", dated products with date < 2099 AND blank-date products are all
    # included → all three distinct-ingredient products qualify → every DIN enriched.
    assert captured["dins"] == matched_dins, captured["dins"]
    assert len(job.summary_records) == 3, job.summary_records


def test_no_file_date_greater_excludes_dated_product(clean_store):
    """A 'greater than 2099' threshold excludes the product whose date is in the
    past → empty survivor set → no enrichment."""
    bundle, _expected, _dins = _date_bundle()
    calls = {"n": 0}

    async def _recorder(din_map, **kw):
        calls["n"] += 1

    job = _run([{"metric": "no_file_date", "operator": "greater", "value": "01/01/2099"}],
               _recorder, bundle)

    assert job.status == "complete", job.error
    assert calls["n"] == 0, "no product clears a 2099 lower bound → no enrichment"
    assert len(job.summary_records) == 0


# ════════════════════════════════════════════════════════════════════════════
# Over-constrained numeric filter → empty Summary, zero PDF fetch
# ════════════════════════════════════════════════════════════════════════════

def test_overconstrained_numeric_filter_empty_and_no_pdf(clean_store):
    calls = {"n": 0}

    async def _recorder(din_map, **kw):
        calls["n"] += 1

    job = _run([{"metric": "approvals", "operator": "above", "value": 9999}],
               _recorder, _two_product_bundle())

    assert job.status == "complete", job.error
    assert calls["n"] == 0, "empty survivor set must trigger NO PDF enrichment"
    wb = openpyxl.load_workbook(io.BytesIO(Path(job.result_path).read_bytes()))
    assert wb.sheetnames == ["Summary", "Detail"]
    assert len(job.summary_records) == 0, "over-constrained filter must yield empty summary"


# ════════════════════════════════════════════════════════════════════════════
# value criterion with NO IQVIA loaded → error naming IQVIA
# ════════════════════════════════════════════════════════════════════════════

def test_value_criterion_without_iqvia_errors(clean_store):
    calls = {"n": 0}

    async def _recorder(din_map, **kw):
        calls["n"] += 1

    job = _run([{"metric": "value", "operator": "above", "value": 1000}],
               _recorder, _two_product_bundle())

    assert job.status == "error"
    assert job.error and "IQVIA" in job.error, job.error
    assert calls["n"] == 0, "no enrichment may run when the IQVIA precondition fails"
    # The error is surfaced via a terminal error SSE event too.
    errs = [e for e in job.events if e.get("status") == "error"]
    assert errs and "IQVIA" in str(errs[-1].get("message", "")), errs
